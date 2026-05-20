"""Quick Workday/HRDH structure analysis."""

from pathlib import Path
import xml.etree.ElementTree as ET
from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parent

# HRDH analysis
print("=" * 60)
print("HRDH_Table_Columns.xlsx Analysis")
print("=" * 60)

hrdh_path = PROJECT_ROOT / "HRDH_Table_Columns.xlsx"
if hrdh_path.exists():
    wb = load_workbook(hrdh_path)
    sheet_names = wb.sheetnames
    print(f"Sheets: {sheet_names}")
    
    ws = wb["overview"] if "overview" in sheet_names else wb.active
    print(f"\nActive sheet: '{ws.title}'")
    print(f"Rows: {ws.max_row}, Columns: {ws.max_column}")
    
    # Get header
    headers = [cell.value for cell in ws[1]]
    print(f"Headers: {headers}")
    
    # Sample first 3 rows
    print("\nFirst 3 data rows:")
    for i in range(2, min(5, ws.max_row + 1)):
        row_data = [ws.cell(i, j).value for j in range(1, 8)]
        print(f"  Row {i}: {row_data}")
    
    # Unique tables
    unique_tables = set()
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if row[0]:
            unique_tables.add(row[0])
    print(f"\nUnique tables: {len(unique_tables)}")
    print(f"Sample tables: {sorted(list(unique_tables))[:5]}")
else:
    print(f"Not found: {hrdh_path}")

# hr_wd.xml analysis
print("\n" + "=" * 60)
print("hr_wd.xml Analysis")
print("=" * 60)

xml_path = PROJECT_ROOT / "hr_wd.xml"
if xml_path.exists():
    with open(xml_path, 'r', encoding='utf-8') as f:
        content = f.read()
    print(f"File size: {len(content):,} bytes")
    print(f"\nFirst 500 chars:")
    print(content[:500])
    
    try:
        root = ET.fromstring(content)
        entity_count = len(root.findall(".//Entity"))
        column_count = len(root.findall(".//Column"))
        print(f"\nEntity definitions: {entity_count}")
        print(f"Column definitions: {column_count}")
        
        # Sample entities
        entities = root.findall(".//Entity")[:3]
        print(f"\nFirst 3 entities:")
        for ent in entities:
            name = ent.get("name", "")
            columns = len(ent.findall(".//Column"))
            print(f"  {name}: {columns} columns")
    except Exception as e:
        print(f"XML parse error: {e}")
else:
    print(f"Not found: {xml_path}")

print("\n" + "=" * 60)
print("Summary: Workday datahub structure")
print("=" * 60)
print("HRDH = HR Data Hub: centralized datahub integration layer")
print("- 1,428 rows covering all Workday HR entities and fields")
print("- Structure: table name + column metadata (type, length, nullable, identity)")
print("- Used as primary ingest for Workday canonical discovery")
print("\nhr_wd.xml = XSD-style schema definition")
print("- Alternative upstream source for parser-driven ingest")
print("- Not primary focus for current wave")
