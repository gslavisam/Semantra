"""Canonical glossary and knowledge runtime management for matching and stewardship."""

from __future__ import annotations

from collections import Counter
from dataclasses import dataclass, field
import csv
import hashlib
import re
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

from app.models.knowledge import (
    CanonicalGlossaryEntry,
    CanonicalGlossaryImportResponse,
    CanonicalPrivacyMetadata,
    KnowledgeConceptBaseRecord,
    KnowledgeConceptDetailResponse,
    KnowledgeConceptFieldContext,
    KnowledgeConceptPromotionResponse,
    KnowledgeConceptPromotionResult,
    KnowledgeConceptSummary,
    KnowledgeRegistryImportResponse,
)
from app.models.mapping import (
    CanonicalConceptMatchDetail,
    CanonicalCoverageColumnMatch,
    CanonicalCoverageProjectSummary,
    CanonicalCoverageSummary,
    CanonicalMappingDetails,
)
from app.models.schema import ColumnProfile, SchemaProfile
from app.services.knowledge_runtime_repository import knowledge_runtime_repository
from app.services.persistence_service import persistence_service
from app.utils.knowledge_text import filter_canonical_aliases, normalize_alias_text, normalize_canonical_alias_text, split_csv_values
from app.utils.normalization import clear_normalization_overrides, configure_normalization_overrides, semantic_token_set


PROJECT_ROOT = Path(__file__).resolve().parents[3]
DEFAULT_METADATA_DICT_PATH = PROJECT_ROOT / "metadata_dict" / "metadata_dict.csv"
DEFAULT_METADATA_WORKBOOK_PATH = PROJECT_ROOT / "metadata_dict" / "metadata_dictionary.xlsx"
DEFAULT_CANONICAL_GLOSSARY_PATH = PROJECT_ROOT / "metadata_dict" / "canonical_glossary_erp.csv"
DEFAULT_SAP_TABLES_PATH = PROJECT_ROOT / "metadata_dict" / "sap_tables_mostUsed.xlsx"
DEFAULT_QAD_TABLES_PATH = PROJECT_ROOT / "metadata_dict" / "qad_tables_mostUsed.xlsx"
DEFAULT_WORKDAY_ENTITIES_PATH = PROJECT_ROOT / "metadata_dict" / "WD_entities_mostUsed.xlsx"
DEFAULT_HRDH_TABLE_COLUMNS_PATH = PROJECT_ROOT / "metadata_dict" / "HRDH_Table_Columns.xlsx"
DEFAULT_QUICKBOOKS_TABLES_PATH = PROJECT_ROOT / "metadata_dict" / "quickbooks_tables_reference.xlsx"
DEFAULT_WD_XSD_OVERLAY_PATH = PROJECT_ROOT / "metadata_dict" / "wd_hr_knowledge_overlay.csv"
DEFAULT_HRDH_OVERLAY_PATH   = PROJECT_ROOT / "metadata_dict" / "hrdh_knowledge_overlay.csv"
DEFAULT_WD_DATAHUB_OVERLAY_PATH = PROJECT_ROOT / "metadata_dict" / "wd_datahub_knowledge_overlay.csv"
DEFAULT_QB_OVERLAY_PATH = PROJECT_ROOT / "metadata_dict" / "qb_knowledge_overlay.csv"
DEFAULT_CANONICAL_FIELD_CONTEXT_ENRICHMENT_PATH = PROJECT_ROOT / "metadata_dict" / "canonical_field_context_enrichment.csv"
DEFAULT_SAP_KNOWLEDGE_AVAILABLE_TAGS_PATH = PROJECT_ROOT / "knowledge_sources" / "generated" / "runtime" / "sap" / "sap_remaining_knowledge_available_tags.csv"
MULTI_VALUE_FIELDS = ("Skracenice", "Alternativni nazivi")
BASE_KNOWLEDGE_REQUIRED_HEADERS = ("Kategorija/Domen", "Naziv (Engleski)")
BASE_KNOWLEDGE_INLINE_EDIT_HEADERS = {
    "domain": "Kategorija/Domen",
    "serbian_name": "Naziv (Srpski)",
    "abbreviations": "Skracenice",
    "alternative_names": "Alternativni nazivi",
    "data_type": "Tip podatka",
    "typical_length": "Tipicna duzina",
    "example_value": "Primer vrednosti",
}
CANONICAL_GLOSSARY_REQUIRED_HEADERS = ("concept_id", "entity", "attribute", "display_name", "description", "data_type", "aliases")
CANONICAL_GLOSSARY_PRIVACY_HEADERS = ("is_pii", "is_gdpr_special_category", "pii_categories", "data_subject_types")
CANONICAL_GLOSSARY_HEADERS = CANONICAL_GLOSSARY_REQUIRED_HEADERS + CANONICAL_GLOSSARY_PRIVACY_HEADERS
CANONICAL_PRIVACY_CLASSIFIER_VERSION = "20260530_whole_set_v3"
DEFAULT_METADATA_EXACT_MATCH_STRENGTH = 0.75
PRIORITY_METADATA_EXACT_MATCH_STRENGTH = 0.9
DEFAULT_METADATA_TOKEN_MATCH_STRENGTH = 0.45
PRIORITY_METADATA_TOKEN_MATCH_STRENGTH = 0.55
DEFAULT_CANONICAL_METADATA_EXACT_MATCH_STRENGTH = 0.75
PRIORITY_CANONICAL_METADATA_EXACT_MATCH_STRENGTH = 0.9
DEFAULT_CANONICAL_METADATA_TOKEN_MATCH_STRENGTH = 0.5
PRIORITY_CANONICAL_METADATA_TOKEN_MATCH_STRENGTH = 0.6
DEFAULT_CANONICAL_BRIDGE_METADATA_MATCH_STRENGTH = 0.6
PRIORITY_CANONICAL_BRIDGE_METADATA_MATCH_STRENGTH = 0.7
GENERIC_ATTRIBUTE_CANONICAL_BRIDGE_STRENGTH = 0.6
ALIAS_FIELDS = (
    "Naziv (Engleski)",
    "Naziv (Srpski)",
    "snake_case",
    "camelCase",
    "PascalCase",
    "kebab-case",
    "UPPER_SNAKE",
    "lowercase",
    "SAP style",
    "Oracle/QAD style",
    "Salesforce style",
    "Odoo style",
    "MS Dynamics style",
)
PERSON_LIKE_CANONICAL_ENTITIES = {"employee", "contact", "candidate", "applicant", "person", "user"}
EMPLOYMENT_RELATED_CANONICAL_ENTITIES = {"absence", "compensation", "payroll", "benefit_enrollment"}
FINANCIAL_ACCOUNT_CANONICAL_ENTITIES = {"bank_account", "supplier_bank_account"}
ENTITY_DEFAULT_DATA_SUBJECT_TYPES = {
    "employee": "employee",
    "contact": "contact",
    "candidate": "candidate",
    "applicant": "candidate",
    "person": "person",
    "user": "user",
    "customer": "customer",
    "supplier": "supplier",
    "vendor": "supplier",
    "absence": "employee",
    "compensation": "employee",
    "payroll": "employee",
    "benefit_enrollment": "employee",
    "bank_account": "party",
    "supplier_bank_account": "supplier",
    "approval": "employee",
    "audit": "employee",
}
PERSON_NAME_ATTRIBUTES = {"name", "full_name", "first_name", "last_name", "middle_name"}
CONTACT_ATTRIBUTES = {"email", "phone", "mobile_phone", "commercial_contact_email"}
IDENTIFIER_ATTRIBUTES = {
    "id",
    "employee_id",
    "manager_id",
    "approver_id",
    "created_by",
    "changed_by",
    "owner_employee_id",
    "national_id",
    "passport_id",
}
DEMOGRAPHIC_ATTRIBUTES = {"gender", "marital_status", "nationality"}
SPECIAL_CATEGORY_TOKENS = {
    "biometric": "biometric",
    "disability": "health",
    "ethnic": "ethnicity",
    "genetic": "genetic",
    "health": "health",
    "medical": "health",
    "religion": "religion",
    "union": "union_membership",
}


def _normalize_alias(value: str) -> str:
    return normalize_alias_text(value)


def _normalize_canonical_alias(value: str) -> str:
    return normalize_canonical_alias_text(value)


def _split_values(value: str) -> list[str]:
    return split_csv_values(value)


def _parse_boolean_flag(value: object) -> bool:
    return str(value or "").strip().lower() in {"1", "true", "yes", "y", "on"}


def _normalize_privacy_tag(value: object) -> str:
    normalized = re.sub(r"\s+", "_", str(value or "").strip().lower())
    return normalized.strip("_")


def _parse_privacy_tags(value: object) -> tuple[str, ...]:
    tags = {
        normalized
        for normalized in (_normalize_privacy_tag(segment) for segment in _split_values(str(value or "")))
        if normalized
    }
    return tuple(sorted(tags))


def _merge_privacy_metadata(*records: CanonicalPrivacyMetadataRecord) -> CanonicalPrivacyMetadataRecord:
    pii_categories = {
        tag
        for record in records
        for tag in record.pii_categories
        if str(tag).strip()
    }
    data_subject_types = {
        tag
        for record in records
        for tag in record.data_subject_types
        if str(tag).strip()
    }
    is_gdpr_special_category = any(record.is_gdpr_special_category for record in records)
    is_pii = any(
        record.is_pii or record.is_gdpr_special_category or record.pii_categories or record.data_subject_types
        for record in records
    )
    return CanonicalPrivacyMetadataRecord(
        is_pii=is_pii or is_gdpr_special_category,
        is_gdpr_special_category=is_gdpr_special_category,
        pii_categories=tuple(sorted(pii_categories)),
        data_subject_types=tuple(sorted(data_subject_types)),
    )


def _canonical_privacy_metadata_model(record: CanonicalPrivacyMetadataRecord) -> CanonicalPrivacyMetadata:
    return CanonicalPrivacyMetadata(
        is_pii=record.is_pii,
        is_gdpr_special_category=record.is_gdpr_special_category,
        pii_categories=list(record.pii_categories),
        data_subject_types=list(record.data_subject_types),
    )


def _infer_canonical_privacy_metadata(
    *,
    concept_id: str,
    entity: str,
    attribute: str,
    display_name: str,
    description: str,
    aliases: Iterable[str],
) -> CanonicalPrivacyMetadataRecord:
    normalized_entity = _normalize_privacy_tag(entity)
    normalized_attribute = _normalize_privacy_tag(attribute)
    stable_text = " ".join(
        part
        for part in [concept_id, entity, attribute, display_name, description]
        if str(part).strip()
    )
    stable_tokens = semantic_token_set(stable_text)
    pii_categories: set[str] = set()
    data_subject_types: set[str] = set()
    is_pii = False
    is_gdpr_special_category = False

    default_subject = ENTITY_DEFAULT_DATA_SUBJECT_TYPES.get(normalized_entity)

    if normalized_entity in PERSON_LIKE_CANONICAL_ENTITIES:
        is_pii = True
        if default_subject:
            data_subject_types.add(default_subject)
        if normalized_entity == "employee":
            pii_categories.add("employment")

    if normalized_entity in EMPLOYMENT_RELATED_CANONICAL_ENTITIES:
        is_pii = True
        pii_categories.add("employment")
        data_subject_types.add("employee")

    if normalized_entity in FINANCIAL_ACCOUNT_CANONICAL_ENTITIES:
        is_pii = True
        pii_categories.add("financial_account")
        if default_subject:
            data_subject_types.add(default_subject)

    if normalized_attribute in PERSON_NAME_ATTRIBUTES and normalized_entity in PERSON_LIKE_CANONICAL_ENTITIES:
        is_pii = True
        pii_categories.add("person_name")
        if default_subject:
            data_subject_types.add(default_subject)

    if normalized_attribute in CONTACT_ATTRIBUTES:
        if not (normalized_entity == "supplier_bank_account" and normalized_attribute == "bank_phone"):
            is_pii = True
            pii_categories.add("contact")
            if normalized_attribute == "commercial_contact_email":
                data_subject_types.add("supplier_contact")
            elif default_subject:
                data_subject_types.add(default_subject)

    if normalized_attribute in IDENTIFIER_ATTRIBUTES:
        if normalized_entity in PERSON_LIKE_CANONICAL_ENTITIES or normalized_attribute in {"approver_id", "created_by", "changed_by", "owner_employee_id"}:
            is_pii = True
            pii_categories.update({"direct_identifier", "employment"} if default_subject == "employee" else {"direct_identifier"})
            if default_subject:
                data_subject_types.add(default_subject)

    if normalized_attribute in {"birth_date", "date_of_birth"} or ({"birth", "date"} <= stable_tokens) or "birthdate" in stable_tokens:
        is_pii = True
        pii_categories.add("date_of_birth")
        if default_subject:
            data_subject_types.add(default_subject)

    if normalized_attribute in {"national_id", "passport_id"}:
        is_pii = True
        pii_categories.update({"direct_identifier", "government_identifier"})
        if default_subject:
            data_subject_types.add(default_subject)

    if normalized_attribute in DEMOGRAPHIC_ATTRIBUTES:
        is_pii = True
        pii_categories.add("demographic")
        if default_subject:
            data_subject_types.add(default_subject)

    if normalized_attribute in {"account_number", "iban", "routing_number"}:
        is_pii = True
        pii_categories.add("financial_account")
        if default_subject:
            data_subject_types.add(default_subject)

    if normalized_entity in {"customer", "supplier", "vendor"} and normalized_attribute in {"email", "phone", "commercial_contact_email"}:
        is_pii = True
        pii_categories.add("contact")
        if default_subject:
            data_subject_types.add(default_subject)

    if normalized_entity == "vendor" and normalized_attribute == "email":
        data_subject_types.add("supplier")

    special_category_tokens = semantic_token_set(
        " ".join(
            part
            for part in [concept_id, attribute, display_name, description]
            if str(part).strip()
        )
    )
    for token, category in SPECIAL_CATEGORY_TOKENS.items():
        if token not in special_category_tokens:
            continue
        is_pii = True
        is_gdpr_special_category = True
        pii_categories.add(category)
        if default_subject:
            data_subject_types.add(default_subject)

    if not is_pii and (pii_categories or data_subject_types or is_gdpr_special_category):
        is_pii = True

    return CanonicalPrivacyMetadataRecord(
        is_pii=is_pii or is_gdpr_special_category,
        is_gdpr_special_category=is_gdpr_special_category,
        pii_categories=tuple(sorted(pii_categories)),
        data_subject_types=tuple(sorted(data_subject_types)),
    )


def _parse_context_patch_note(value: str) -> dict[str, str] | None:
    parts = [segment.strip() for segment in str(value or "").split(";") if segment.strip()]
    if not parts:
        return None

    parsed: dict[str, str] = {}
    for part in parts:
        if "=" not in part:
            continue
        key, raw_value = part.split("=", 1)
        normalized_key = key.strip().lower()
        if not normalized_key:
            continue
        parsed[normalized_key] = raw_value.strip()

    enabled = parsed.get("context_patch", "").strip().lower()
    if enabled not in {"true", "1", "yes"}:
        return None
    return parsed


@dataclass(frozen=True)
class KnowledgeFieldContext:
    """Describes one system, object, and field context attached to a knowledge concept."""

    system: str
    object_name: str
    field_name: str
    category: str = ""
    object_description: str = ""
    field_description: str = ""
    note: str = ""


@dataclass(frozen=True)
class KnowledgeConcept:
    """Knowledge-layer concept with aliases and optional field contexts used during matching."""

    concept_id: str
    domain: str
    canonical_name: str
    aliases: frozenset[str]
    contexts: tuple[KnowledgeFieldContext, ...] = ()
    context_terms: frozenset[str] = field(default_factory=frozenset)


@dataclass(frozen=True)
class ConceptMatch:
    """Scored knowledge-concept match result for a profiled schema field."""

    concept_id: str
    strength: float
    matched_aliases: tuple[str, ...] = ()
    contexts: tuple[KnowledgeFieldContext, ...] = ()


@dataclass(frozen=True)
class CanonicalBusinessConcept:
    """Canonical business concept entry loaded from the glossary and overlay runtime."""

    concept_id: str
    entity: str
    attribute: str
    display_name: str
    description: str = ""
    data_type: str = ""
    aliases: frozenset[str] = field(default_factory=frozenset)
    privacy: CanonicalPrivacyMetadataRecord = field(default_factory=lambda: CanonicalPrivacyMetadataRecord())


@dataclass(frozen=True)
class CanonicalPrivacyMetadataRecord:
    """Normalized privacy metadata attached to one canonical concept."""

    is_pii: bool = False
    is_gdpr_special_category: bool = False
    pii_categories: tuple[str, ...] = ()
    data_subject_types: tuple[str, ...] = ()


@dataclass(frozen=True)
class CanonicalConceptMatch:
    """Scored canonical-concept match result for one profiled field."""

    concept_id: str
    strength: float
    matched_aliases: tuple[str, ...] = ()


class MetadataKnowledgeService:
    """Owns the knowledge and canonical runtime used by mapping, stewardship, and reuse flows."""

    def __init__(self, csv_path: Path | None = None) -> None:
        self.csv_path = csv_path or DEFAULT_METADATA_DICT_PATH
        self.metadata_workbook_path = DEFAULT_METADATA_WORKBOOK_PATH
        self.canonical_glossary_path = DEFAULT_CANONICAL_GLOSSARY_PATH
        self.sap_tables_path = DEFAULT_SAP_TABLES_PATH
        self.qad_tables_path = DEFAULT_QAD_TABLES_PATH
        self.workday_entities_path = DEFAULT_WORKDAY_ENTITIES_PATH
        self._concepts_by_id: dict[str, KnowledgeConcept] = {}
        self._alias_to_concepts: dict[str, set[str]] = {}
        self._field_alias_to_contexts: dict[str, list[tuple[str, KnowledgeFieldContext]]] = {}
        self._canonical_concepts_by_id: dict[str, CanonicalBusinessConcept] = {}
        self._canonical_alias_to_concepts: dict[str, set[str]] = {}
        self._active_overlay_name: str | None = None
        self._overlay_aliases_by_concept: dict[str, set[str]] = {}
        self._overlay_canonical_aliases_by_concept: dict[str, set[str]] = {}
        self._concept_match_cache: dict[tuple[bool, tuple[str, str, str, str, tuple[str, ...]]], list[ConceptMatch]] = {}
        self._canonical_concept_match_cache: dict[
            tuple[bool, tuple[str, str, str, str, tuple[str, ...]]],
            list[CanonicalConceptMatch],
        ] = {}
        self._last_runtime_source = "source_files"
        # Bridge: KC concept_id → set of CC concept_ids reachable via shared aliases.
        # Built after every full load so match_canonical_concepts() can use KC signal.
        self._kc_to_cc: dict[str, set[str]] = {}
        self.refresh()

    @property
    def is_available(self) -> bool:
        return bool(self._concepts_by_id)

    @property
    def concept_count(self) -> int:
        return len(self._concepts_by_id)

    @property
    def canonical_concept_count(self) -> int:
        return len(self._canonical_concepts_by_id)

    @property
    def runtime_source(self) -> str:
        return self._last_runtime_source

    def current_source_hash(self) -> str:
        return self._compute_source_hash()

    def concepts_for_alias(self, alias: str) -> list[str]:
        """Return knowledge concept ids currently linked to a normalized alias."""

        normalized_alias = _normalize_alias(alias)
        if not normalized_alias:
            return []
        return sorted(self._alias_to_concepts.get(normalized_alias, set()))

    def resolve_canonical_concept_id(self, term: str) -> str | None:
        """Resolve free text, display name, or alias text to a canonical concept id when possible."""

        normalized_term = _normalize_alias(term)
        if not normalized_term:
            return None

        exact_term = term.strip()
        if exact_term in self._canonical_concepts_by_id:
            return exact_term

        for concept_id, concept in self._canonical_concepts_by_id.items():
            if normalized_term == _normalize_alias(concept_id.replace(".", " ")):
                return concept_id
            if normalized_term == _normalize_alias(concept.display_name):
                return concept_id
            if normalized_term in concept.aliases:
                return concept_id
        return None

    def export_canonical_glossary_csv(self) -> str:
        if not self.canonical_glossary_path.exists():
            return ",".join(CANONICAL_GLOSSARY_HEADERS) + "\n"

        payload = self.canonical_glossary_path.read_text(encoding="utf-8-sig")
        return payload if payload.endswith("\n") else payload + "\n"

    def export_base_knowledge_csv(self) -> str:
        if not self.csv_path.exists():
            return ""

        payload = self.csv_path.read_text(encoding="utf-8-sig")
        return payload if payload.endswith("\n") else payload + "\n"

    def canonical_privacy_concepts_for_mapping(
        self,
        source: ColumnProfile,
        target: ColumnProfile,
        *,
        prefer_metadata_text: bool = False,
    ) -> list[dict[str, object]]:
        """Return shared canonical concepts that carry privacy metadata for one mapping pair."""

        details = self.canonical_mapping_details(source, target, prefer_metadata_text=prefer_metadata_text)
        privacy_concepts: list[dict[str, object]] = []
        for match in details.shared_concepts:
            concept = self._canonical_concepts_by_id.get(match.concept_id)
            if concept is None:
                continue
            privacy = concept.privacy
            if not (
                privacy.is_pii
                or privacy.is_gdpr_special_category
                or privacy.pii_categories
                or privacy.data_subject_types
            ):
                continue
            privacy_concepts.append(
                {
                    "concept_id": concept.concept_id,
                    "display_name": concept.display_name,
                    "strength": match.strength,
                    "is_pii": privacy.is_pii,
                    "is_gdpr_special_category": privacy.is_gdpr_special_category,
                    "pii_categories": list(privacy.pii_categories),
                    "data_subject_types": list(privacy.data_subject_types),
                }
            )
        return privacy_concepts

    def _linked_canonical_privacy(self, concept_ids: Iterable[str]) -> CanonicalPrivacyMetadata:
        records = [
            self._canonical_concepts_by_id[concept_id].privacy
            for concept_id in concept_ids
            if concept_id in self._canonical_concepts_by_id
        ]
        if not records:
            return CanonicalPrivacyMetadata()
        return _canonical_privacy_metadata_model(_merge_privacy_metadata(*records))

    def list_knowledge_concepts(self) -> list[KnowledgeConceptSummary]:
        base_concept_ids = self._base_knowledge_concept_ids()
        concepts: list[KnowledgeConceptSummary] = []
        for concept in sorted(self._concepts_by_id.values(), key=lambda item: item.concept_id):
            source_systems = sorted({context.system.strip() for context in concept.contexts if context.system.strip()})
            linked_canonical_concepts = sorted(self._kc_to_cc.get(concept.concept_id, set()))
            editable = concept.concept_id in base_concept_ids
            concepts.append(
                KnowledgeConceptSummary(
                    concept_id=concept.concept_id,
                    domain=concept.domain,
                    canonical_name=concept.canonical_name,
                    source=self._classify_knowledge_concept_source(concept.concept_id, editable=editable),
                    editable=editable,
                    alias_count=len(concept.aliases),
                    field_context_count=len(concept.contexts),
                    linked_canonical_concept_count=len(linked_canonical_concepts),
                    source_systems=source_systems,
                    linked_canonical_concepts=linked_canonical_concepts,
                    linked_privacy=self._linked_canonical_privacy(linked_canonical_concepts),
                    aliases=sorted(concept.aliases),
                )
            )
        return concepts

    def get_knowledge_concept(self, concept_id: str) -> KnowledgeConceptDetailResponse | None:
        normalized_concept_id = str(concept_id or "").strip()
        concept = self._concepts_by_id.get(normalized_concept_id)
        if concept is None:
            return None

        base_concept_ids = self._base_knowledge_concept_ids()
        editable = concept.concept_id in base_concept_ids
        linked_canonical_concepts = sorted(self._kc_to_cc.get(concept.concept_id, set()))
        summary = KnowledgeConceptSummary(
            concept_id=concept.concept_id,
            domain=concept.domain,
            canonical_name=concept.canonical_name,
            source=self._classify_knowledge_concept_source(concept.concept_id, editable=editable),
            editable=editable,
            alias_count=len(concept.aliases),
            field_context_count=len(concept.contexts),
            linked_canonical_concept_count=len(linked_canonical_concepts),
            source_systems=sorted({context.system.strip() for context in concept.contexts if context.system.strip()}),
            linked_canonical_concepts=linked_canonical_concepts,
            linked_privacy=self._linked_canonical_privacy(linked_canonical_concepts),
            aliases=sorted(concept.aliases),
        )
        return KnowledgeConceptDetailResponse(
            concept=summary,
            field_contexts=[
                KnowledgeConceptFieldContext(
                    system=context.system,
                    object_name=context.object_name,
                    field_name=context.field_name,
                    category=context.category,
                    object_description=context.object_description,
                    field_description=context.field_description,
                    note=context.note,
                )
                for context in concept.contexts
            ],
            base_record=self._build_base_knowledge_record(normalized_concept_id),
        )

    def update_base_knowledge_concept(
        self,
        concept_id: str,
        *,
        domain: str,
        serbian_name: str,
        abbreviations: str,
        alternative_names: str,
        data_type: str,
        typical_length: str,
        example_value: str,
    ) -> KnowledgeConceptDetailResponse:
        fieldnames, rows = self._read_base_knowledge_rows()
        row = self._base_knowledge_row(rows, concept_id)
        if row is None:
            raise KeyError(f"Unknown editable base knowledge concept: {concept_id}")

        row[BASE_KNOWLEDGE_INLINE_EDIT_HEADERS["domain"]] = str(domain or "").strip()
        row[BASE_KNOWLEDGE_INLINE_EDIT_HEADERS["serbian_name"]] = str(serbian_name or "").strip()
        row[BASE_KNOWLEDGE_INLINE_EDIT_HEADERS["abbreviations"]] = str(abbreviations or "").strip()
        row[BASE_KNOWLEDGE_INLINE_EDIT_HEADERS["alternative_names"]] = str(alternative_names or "").strip()
        row[BASE_KNOWLEDGE_INLINE_EDIT_HEADERS["data_type"]] = str(data_type or "").strip()
        row[BASE_KNOWLEDGE_INLINE_EDIT_HEADERS["typical_length"]] = str(typical_length or "").strip()
        row[BASE_KNOWLEDGE_INLINE_EDIT_HEADERS["example_value"]] = str(example_value or "").strip()

        self._write_base_knowledge_rows(fieldnames, rows)
        self.refresh()
        detail = self.get_knowledge_concept(concept_id)
        if detail is None:
            raise KeyError(f"Knowledge concept disappeared after update: {concept_id}")
        return detail

    def promote_knowledge_concepts_to_canonical_glossary(
        self,
        concept_ids: Iterable[str],
        *,
        target_concept_id: str | None = None,
    ) -> KnowledgeConceptPromotionResponse:
        if not self.canonical_glossary_path.exists():
            raise KeyError("Canonical glossary file is not available.")

        with self.canonical_glossary_path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            if not reader.fieldnames:
                raise ValueError("Canonical glossary CSV must include a header row.")
            rows = [{header: str(row.get(header) or "").strip() for header in CANONICAL_GLOSSARY_HEADERS} for row in reader]

        rows_by_concept_id = {row["concept_id"]: row for row in rows if row.get("concept_id")}
        results: list[KnowledgeConceptPromotionResult] = []
        changed = False

        for raw_concept_id in concept_ids:
            normalized_concept_id = str(raw_concept_id or "").strip()
            concept = self._concepts_by_id.get(normalized_concept_id)
            if concept is None:
                results.append(
                    KnowledgeConceptPromotionResult(
                        knowledge_concept_id=normalized_concept_id or "unknown",
                        status="error",
                        message="Knowledge concept was not found in the current runtime.",
                    )
                )
                continue

            linked_targets = sorted(self._kc_to_cc.get(normalized_concept_id, set()))
            resolved_target_concept_id = str(target_concept_id or "").strip() or None
            if resolved_target_concept_id is None:
                if len(linked_targets) == 1:
                    resolved_target_concept_id = linked_targets[0]
                elif not linked_targets:
                    results.append(
                        KnowledgeConceptPromotionResult(
                            knowledge_concept_id=normalized_concept_id,
                            status="skipped",
                            message="No linked canonical concept is available for automatic promotion.",
                        )
                    )
                    continue
                else:
                    results.append(
                        KnowledgeConceptPromotionResult(
                            knowledge_concept_id=normalized_concept_id,
                            status="skipped",
                            message="Multiple linked canonical concepts are available; choose one explicitly for single-concept promotion.",
                        )
                    )
                    continue

            target_runtime_concept = self._canonical_concepts_by_id.get(resolved_target_concept_id)
            target_row = rows_by_concept_id.get(resolved_target_concept_id)
            concept_created = False
            if target_row is None:
                if target_runtime_concept is None:
                    results.append(
                        KnowledgeConceptPromotionResult(
                            knowledge_concept_id=normalized_concept_id,
                            target_concept_id=resolved_target_concept_id,
                            status="error",
                            message="Target canonical concept is not available in the stable glossary or runtime.",
                        )
                    )
                    continue
                target_row = {
                    "concept_id": target_runtime_concept.concept_id,
                    "entity": target_runtime_concept.entity,
                    "attribute": target_runtime_concept.attribute,
                    "display_name": target_runtime_concept.display_name,
                    "description": target_runtime_concept.description,
                    "data_type": target_runtime_concept.data_type,
                    "aliases": "",
                }
                rows.append(target_row)
                rows_by_concept_id[resolved_target_concept_id] = target_row
                concept_created = True
                changed = True

            current_aliases = _split_values(target_row.get("aliases") or "")
            existing_aliases = {
                normalized
                for normalized in (_normalize_canonical_alias(value) for value in current_aliases)
                if normalized
            }
            candidate_aliases = filter_canonical_aliases([concept.canonical_name, *sorted(concept.aliases)])
            aliases_to_add = [
                alias
                for alias in candidate_aliases
                if _normalize_canonical_alias(alias) and _normalize_canonical_alias(alias) not in existing_aliases
            ]

            if aliases_to_add:
                target_row["aliases"] = ", ".join(filter_canonical_aliases([*current_aliases, *aliases_to_add]))
                changed = True
                message = f"Promoted {len(aliases_to_add)} aliases into {resolved_target_concept_id}."
                status = "promoted"
            elif concept_created:
                message = f"Created stable canonical concept {resolved_target_concept_id}; all promoted aliases were already present."
                status = "promoted"
            else:
                message = f"All candidate aliases are already present in {resolved_target_concept_id}."
                status = "skipped"

            results.append(
                KnowledgeConceptPromotionResult(
                    knowledge_concept_id=normalized_concept_id,
                    target_concept_id=resolved_target_concept_id,
                    status=status,
                    alias_count=len(candidate_aliases),
                    aliases_added=len(aliases_to_add),
                    concept_created=concept_created,
                    message=message,
                )
            )

        if changed:
            with self.canonical_glossary_path.open("w", encoding="utf-8", newline="") as handle:
                writer = csv.DictWriter(handle, fieldnames=CANONICAL_GLOSSARY_HEADERS)
                writer.writeheader()
                writer.writerows(rows)
            self._refresh_after_canonical_authoring_change()

        return KnowledgeConceptPromotionResponse(
            promoted_count=sum(1 for item in results if item.status == "promoted"),
            skipped_count=sum(1 for item in results if item.status != "promoted"),
            results=results,
        )

    def list_canonical_glossary_entries(self) -> list[CanonicalGlossaryEntry]:
        return [
            CanonicalGlossaryEntry(
                concept_id=concept.concept_id,
                entity=concept.entity,
                attribute=concept.attribute,
                display_name=concept.display_name,
                description=concept.description,
                data_type=concept.data_type,
                aliases=sorted(concept.aliases),
                privacy=_canonical_privacy_metadata_model(concept.privacy),
            )
            for concept in sorted(self._canonical_concepts_by_id.values(), key=lambda item: item.concept_id)
        ]

    def import_canonical_glossary_csv(
        self,
        payload: bytes,
        filename: str | None = None,
    ) -> CanonicalGlossaryImportResponse:
        if filename and not filename.lower().endswith(".csv"):
            raise ValueError("Canonical glossary import currently supports CSV files only.")

        decoded = self._decode_csv_payload(payload)
        reader = csv.DictReader(decoded.splitlines())
        if not reader.fieldnames:
            raise ValueError("Canonical glossary CSV must include a header row.")

        missing_headers = [header for header in CANONICAL_GLOSSARY_REQUIRED_HEADERS if header not in reader.fieldnames]
        if missing_headers:
            missing_label = ", ".join(missing_headers)
            raise ValueError(f"Canonical glossary CSV is missing required columns: {missing_label}.")

        parsed_rows: list[dict[str, str]] = []
        for row_number, row in enumerate(reader, start=2):
            concept_id = str(row.get("concept_id") or "").strip()
            display_name = str(row.get("display_name") or "").strip()
            if not concept_id:
                raise ValueError(f"Canonical glossary row {row_number} is missing concept_id.")
            if not display_name:
                raise ValueError(f"Canonical glossary row {row_number} is missing display_name.")
            parsed_row = {header: str(row.get(header) or "").strip() for header in CANONICAL_GLOSSARY_HEADERS}
            parsed_row["aliases"] = ", ".join(filter_canonical_aliases(_split_values(parsed_row.get("aliases") or "")))
            parsed_row["is_pii"] = "true" if _parse_boolean_flag(parsed_row.get("is_pii")) else ""
            parsed_row["is_gdpr_special_category"] = "true" if _parse_boolean_flag(parsed_row.get("is_gdpr_special_category")) else ""
            parsed_row["pii_categories"] = ", ".join(_parse_privacy_tags(parsed_row.get("pii_categories")))
            parsed_row["data_subject_types"] = ", ".join(_parse_privacy_tags(parsed_row.get("data_subject_types")))
            parsed_rows.append(parsed_row)

        with self.canonical_glossary_path.open("w", encoding="utf-8", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=CANONICAL_GLOSSARY_HEADERS)
            writer.writeheader()
            writer.writerows(parsed_rows)
        self._refresh_after_canonical_authoring_change()
        return CanonicalGlossaryImportResponse(
            imported_row_count=len(parsed_rows),
            canonical_concept_count=self.canonical_concept_count,
            source_filename=filename,
        )

    def import_base_knowledge_csv(
        self,
        payload: bytes,
        filename: str | None = None,
    ) -> KnowledgeRegistryImportResponse:
        if filename and not filename.lower().endswith(".csv"):
            raise ValueError("Knowledge registry import currently supports CSV files only.")

        decoded = self._decode_csv_payload(payload)
        reader = csv.DictReader(decoded.splitlines())
        if not reader.fieldnames:
            raise ValueError("Knowledge registry CSV must include a header row.")

        fieldnames = [str(header or "").strip() for header in reader.fieldnames if str(header or "").strip()]
        missing_headers = [header for header in BASE_KNOWLEDGE_REQUIRED_HEADERS if header not in fieldnames]
        if missing_headers:
            missing_label = ", ".join(missing_headers)
            raise ValueError(f"Knowledge registry CSV is missing required columns: {missing_label}.")

        parsed_rows: list[dict[str, str]] = []
        for row_number, row in enumerate(reader, start=2):
            concept_id = str(row.get("Naziv (Engleski)") or "").strip()
            if not concept_id:
                raise ValueError(f"Knowledge registry row {row_number} is missing Naziv (Engleski).")
            parsed_rows.append({header: str(row.get(header) or "").strip() for header in fieldnames})

        with self.csv_path.open("w", encoding="utf-8", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(parsed_rows)
        self.refresh()
        return KnowledgeRegistryImportResponse(
            imported_row_count=len(parsed_rows),
            knowledge_concept_count=self.concept_count,
            source_filename=filename,
        )

    def promote_overlay_alias_to_canonical_glossary(
        self,
        concept_id: str,
        alias: str,
        *,
        display_name: str | None = None,
        description: str | None = None,
        data_type: str | None = None,
    ) -> tuple[CanonicalGlossaryEntry, bool, bool]:
        normalized_concept_id = str(concept_id or "").strip()
        normalized_alias = _normalize_canonical_alias(alias)
        if not normalized_concept_id:
            raise ValueError("Canonical glossary promotion requires concept_id.")
        if not normalized_alias:
            raise ValueError("Canonical glossary promotion requires a non-empty, non-numeric alias.")

        if not self.canonical_glossary_path.exists():
            raise KeyError("Canonical glossary file is not available.")

        with self.canonical_glossary_path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            if not reader.fieldnames:
                raise ValueError("Canonical glossary CSV must include a header row.")
            rows = [{header: str(row.get(header) or "").strip() for header in CANONICAL_GLOSSARY_HEADERS} for row in reader]

        matching_row = next((row for row in rows if row.get("concept_id") == normalized_concept_id), None)
        concept_created = matching_row is None
        if matching_row is None:
            matching_row = {
                "concept_id": normalized_concept_id,
                "entity": normalized_concept_id.split(".", 1)[0] if "." in normalized_concept_id else "general",
                "attribute": normalized_concept_id.split(".", 1)[1] if "." in normalized_concept_id else normalized_concept_id,
                "display_name": str(display_name or normalized_concept_id).strip() or normalized_concept_id,
                "description": str(description or "").strip(),
                "data_type": str(data_type or "").strip(),
                "aliases": "",
                "is_pii": "",
                "is_gdpr_special_category": "",
                "pii_categories": "",
                "data_subject_types": "",
            }
            rows.append(matching_row)

        existing_aliases = {
            normalized
            for normalized in (_normalize_canonical_alias(value) for value in _split_values(matching_row.get("aliases") or ""))
            if normalized
        }
        alias_added = normalized_alias not in existing_aliases
        existing_aliases.add(normalized_alias)
        matching_row["aliases"] = ", ".join(sorted(existing_aliases))

        with self.canonical_glossary_path.open("w", encoding="utf-8", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=CANONICAL_GLOSSARY_HEADERS)
            writer.writeheader()
            writer.writerows(rows)

        self._refresh_after_canonical_authoring_change()

        concept = self._canonical_concepts_by_id.get(normalized_concept_id)
        if concept is None:
            raise KeyError(f"Canonical concept '{normalized_concept_id}' was not available after glossary refresh.")
        return (
            CanonicalGlossaryEntry(
                concept_id=concept.concept_id,
                entity=concept.entity,
                attribute=concept.attribute,
                display_name=concept.display_name,
                description=concept.description,
                data_type=concept.data_type,
                aliases=sorted(concept.aliases),
                privacy=_canonical_privacy_metadata_model(concept.privacy),
            ),
            alias_added,
            concept_created,
        )

    def refresh(self) -> None:
        self._clear_runtime_state()

        # --- DB-first loading (skip expensive file parsing when DB is current) ---
        current_hash = self._compute_source_hash()
        seed_meta = knowledge_runtime_repository.get_seed_meta()
        if seed_meta and seed_meta["source_hash"] == current_hash:
            self._load_from_db()
            self._last_runtime_source = "sqlite_cache"
        else:
            self._load_canonical_glossary()
            self._load()
            self._seed_to_db(current_hash)
            self._last_runtime_source = "source_files"

        self._apply_active_overlay()
        self._rebuild_kc_cc_bridge()

    def _profile_metadata_candidates(self, profile: ColumnProfile) -> tuple[set[str], set[str]]:
        metadata_candidates: set[str] = set()
        metadata_tokens: set[str] = set()
        for value in (profile.description, profile.declared_type):
            text = str(value or "").strip()
            normalized = _normalize_alias(text)
            if not normalized:
                continue
            metadata_candidates.add(normalized)
            metadata_tokens.update(token for token in normalized.split() if token)
            metadata_tokens.update(semantic_token_set(text))
        return metadata_candidates, metadata_tokens

    def _profile_match_cache_key(
        self,
        profile: ColumnProfile,
        *,
        prefer_metadata_text: bool,
    ) -> tuple[bool, tuple[str, str, str, str, tuple[str, ...]]]:
        return (
            bool(prefer_metadata_text),
            (
                str(profile.name),
                str(profile.normalized_name),
                str(profile.description),
                str(profile.declared_type),
                tuple(str(token) for token in profile.tokenized_name),
            ),
        )

    def _metadata_match_strengths(self, *, prefer_metadata_text: bool = False) -> tuple[float, float, float, float, float]:
        if prefer_metadata_text:
            return (
                PRIORITY_METADATA_EXACT_MATCH_STRENGTH,
                PRIORITY_METADATA_TOKEN_MATCH_STRENGTH,
                PRIORITY_CANONICAL_METADATA_EXACT_MATCH_STRENGTH,
                PRIORITY_CANONICAL_METADATA_TOKEN_MATCH_STRENGTH,
                PRIORITY_CANONICAL_BRIDGE_METADATA_MATCH_STRENGTH,
            )
        return (
            DEFAULT_METADATA_EXACT_MATCH_STRENGTH,
            DEFAULT_METADATA_TOKEN_MATCH_STRENGTH,
            DEFAULT_CANONICAL_METADATA_EXACT_MATCH_STRENGTH,
            DEFAULT_CANONICAL_METADATA_TOKEN_MATCH_STRENGTH,
            DEFAULT_CANONICAL_BRIDGE_METADATA_MATCH_STRENGTH,
        )

    def _profile_matches_canonical_attribute(self, profile: ColumnProfile, concept: CanonicalBusinessConcept) -> bool:
        attribute_alias = _normalize_canonical_alias(concept.attribute)
        if not attribute_alias or len(attribute_alias.split()) < 2:
            return False
        profile_names = {
            _normalize_canonical_alias(profile.name),
            _normalize_canonical_alias(profile.normalized_name),
        }
        return attribute_alias in profile_names

    def _pairwise_canonical_attribute_bridge(
        self,
        source: ColumnProfile,
        target: ColumnProfile,
        *,
        prefer_metadata_text: bool = False,
    ) -> dict[str, float]:
        """Return weak canonical bridge scores when a generic target name matches a source concept attribute.

        This keeps generic targets like `region_code` or `payment_terms_id` from losing all
        business-meaning evidence when the source already resolves strongly to a richer canonical concept.
        """

        bridges: dict[str, float] = {}
        for match in self.match_canonical_concepts(source, prefer_metadata_text=prefer_metadata_text):
            concept = self._canonical_concepts_by_id.get(match.concept_id)
            if concept is None:
                continue
            if not self._profile_matches_canonical_attribute(target, concept):
                continue
            bridges[match.concept_id] = max(
                bridges.get(match.concept_id, 0.0),
                min(match.strength, GENERIC_ATTRIBUTE_CANONICAL_BRIDGE_STRENGTH),
            )
        return bridges

    def expand_semantic_tokens(self, profile: ColumnProfile, *, prefer_metadata_text: bool = False) -> set[str]:
        tokens = set(profile.tokenized_name)
        if prefer_metadata_text:
            _metadata_candidates, metadata_tokens = self._profile_metadata_candidates(profile)
            tokens.update(metadata_tokens)
        for match in self.match_concepts(profile, prefer_metadata_text=prefer_metadata_text):
            concept = self._concepts_by_id.get(match.concept_id)
            if not concept:
                continue
            tokens.add(match.concept_id)
            for alias in concept.aliases:
                tokens.update(alias.split())
            tokens.update(concept.context_terms)
        return tokens

    def knowledge_alignment(self, source: ColumnProfile, target: ColumnProfile, *, prefer_metadata_text: bool = False) -> float:
        """Score how strongly the source field's knowledge concepts align with the target field."""

        source_matches = {match.concept_id: match for match in self.match_concepts(source, prefer_metadata_text=prefer_metadata_text)}
        target_matches = {match.concept_id: match for match in self.match_concepts(target, prefer_metadata_text=prefer_metadata_text)}
        source_canonical_matches = {
            match.concept_id: match for match in self.match_canonical_concepts(source, prefer_metadata_text=prefer_metadata_text)
        }
        target_canonical_matches = {
            match.concept_id: match for match in self.match_canonical_concepts(target, prefer_metadata_text=prefer_metadata_text)
        }
        shared = set(source_matches) & set(target_matches)

        candidate_scores: list[float] = []
        for concept_id in shared:
            source_match = source_matches[concept_id]
            target_match = target_matches[concept_id]
            base_score = min(source_match.strength, target_match.strength)
            if source_match.contexts and target_match.contexts:
                base_score = min(1.0, base_score + 0.15)
            elif source_match.contexts or target_match.contexts:
                base_score = min(1.0, base_score + 0.05)
            candidate_scores.append(base_score)

        # Bridge extension: when target is a canonical concept, check if any
        # source KC concept bridges (via _kc_to_cc) directly to that CC target.
        # This allows SAP/QAD fields with rich KC aliases to produce a strong
        # knowledge signal even when the CC target has no KC representation.
        target_cc_id = self.resolve_canonical_concept_id(target.name)
        if target_cc_id == target.name:
            for kc_id, source_match in source_matches.items():
                if target_cc_id in self._kc_to_cc.get(kc_id, set()):
                    candidate_scores.append(source_match.strength * 0.85)

        # Metadata-driven canonical matches should also feed the KC->CC bridge,
        # otherwise the same concept can show up in canonical_details while the
        # knowledge score incorrectly stays at zero.
        for kc_id, source_match in source_matches.items():
            for target_cc_id, target_cc_match in target_canonical_matches.items():
                if not set(source_match.matched_aliases).intersection(target_cc_match.matched_aliases):
                    continue
                if target_cc_id in self._kc_to_cc.get(kc_id, set()):
                    candidate_scores.append(min(source_match.strength * 0.85, target_cc_match.strength))

        for kc_id, target_match in target_matches.items():
            for source_cc_id, source_cc_match in source_canonical_matches.items():
                if not set(target_match.matched_aliases).intersection(source_cc_match.matched_aliases):
                    continue
                if source_cc_id in self._kc_to_cc.get(kc_id, set()):
                    candidate_scores.append(min(target_match.strength * 0.85, source_cc_match.strength))

        for concept_id, bridge_strength in self._pairwise_canonical_attribute_bridge(
            source,
            target,
            prefer_metadata_text=prefer_metadata_text,
        ).items():
            linked_source_strength = max(
                (
                    source_match.strength
                    for kc_id, source_match in source_matches.items()
                    if concept_id in self._kc_to_cc.get(kc_id, set())
                ),
                default=0.0,
            )
            if linked_source_strength > 0:
                candidate_scores.append(min(linked_source_strength * 0.85, bridge_strength))

        for concept_id, bridge_strength in self._pairwise_canonical_attribute_bridge(
            target,
            source,
            prefer_metadata_text=prefer_metadata_text,
        ).items():
            linked_target_strength = max(
                (
                    target_match.strength
                    for kc_id, target_match in target_matches.items()
                    if concept_id in self._kc_to_cc.get(kc_id, set())
                ),
                default=0.0,
            )
            if linked_target_strength > 0:
                candidate_scores.append(min(linked_target_strength * 0.85, bridge_strength))

        return round(max(candidate_scores), 4) if candidate_scores else 0.0

    def canonical_alignment(self, source: ColumnProfile, target: ColumnProfile, *, prefer_metadata_text: bool = False) -> float:
        """Score how strongly the source and target align on canonical business concepts."""

        source_matches = {match.concept_id: match for match in self.match_canonical_concepts(source, prefer_metadata_text=prefer_metadata_text)}
        target_matches = {match.concept_id: match for match in self.match_canonical_concepts(target, prefer_metadata_text=prefer_metadata_text)}
        shared = set(source_matches) & set(target_matches)
        candidate_scores = [min(source_matches[concept_id].strength, target_matches[concept_id].strength) for concept_id in shared]
        candidate_scores.extend(
            self._pairwise_canonical_attribute_bridge(
                source,
                target,
                prefer_metadata_text=prefer_metadata_text,
            ).values()
        )
        candidate_scores.extend(
            self._pairwise_canonical_attribute_bridge(
                target,
                source,
                prefer_metadata_text=prefer_metadata_text,
            ).values()
        )
        if not candidate_scores:
            return 0.0

        return round(max(candidate_scores), 4)

    def canonical_mapping_details(
        self,
        source: ColumnProfile,
        target: ColumnProfile,
        *,
        prefer_metadata_text: bool = False,
    ) -> CanonicalMappingDetails:
        """Build the canonical detail payload attached to mapping candidates and explanations."""

        source_matches = {match.concept_id: match for match in self.match_canonical_concepts(source, prefer_metadata_text=prefer_metadata_text)}
        target_matches = {match.concept_id: match for match in self.match_canonical_concepts(target, prefer_metadata_text=prefer_metadata_text)}
        shared = set(source_matches) & set(target_matches)

        return CanonicalMappingDetails(
            source_concepts=self._canonical_match_details(source_matches),
            target_concepts=self._canonical_match_details(target_matches),
            shared_concepts=self._canonical_match_details(
                {
                    concept_id: CanonicalConceptMatch(
                        concept_id=concept_id,
                        strength=min(source_matches[concept_id].strength, target_matches[concept_id].strength),
                    )
                    for concept_id in shared
                }
            ),
        )

    def canonical_coverage(self, schema: SchemaProfile, *, prefer_metadata_text: bool = False) -> CanonicalCoverageSummary:
        matched_columns_detail: list[CanonicalCoverageColumnMatch] = []
        unmatched_columns: list[str] = []

        for column in schema.columns:
            matches = self.match_canonical_concepts(column, prefer_metadata_text=prefer_metadata_text)
            if matches:
                matched_columns_detail.append(
                    CanonicalCoverageColumnMatch(
                        column=column.name,
                        concept_ids=[match.concept_id for match in matches],
                    )
                )
            else:
                unmatched_columns.append(column.name)

        total_columns = len(schema.columns)
        matched_columns = len(matched_columns_detail)
        coverage_ratio = round((matched_columns / total_columns), 4) if total_columns else 0.0
        return CanonicalCoverageSummary(
            total_columns=total_columns,
            matched_columns=matched_columns,
            coverage_ratio=coverage_ratio,
            unmatched_columns=unmatched_columns,
            matched_columns_detail=matched_columns_detail,
        )

    def canonical_project_coverage(
        self,
        source_coverage: CanonicalCoverageSummary,
        target_coverage: CanonicalCoverageSummary,
    ) -> CanonicalCoverageProjectSummary:
        source_concepts = {
            concept_id
            for detail in source_coverage.matched_columns_detail
            for concept_id in detail.concept_ids
        }
        target_concepts = {
            concept_id
            for detail in target_coverage.matched_columns_detail
            for concept_id in detail.concept_ids
        }
        all_concepts = sorted(source_concepts | target_concepts)
        shared_concepts = sorted(source_concepts & target_concepts)
        source_only_concepts = sorted(source_concepts - target_concepts)
        target_only_concepts = sorted(target_concepts - source_concepts)
        total_columns = source_coverage.total_columns + target_coverage.total_columns
        matched_columns = source_coverage.matched_columns + target_coverage.matched_columns

        return CanonicalCoverageProjectSummary(
            total_columns=total_columns,
            matched_columns=matched_columns,
            coverage_ratio=round((matched_columns / total_columns), 4) if total_columns else 0.0,
            concept_count=len(all_concepts),
            shared_concept_count=len(shared_concepts),
            concepts=all_concepts,
            shared_concepts=shared_concepts,
            source_only_concepts=source_only_concepts,
            target_only_concepts=target_only_concepts,
        )

    def explain_alignment(
        self,
        source: ColumnProfile,
        target: ColumnProfile,
        *,
        prefer_metadata_text: bool = False,
    ) -> list[str]:
        source_matches = {match.concept_id: match for match in self.match_concepts(source, prefer_metadata_text=prefer_metadata_text)}
        target_matches = {match.concept_id: match for match in self.match_concepts(target, prefer_metadata_text=prefer_metadata_text)}
        shared = set(source_matches) & set(target_matches)
        explanations: list[str] = []
        for concept_id in sorted(shared):
            concept = self._concepts_by_id.get(concept_id)
            if concept is None:
                continue
            explanations.append(
                f"Internal metadata dictionary aligns both fields to concept '{concept.canonical_name}' in domain '{concept.domain}'."
            )
            overlay_aliases = self._overlay_aliases_by_concept.get(concept_id, set())
            matched_overlay_aliases = sorted(
                overlay_aliases.intersection(source_matches[concept_id].matched_aliases)
                | overlay_aliases.intersection(target_matches[concept_id].matched_aliases)
            )
            if matched_overlay_aliases and self._active_overlay_name:
                explanations.append(
                    f"Custom knowledge overlay '{self._active_overlay_name}' matched alias(es): {', '.join(matched_overlay_aliases)}."
                )
            source_contexts = source_matches[concept_id].contexts
            target_contexts = target_matches[concept_id].contexts
            if source_contexts or target_contexts:
                source_label = ", ".join(self._format_context(context) for context in source_contexts[:2]) or "no explicit source context"
                target_label = ", ".join(self._format_context(context) for context in target_contexts[:2]) or "no explicit target context"
                explanations.append(f"Context prior: source {source_label}; target {target_label}.")
        if not shared:
            for concept_id in sorted(
                self._pairwise_canonical_attribute_bridge(
                    source,
                    target,
                    prefer_metadata_text=prefer_metadata_text,
                )
            ):
                concept = self._canonical_concepts_by_id.get(concept_id)
                if concept is None:
                    continue
                explanations.append(
                    f"Internal metadata dictionary reused the source business concept bridge because target name '{target.name}' exactly matches canonical attribute '{concept.attribute}' for '{concept.display_name}'."
                )
        return explanations

    def explain_canonical_alignment(
        self,
        source: ColumnProfile,
        target: ColumnProfile,
        *,
        prefer_metadata_text: bool = False,
    ) -> list[str]:
        source_matches = {match.concept_id: match for match in self.match_canonical_concepts(source, prefer_metadata_text=prefer_metadata_text)}
        target_matches = {match.concept_id: match for match in self.match_canonical_concepts(target, prefer_metadata_text=prefer_metadata_text)}
        shared = set(source_matches) & set(target_matches)
        explanations: list[str] = []
        for concept_id in sorted(shared):
            concept = self._canonical_concepts_by_id.get(concept_id)
            if concept is None:
                continue
            explanations.append(
                f"Canonical glossary aligns both fields to business concept '{concept.display_name}' ({concept.concept_id})."
            )
            overlay_aliases = self._overlay_canonical_aliases_by_concept.get(concept_id, set())
            matched_overlay_aliases = sorted(
                overlay_aliases.intersection(source_matches[concept_id].matched_aliases)
                | overlay_aliases.intersection(target_matches[concept_id].matched_aliases)
            )
            if matched_overlay_aliases and self._active_overlay_name:
                explanations.append(
                    f"Custom knowledge overlay '{self._active_overlay_name}' extended canonical concept '{concept.display_name}' with alias(es): {', '.join(matched_overlay_aliases)}."
                )
        if not shared:
            for concept_id in sorted(
                self._pairwise_canonical_attribute_bridge(
                    source,
                    target,
                    prefer_metadata_text=prefer_metadata_text,
                )
            ):
                concept = self._canonical_concepts_by_id.get(concept_id)
                if concept is None:
                    continue
                explanations.append(
                    f"Canonical glossary attribute bridge aligns source concept '{concept.display_name}' ({concept.concept_id}) to the generic target attribute '{concept.attribute}'."
                )
        return explanations

    def describe_profile(self, profile: ColumnProfile) -> list[str]:
        descriptions: list[str] = []
        for match in self.match_concepts(profile):
            concept = self._concepts_by_id.get(match.concept_id)
            if concept is None:
                continue
            prefix = f"{profile.name} matches internal concept '{concept.canonical_name}'"
            if match.contexts:
                context_details = ", ".join(self._format_context(context) for context in match.contexts[:3])
                descriptions.append(f"{prefix} via {context_details}.")
            else:
                descriptions.append(f"{prefix} via curated aliases and multilingual metadata.")
        return descriptions

    def match_concepts(self, profile: ColumnProfile, *, prefer_metadata_text: bool = False) -> list[ConceptMatch]:
        """Match one schema profile against the knowledge-layer concept index."""

        cache_key = self._profile_match_cache_key(profile, prefer_metadata_text=prefer_metadata_text)
        cached = self._concept_match_cache.get(cache_key)
        if cached is not None:
            return cached

        normalized_name = _normalize_alias(profile.name)
        normalized_profile_name = _normalize_alias(profile.normalized_name)
        profile_tokens = {token for token in normalized_profile_name.split() if token}
        metadata_candidates, metadata_tokens = self._profile_metadata_candidates(profile)
        metadata_exact_strength, metadata_token_strength, _canonical_metadata_exact_strength, _canonical_metadata_token_strength, _canonical_bridge_metadata_strength = self._metadata_match_strengths(
            prefer_metadata_text=prefer_metadata_text
        )

        strengths: dict[str, float] = {}
        matched_aliases: dict[str, set[str]] = {}
        matched_contexts: dict[str, list[KnowledgeFieldContext]] = {}
        for candidate_name in {normalized_name, normalized_profile_name}:
            for concept_id in self._alias_to_concepts.get(candidate_name, set()):
                strengths[concept_id] = max(strengths.get(concept_id, 0.0), 1.0)
                matched_aliases.setdefault(concept_id, set()).add(candidate_name)
            for concept_id, context in self._field_alias_to_contexts.get(candidate_name, []):
                strengths[concept_id] = max(strengths.get(concept_id, 0.0), 1.0)
                matched_aliases.setdefault(concept_id, set()).add(candidate_name)
                matched_contexts.setdefault(concept_id, []).append(context)

        for candidate_name in metadata_candidates:
            for concept_id in self._alias_to_concepts.get(candidate_name, set()):
                strengths[concept_id] = max(strengths.get(concept_id, 0.0), metadata_exact_strength)
                matched_aliases.setdefault(concept_id, set()).add(candidate_name)
            for concept_id, context in self._field_alias_to_contexts.get(candidate_name, []):
                strengths[concept_id] = max(strengths.get(concept_id, 0.0), metadata_exact_strength)
                matched_aliases.setdefault(concept_id, set()).add(candidate_name)
                matched_contexts.setdefault(concept_id, []).append(context)

        for alias, concept_ids in self._alias_to_concepts.items():
            alias_tokens = set(alias.split())
            if len(alias_tokens) < 2:
                continue
            if alias_tokens.issubset(profile_tokens):
                for concept_id in concept_ids:
                    strengths[concept_id] = max(strengths.get(concept_id, 0.0), 0.7)
                    matched_aliases.setdefault(concept_id, set()).add(alias)

        for alias, concept_ids in self._alias_to_concepts.items():
            alias_tokens = set(alias.split())
            if len(alias_tokens) < 2:
                continue
            if alias_tokens.issubset(metadata_tokens):
                for concept_id in concept_ids:
                    strengths[concept_id] = max(strengths.get(concept_id, 0.0), metadata_token_strength)
                    matched_aliases.setdefault(concept_id, set()).add(alias)

        results = [
            ConceptMatch(
                concept_id=concept_id,
                strength=strength,
                matched_aliases=tuple(sorted(matched_aliases.get(concept_id, set()))),
                contexts=tuple(self._unique_contexts(matched_contexts.get(concept_id, []))),
            )
            for concept_id, strength in sorted(strengths.items())
        ]
        self._concept_match_cache[cache_key] = results
        return results

    def match_canonical_concepts(self, profile: ColumnProfile, *, prefer_metadata_text: bool = False) -> list[CanonicalConceptMatch]:
        """Match one schema profile against the canonical business concept registry."""

        cache_key = self._profile_match_cache_key(profile, prefer_metadata_text=prefer_metadata_text)
        cached = self._canonical_concept_match_cache.get(cache_key)
        if cached is not None:
            return cached

        normalized_name = _normalize_alias(profile.name)
        normalized_profile_name = _normalize_alias(profile.normalized_name)
        profile_tokens = {token for token in normalized_profile_name.split() if token}
        metadata_candidates, metadata_tokens = self._profile_metadata_candidates(profile)
        _metadata_exact_strength, _metadata_token_strength, canonical_metadata_exact_strength, canonical_metadata_token_strength, canonical_bridge_metadata_strength = self._metadata_match_strengths(
            prefer_metadata_text=prefer_metadata_text
        )

        strengths: dict[str, float] = {}
        matched_aliases: dict[str, set[str]] = {}

        # --- Phase 1a: exact direct CC match ---
        for candidate_name in {normalized_name, normalized_profile_name}:
            for concept_id in self._canonical_alias_to_concepts.get(candidate_name, set()):
                strengths[concept_id] = max(strengths.get(concept_id, 0.0), 1.0)
                matched_aliases.setdefault(concept_id, set()).add(candidate_name)

        for candidate_name in metadata_candidates:
            for concept_id in self._canonical_alias_to_concepts.get(candidate_name, set()):
                strengths[concept_id] = max(strengths.get(concept_id, 0.0), canonical_metadata_exact_strength)
                matched_aliases.setdefault(concept_id, set()).add(candidate_name)

        # --- Phase 1b: token-subset direct CC match ---
        for alias, concept_ids in self._canonical_alias_to_concepts.items():
            alias_tokens = set(alias.split())
            if len(alias_tokens) < 2:
                continue
            if alias_tokens.issubset(profile_tokens):
                for concept_id in concept_ids:
                    strengths[concept_id] = max(strengths.get(concept_id, 0.0), 0.75)
                    matched_aliases.setdefault(concept_id, set()).add(alias)

        for alias, concept_ids in self._canonical_alias_to_concepts.items():
            alias_tokens = set(alias.split())
            if len(alias_tokens) < 2:
                continue
            if alias_tokens.issubset(metadata_tokens):
                for concept_id in concept_ids:
                    strengths[concept_id] = max(strengths.get(concept_id, 0.0), canonical_metadata_token_strength)
                    matched_aliases.setdefault(concept_id, set()).add(alias)

        # --- Phase 2a: KC→CC bridge — exact alias match via KC space ---
        # Reaches CC concepts whose aliases overlap with the rich KC alias set
        # (multilingual names, SAP/QAD/WD field names from metadata_dict.csv).
        for candidate_name in {normalized_name, normalized_profile_name}:
            for kc_id in self._alias_to_concepts.get(candidate_name, set()):
                for cc_id in self._kc_to_cc.get(kc_id, set()).intersection(
                    self._canonical_alias_to_concepts.get(candidate_name, set())
                ):
                    if cc_id not in strengths:  # don't downgrade a direct Phase-1 match
                        strengths[cc_id] = 0.8
                        matched_aliases.setdefault(cc_id, set()).add(candidate_name)

        for candidate_name in metadata_candidates:
            for kc_id in self._alias_to_concepts.get(candidate_name, set()):
                for cc_id in self._kc_to_cc.get(kc_id, set()).intersection(
                    self._canonical_alias_to_concepts.get(candidate_name, set())
                ):
                    if cc_id not in strengths:
                        strengths[cc_id] = canonical_bridge_metadata_strength
                        matched_aliases.setdefault(cc_id, set()).add(candidate_name)

        # --- Phase 2b: KC→CC bridge — token-subset via KC aliases ---
        if profile_tokens:
            for alias, kc_ids in self._alias_to_concepts.items():
                alias_tokens = set(alias.split())
                if len(alias_tokens) < 2:
                    continue
                if alias_tokens.issubset(profile_tokens):
                    for kc_id in kc_ids:
                        for cc_id in self._kc_to_cc.get(kc_id, set()).intersection(
                            self._canonical_alias_to_concepts.get(alias, set())
                        ):
                            if cc_id not in strengths:
                                strengths[cc_id] = 0.6
                                matched_aliases.setdefault(cc_id, set()).add(alias)

        if metadata_tokens:
            for alias, kc_ids in self._alias_to_concepts.items():
                alias_tokens = set(alias.split())
                if len(alias_tokens) < 2:
                    continue
                if alias_tokens.issubset(metadata_tokens):
                    for kc_id in kc_ids:
                        for cc_id in self._kc_to_cc.get(kc_id, set()).intersection(
                            self._canonical_alias_to_concepts.get(alias, set())
                        ):
                            if cc_id not in strengths:
                                strengths[cc_id] = 0.5
                                matched_aliases.setdefault(cc_id, set()).add(alias)

        results = [
            CanonicalConceptMatch(
                concept_id=concept_id,
                strength=strength,
                matched_aliases=tuple(sorted(matched_aliases.get(concept_id, set()))),
            )
            for concept_id, strength in sorted(strengths.items())
        ]
        self._canonical_concept_match_cache[cache_key] = results
        return results

    def _canonical_match_details(
        self,
        matches_by_id: dict[str, CanonicalConceptMatch],
    ) -> list[CanonicalConceptMatchDetail]:
        details: list[CanonicalConceptMatchDetail] = []
        for concept_id, match in sorted(
            matches_by_id.items(),
            key=lambda item: (-item[1].strength, item[0]),
        ):
            concept = self._canonical_concepts_by_id.get(concept_id)
            details.append(
                CanonicalConceptMatchDetail(
                    concept_id=concept_id,
                    display_name=concept.display_name if concept is not None else concept_id,
                    strength=round(match.strength, 4),
                )
            )
        return details

    @property
    def kc_to_cc_bridge_size(self) -> int:
        """Number of KC concepts that have at least one CC reachable via alias bridge."""
        return len(self._kc_to_cc)

    def _rebuild_kc_cc_bridge(self) -> None:
        """Build (or rebuild) the alias-overlap bridge from KC concepts to CC concepts.

        For every KC alias that is also a CC alias, we create a link
        KC_concept_id → CC_concept_id.  This lets match_canonical_concepts()
        use the rich KC alias space (SAP/QAD/WD field names, multilingual terms)
        to reach CC concepts that would otherwise be unreachable.
        """
        self._kc_to_cc.clear()
        for kc_concept_id, kc_concept in self._concepts_by_id.items():
            for alias in kc_concept.aliases:
                for cc_id in self._canonical_alias_to_concepts.get(alias, ()):
                    self._kc_to_cc.setdefault(kc_concept_id, set()).add(cc_id)
        # Also bridge through field_alias_to_contexts where the stored concept_id
        # happens to be a CC concept_id (registered by _auto_register_sap_field_aliases).
        for alias, entries in self._field_alias_to_contexts.items():
            canonical_ids = {stored_id for stored_id, _ctx in entries if stored_id in self._canonical_concepts_by_id}
            has_noncanonical_entry = any(stored_id not in self._canonical_concepts_by_id for stored_id, _ctx in entries)
            if len(canonical_ids) != 1 or has_noncanonical_entry:
                continue
            # Promote a field-context alias into the direct canonical alias index only
            # when the alias is unambiguous across the loaded runtime.
            self._canonical_alias_to_concepts.setdefault(alias, set()).update(canonical_ids)

    def reseed_from_files(self) -> dict:
        """Force reload from all source files and re-persist to DB.

        Returns stats about the newly seeded knowledge base.
        """
        self._clear_runtime_state()
        self._load_canonical_glossary()
        self._load()
        self._seed_to_db(self._compute_source_hash())
        self._last_runtime_source = "source_files"
        self._apply_active_overlay()
        self._rebuild_kc_cc_bridge()
        return {
            "concept_count":         len(self._concepts_by_id),
            "canonical_count":       len(self._canonical_concepts_by_id),
            "alias_count":           len(self._alias_to_concepts),
            "canonical_alias_count": len(self._canonical_alias_to_concepts),
        }

    # ------------------------------------------------------------------
    # Internal DB-persistence helpers
    # ------------------------------------------------------------------

    _SOURCE_PATHS: tuple[str, ...] = (
        "DEFAULT_METADATA_DICT_PATH",
        "DEFAULT_METADATA_WORKBOOK_PATH",
        "DEFAULT_CANONICAL_GLOSSARY_PATH",
        "DEFAULT_SAP_TABLES_PATH",
        "DEFAULT_QAD_TABLES_PATH",
        "DEFAULT_WORKDAY_ENTITIES_PATH",
        "DEFAULT_HRDH_TABLE_COLUMNS_PATH",
        "DEFAULT_QUICKBOOKS_TABLES_PATH",
        "DEFAULT_WD_XSD_OVERLAY_PATH",
        "DEFAULT_HRDH_OVERLAY_PATH",
        "DEFAULT_WD_DATAHUB_OVERLAY_PATH",
        "DEFAULT_QB_OVERLAY_PATH",
        "DEFAULT_CANONICAL_FIELD_CONTEXT_ENRICHMENT_PATH",
        "DEFAULT_SAP_KNOWLEDGE_AVAILABLE_TAGS_PATH",
    )

    def _compute_source_hash(self) -> str:
        """Return a stable hash of all source-file mtime+size pairs.

        When any source file changes, the hash changes → triggers re-seed.
        Missing files contribute the string 'missing'.
        """
        import hashlib as _hl
        h = _hl.md5(usedforsecurity=False)
        h.update(f"privacy_classifier:{CANONICAL_PRIVACY_CLASSIFIER_VERSION}".encode())
        import sys
        module = sys.modules[__name__]
        for attr_name in self._SOURCE_PATHS:
            path: Path | None = getattr(module, attr_name, None)
            if path is None:
                # Fall back to instance attribute (canonical_glossary_path etc.)
                path = getattr(self, attr_name.lower().replace("default_", "").replace("_path", "") + "_path", None)
            if path is not None and path.exists():
                stat = path.stat()
                h.update(f"{path}:{stat.st_mtime}:{stat.st_size}".encode())
            else:
                h.update(f"{attr_name}:missing".encode())
        return h.hexdigest()

    def _clear_runtime_state(self) -> None:
        self._concepts_by_id.clear()
        self._alias_to_concepts.clear()
        self._field_alias_to_contexts.clear()
        self._canonical_concepts_by_id.clear()
        self._canonical_alias_to_concepts.clear()
        self._concept_match_cache.clear()
        self._canonical_concept_match_cache.clear()
        self._active_overlay_name = None
        self._overlay_aliases_by_concept.clear()
        self._overlay_canonical_aliases_by_concept.clear()
        clear_normalization_overrides()

    def _read_base_knowledge_rows(self) -> tuple[list[str], list[dict[str, str]]]:
        if not self.csv_path.exists():
            raise KeyError("Base knowledge registry file is not available.")

        decoded = self._decode_csv_payload(self.csv_path.read_bytes())
        reader = csv.DictReader(decoded.splitlines())
        if not reader.fieldnames:
            raise ValueError("Knowledge registry CSV must include a header row.")
        fieldnames = [str(header or "").strip() for header in reader.fieldnames if str(header or "").strip()]
        rows = [{header: str(row.get(header) or "").strip() for header in fieldnames} for row in reader]
        return fieldnames, rows

    def _write_base_knowledge_rows(self, fieldnames: list[str], rows: list[dict[str, str]]) -> None:
        with self.csv_path.open("w", encoding="utf-8", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(rows)

    def _base_knowledge_row(self, rows: Iterable[dict[str, str]], concept_id: str) -> dict[str, str] | None:
        normalized_concept_id = _normalize_alias(str(concept_id or ""))
        for row in rows:
            row_concept_id = _normalize_alias(str(row.get("Naziv (Engleski)") or ""))
            if row_concept_id == normalized_concept_id:
                return row
        return None

    def _build_base_knowledge_record(self, concept_id: str) -> KnowledgeConceptBaseRecord | None:
        try:
            _fieldnames, rows = self._read_base_knowledge_rows()
        except (KeyError, ValueError):
            return None

        row = self._base_knowledge_row(rows, concept_id)
        if row is None:
            return None
        return KnowledgeConceptBaseRecord(
            domain=str(row.get("Kategorija/Domen") or "").strip(),
            english_name=str(row.get("Naziv (Engleski)") or "").strip(),
            serbian_name=str(row.get("Naziv (Srpski)") or "").strip(),
            abbreviations=str(row.get("Skracenice") or "").strip(),
            alternative_names=str(row.get("Alternativni nazivi") or "").strip(),
            data_type=str(row.get("Tip podatka") or "").strip(),
            typical_length=str(row.get("Tipicna duzina") or "").strip(),
            example_value=str(row.get("Primer vrednosti") or "").strip(),
        )

    def _base_knowledge_concept_ids(self) -> set[str]:
        if not self.csv_path.exists():
            return set()

        for encoding in ("utf-8-sig", "utf-8", "cp1250", "cp1252", "latin-1"):
            try:
                with self.csv_path.open("r", encoding=encoding, newline="") as handle:
                    reader = csv.DictReader(handle)
                    return {
                        _normalize_alias(str(row.get("Naziv (Engleski)") or ""))
                        for row in reader
                        if _normalize_alias(str(row.get("Naziv (Engleski)") or ""))
                    }
            except UnicodeDecodeError:
                continue
        return set()

    def _classify_knowledge_concept_source(self, concept_id: str, *, editable: bool) -> str:
        if editable:
            return "base_registry"
        if concept_id.startswith("sap_knowledge_available."):
            return "generated_runtime"
        return "derived_runtime"

    def _refresh_after_canonical_authoring_change(self) -> None:
        """Reload the canonical slice without reparsing the full metadata source stack."""
        seed_meta = knowledge_runtime_repository.get_seed_meta()
        knowledge_dicts, _canonical_dicts, canonical_context_dicts = knowledge_runtime_repository.load_runtime_snapshot()
        if seed_meta is None or not knowledge_dicts:
            self.refresh()
            return

        self._clear_runtime_state()
        self._load_persisted_knowledge_concepts(knowledge_dicts)
        self._load_canonical_glossary()
        self._load_persisted_canonical_field_contexts(canonical_context_dicts)
        canonical_field_contexts: list[tuple[str, KnowledgeFieldContext]] = []
        seen_canonical_contexts: set[tuple[str, str, str, str, str, str, str, str]] = set()
        for entries in self._field_alias_to_contexts.values():
            for concept_id, context in entries:
                if concept_id not in self._canonical_concepts_by_id:
                    continue
                context_key = (
                    concept_id,
                    context.system,
                    context.object_name,
                    context.field_name,
                    context.category,
                    context.object_description,
                    context.field_description,
                    context.note,
                )
                if context_key in seen_canonical_contexts:
                    continue
                seen_canonical_contexts.add(context_key)
                canonical_field_contexts.append((concept_id, context))

        knowledge_runtime_repository.sync_canonical_runtime(
            list(self._canonical_concepts_by_id.values()),
            canonical_field_contexts,
            source_hash=self._compute_source_hash(),
            concept_count=len(self._concepts_by_id),
        )
        self._last_runtime_source = "canonical_authoring_sync"
        self._apply_active_overlay()
        self._rebuild_kc_cc_bridge()

    def _load_from_db(self) -> None:
        """Populate in-memory knowledge state directly from the SQLite DB."""
        knowledge_dicts, canonical_dicts, canonical_context_dicts = knowledge_runtime_repository.load_runtime_snapshot()

        self._load_persisted_canonical_concepts(canonical_dicts)
        self._load_persisted_knowledge_concepts(knowledge_dicts)
        self._load_persisted_canonical_field_contexts(canonical_context_dicts)

    def _load_persisted_canonical_concepts(self, canonical_dicts: list[dict]) -> None:
        for d in canonical_dicts:
            canonical_aliases = frozenset(filter_canonical_aliases(d["aliases"]))
            for alias in canonical_aliases:
                self._canonical_alias_to_concepts.setdefault(alias, set()).add(d["concept_id"])
            self._canonical_concepts_by_id[d["concept_id"]] = CanonicalBusinessConcept(
                concept_id=d["concept_id"],
                entity=d["entity"],
                attribute=d["attribute"],
                display_name=d["display_name"],
                description=d["description"],
                data_type=d["data_type"],
                aliases=canonical_aliases,
                privacy=CanonicalPrivacyMetadataRecord(
                    is_pii=bool(d.get("is_pii")),
                    is_gdpr_special_category=bool(d.get("is_gdpr_special_category")),
                    pii_categories=tuple(str(value) for value in d.get("pii_categories", [])),
                    data_subject_types=tuple(str(value) for value in d.get("data_subject_types", [])),
                ),
            )

    def _load_persisted_knowledge_concepts(self, knowledge_dicts: list[dict]) -> None:
        for d in knowledge_dicts:
            contexts = tuple(
                KnowledgeFieldContext(
                    system=ctx["system"],
                    object_name=ctx["object_name"],
                    field_name=ctx["field_name"],
                    category=ctx["category"],
                    object_description=ctx["object_description"],
                    field_description=ctx["field_description"],
                    note=ctx["note"],
                )
                for ctx in d["contexts"]
            )
            concept = KnowledgeConcept(
                concept_id=d["concept_id"],
                domain=d["domain"],
                canonical_name=d["canonical_name"],
                aliases=frozenset(d["aliases"]),
                contexts=contexts,
                context_terms=frozenset(d["context_terms"]),
            )
            self._concepts_by_id[d["concept_id"]] = concept
            for alias in concept.aliases:
                self._alias_to_concepts.setdefault(alias, set()).add(d["concept_id"])
            for ctx in contexts:
                alias = _normalize_alias(ctx.field_name)
                if alias:
                    self._field_alias_to_contexts.setdefault(alias, []).append((d["concept_id"], ctx))

    def _load_persisted_canonical_field_contexts(self, canonical_context_dicts: list[dict]) -> None:
        for ctx in canonical_context_dicts:
            concept_id = ctx["concept_id"]
            if concept_id not in self._canonical_concepts_by_id:
                continue
            context = KnowledgeFieldContext(
                system=ctx["system"],
                object_name=ctx["object_name"],
                field_name=ctx["field_name"],
                category=ctx["category"],
                object_description=ctx["object_description"],
                field_description=ctx["field_description"],
                note=ctx["note"],
            )
            alias = _normalize_alias(context.field_name)
            if alias:
                self._field_alias_to_contexts.setdefault(alias, []).append((concept_id, context))

    def _seed_to_db(self, source_hash: str) -> None:
        """Persist current in-memory knowledge to the DB."""
        canonical_field_contexts: list[tuple[str, KnowledgeFieldContext]] = []
        seen_canonical_contexts: set[tuple[str, str, str, str, str, str, str, str]] = set()
        for entries in self._field_alias_to_contexts.values():
            for concept_id, context in entries:
                if concept_id not in self._canonical_concepts_by_id:
                    continue
                context_key = (
                    concept_id,
                    context.system,
                    context.object_name,
                    context.field_name,
                    context.category,
                    context.object_description,
                    context.field_description,
                    context.note,
                )
                if context_key in seen_canonical_contexts:
                    continue
                seen_canonical_contexts.add(context_key)
                canonical_field_contexts.append((concept_id, context))

        knowledge_runtime_repository.replace_runtime_snapshot(
            list(self._concepts_by_id.values()),
            list(self._canonical_concepts_by_id.values()),
            canonical_field_contexts,
            source_hash=source_hash,
        )

    def _load_canonical_glossary(self) -> None:
        if not self.canonical_glossary_path.exists():
            return

        with self.canonical_glossary_path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            for row in reader:
                concept_id = str(row.get("concept_id") or "").strip()
                if not concept_id:
                    continue
                display_name = str(row.get("display_name") or concept_id).strip()
                aliases = set(
                    filter_canonical_aliases(
                        (
                            display_name,
                            concept_id.replace(".", " "),
                            *_split_values(str(row.get("aliases") or "")),
                        )
                    )
                )
                self._register_canonical_concept(
                    concept_id=concept_id,
                    entity=str(row.get("entity") or "general").strip() or "general",
                    attribute=str(row.get("attribute") or concept_id.split(".")[-1]).strip() or concept_id.split(".")[-1],
                    display_name=display_name,
                    description=str(row.get("description") or "").strip(),
                    data_type=str(row.get("data_type") or "").strip(),
                    aliases=aliases,
                    privacy=_merge_privacy_metadata(
                        CanonicalPrivacyMetadataRecord(
                            is_pii=_parse_boolean_flag(row.get("is_pii")),
                            is_gdpr_special_category=_parse_boolean_flag(row.get("is_gdpr_special_category")),
                            pii_categories=_parse_privacy_tags(row.get("pii_categories")),
                            data_subject_types=_parse_privacy_tags(row.get("data_subject_types")),
                        ),
                        _infer_canonical_privacy_metadata(
                            concept_id=concept_id,
                            entity=str(row.get("entity") or "general").strip() or "general",
                            attribute=str(row.get("attribute") or concept_id.split(".")[-1]).strip() or concept_id.split(".")[-1],
                            display_name=display_name,
                            description=str(row.get("description") or "").strip(),
                            aliases=aliases,
                        ),
                    ),
                )

    def _load(self) -> None:
        if not self.csv_path.exists():
            return

        for encoding in ("utf-8-sig", "utf-8", "cp1250", "cp1252", "latin-1"):
            try:
                with self.csv_path.open("r", encoding=encoding, newline="") as handle:
                    reader = csv.DictReader(handle)
                    for row in reader:
                        self._register_csv_row(row)
                self._load_workbook_contexts()
                self._load_sap_knowledge_available_tags(DEFAULT_SAP_KNOWLEDGE_AVAILABLE_TAGS_PATH)
                return
            except UnicodeDecodeError:
                self._concepts_by_id.clear()
                self._alias_to_concepts.clear()
                self._field_alias_to_contexts.clear()

    def _register_csv_row(self, row: dict[str, object]) -> None:
        concept_id = _normalize_alias(str(row.get("Naziv (Engleski)", "")))
        if not concept_id:
            return
        domain = str(row.get("Kategorija/Domen") or "general").strip()
        canonical_name = str(row.get("Naziv (Engleski)") or concept_id).strip()
        aliases = self._extract_aliases(row, ALIAS_FIELDS, MULTI_VALUE_FIELDS)
        self._register_concept(concept_id, domain, canonical_name, aliases=aliases)

    def _apply_active_overlay(self) -> None:
        active_overlay = persistence_service.get_active_knowledge_overlay_version()
        if active_overlay is None or active_overlay.overlay_id is None:
            return

        self._active_overlay_name = active_overlay.name
        entries = persistence_service.get_knowledge_overlay_entries(active_overlay.overlay_id)
        overlay_abbreviations: dict[str, str] = {}
        overlay_synonyms: dict[str, set[str]] = {}

        for entry in entries:
            canonical_term = entry.normalized_canonical_term
            alias = entry.normalized_alias
            if not canonical_term or not alias:
                continue

            if entry.entry_type == "field_alias":
                self._register_concept(
                    canonical_term,
                    entry.domain or "overlay",
                    entry.canonical_term,
                    aliases={alias},
                )
                self._overlay_aliases_by_concept.setdefault(canonical_term, set()).add(alias)

                context_patch = _parse_context_patch_note(entry.note or "")
                canonical_concept_id = entry.canonical_concept_id or self.resolve_canonical_concept_id(entry.canonical_term)
                if context_patch and canonical_concept_id and canonical_concept_id in self._canonical_concepts_by_id:
                    object_name = str(context_patch.get("object") or "").strip()
                    field_name = str(context_patch.get("field") or "").strip()
                    if (not object_name or not field_name) and "." in entry.alias:
                        object_part, field_part = entry.alias.split(".", 1)
                        object_name = object_name or object_part.strip()
                        field_name = field_name or field_part.strip()
                    context = KnowledgeFieldContext(
                        system=str(context_patch.get("system") or entry.source_system or "").strip(),
                        object_name=object_name,
                        field_name=field_name or entry.alias,
                        category=self._canonical_concepts_by_id[canonical_concept_id].entity,
                        object_description=str(context_patch.get("object_description") or "").strip(),
                        field_description=str(context_patch.get("field_description") or "").strip(),
                        note=str(context_patch.get("note") or "").strip(),
                    )
                    self._register_field_context(context.field_name, canonical_concept_id, context)
            elif entry.entry_type == "concept_alias":
                canonical_concept_id = entry.canonical_concept_id or self.resolve_canonical_concept_id(entry.canonical_term)
                if canonical_concept_id is None:
                    continue
                if canonical_concept_id not in self._canonical_concepts_by_id:
                    self._register_canonical_concept(
                        concept_id=canonical_concept_id,
                        entity=canonical_concept_id.split(".", 1)[0] if "." in canonical_concept_id else "general",
                        attribute=canonical_concept_id.split(".", 1)[1] if "." in canonical_concept_id else canonical_concept_id,
                        display_name=entry.canonical_term,
                        description=f"Overlay-only canonical concept proposed by {entry.source_system or 'knowledge overlay'}.",
                        data_type="",
                        aliases={_normalize_alias(entry.canonical_term)},
                    )
                self._register_canonical_concept(
                    concept_id=canonical_concept_id,
                    entity=self._canonical_concepts_by_id.get(canonical_concept_id).entity if canonical_concept_id in self._canonical_concepts_by_id else "general",
                    attribute=self._canonical_concepts_by_id.get(canonical_concept_id).attribute if canonical_concept_id in self._canonical_concepts_by_id else canonical_concept_id.split(".")[-1],
                    display_name=self._canonical_concepts_by_id.get(canonical_concept_id).display_name if canonical_concept_id in self._canonical_concepts_by_id else entry.canonical_term,
                    description=self._canonical_concepts_by_id.get(canonical_concept_id).description if canonical_concept_id in self._canonical_concepts_by_id else "",
                    data_type=self._canonical_concepts_by_id.get(canonical_concept_id).data_type if canonical_concept_id in self._canonical_concepts_by_id else "",
                    aliases={alias},
                )
                self._overlay_canonical_aliases_by_concept.setdefault(canonical_concept_id, set()).add(alias)
            elif entry.entry_type == "abbreviation":
                overlay_abbreviations[alias] = canonical_term
            elif entry.entry_type == "synonym":
                overlay_synonyms.setdefault(canonical_term, set()).add(alias)

        configure_normalization_overrides(overlay_abbreviations, overlay_synonyms)

    def _load_workbook_contexts(self) -> None:
        sap_descriptions, sap_field_descriptions = self._load_sap_table_descriptions()
        qad_descriptions = self._load_qad_table_descriptions()
        workday_entities, workday_fields = self._load_workday_entity_descriptions()
        self._load_metadata_mapping_sheet(sap_descriptions, sap_field_descriptions, qad_descriptions, workday_entities, workday_fields)
        self._auto_register_sap_field_aliases(sap_field_descriptions)
        self._load_csv_knowledge_overlay(DEFAULT_WD_XSD_OVERLAY_PATH, source_tag="Workday_Webservice")
        self._load_csv_knowledge_overlay(DEFAULT_HRDH_OVERLAY_PATH, source_tag="HRDH")
        self._load_csv_knowledge_overlay(DEFAULT_WD_DATAHUB_OVERLAY_PATH, source_tag="Workday_HRDH")
        self._load_csv_knowledge_overlay(DEFAULT_QB_OVERLAY_PATH, source_tag="QuickBooks")
        self._load_canonical_field_context_enrichment(DEFAULT_CANONICAL_FIELD_CONTEXT_ENRICHMENT_PATH)

    def _load_canonical_field_context_enrichment(self, csv_path: Path) -> None:
        if not csv_path.exists():
            return

        with csv_path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            for row in reader:
                concept_id = str(row.get("concept_id") or "").strip()
                field_name = str(row.get("field_name") or "").strip()
                if not concept_id or not field_name or concept_id not in self._canonical_concepts_by_id:
                    continue
                concept = self._canonical_concepts_by_id[concept_id]
                context = KnowledgeFieldContext(
                    system=str(row.get("system") or "").strip(),
                    object_name=str(row.get("object_name") or "").strip(),
                    field_name=field_name,
                    category=str(row.get("category") or concept.entity).strip(),
                    object_description=str(row.get("object_description") or "").strip(),
                    field_description=str(row.get("field_description") or "").strip(),
                    note=str(row.get("note") or "").strip(),
                )
                self._register_field_context(field_name, concept_id, context)

    def _load_sap_knowledge_available_tags(self, csv_path: Path) -> None:
        if not csv_path.exists():
            return

        grouped_rows: dict[str, list[dict[str, str]]] = {}
        display_names: dict[str, str] = {}

        with csv_path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            for row in reader:
                description = str(row.get("sap_description") or "").strip()
                field_name = str(row.get("sap_field") or "").strip()
                if not description and not field_name:
                    continue

                group_label = description or field_name
                group_key = _normalize_alias(group_label)
                if not group_key:
                    continue

                grouped_rows.setdefault(group_key, []).append({
                    "sap_module": str(row.get("sap_module") or "").strip(),
                    "sap_table": str(row.get("sap_table") or "").strip(),
                    "sap_field": field_name,
                    "sap_description": description,
                    "knowledge_tag_level": str(row.get("knowledge_tag_level") or "").strip(),
                    "knowledge_tag_reason": str(row.get("knowledge_tag_reason") or "").strip(),
                    "knowledge_tag_note": str(row.get("knowledge_tag_note") or "").strip(),
                })
                display_names.setdefault(group_key, group_label)

        for group_key, rows in grouped_rows.items():
            display_name = display_names.get(group_key) or group_key
            concept_id = self._sap_knowledge_available_concept_id(group_key)
            dominant_module = Counter(row.get("sap_module") or "UNKNOWN" for row in rows).most_common(1)[0][0]
            contexts: list[KnowledgeFieldContext] = []
            for row in rows:
                field_name = row.get("sap_field") or ""
                if not field_name:
                    continue
                contexts.append(
                    KnowledgeFieldContext(
                        system="SAP",
                        object_name=row.get("sap_table") or "",
                        field_name=field_name,
                        category="sap_knowledge_available",
                        object_description="",
                        field_description=row.get("sap_description") or "",
                        note=row.get("knowledge_tag_note") or f"module={dominant_module}",
                    )
                )
            self._register_concept(
                concept_id,
                f"sap_knowledge_available:{dominant_module or 'UNKNOWN'}",
                display_name,
                aliases={display_name},
                contexts=contexts,
            )

    def _sap_knowledge_available_concept_id(self, normalized_label: str) -> str:
        slug = re.sub(r"[^a-z0-9]+", "_", normalized_label.lower()).strip("_")
        if not slug:
            slug = hashlib.md5(normalized_label.encode("utf-8"), usedforsecurity=False).hexdigest()[:12]
        return f"sap_knowledge_available.{slug[:80]}"

    def _auto_register_sap_field_aliases(self, field_descriptions: dict[tuple[str, str], str]) -> None:
        """Auto-register SAP fields from Tbls_Clm as canonical concept aliases and field contexts."""
        field_canonical_targets: dict[str, set[str]] = {}
        field_has_unresolved_meaning: dict[str, bool] = {}
        for (_table, field), description in field_descriptions.items():
            normalized_field = _normalize_alias(field)
            if not normalized_field:
                continue
            canonical_id = self.resolve_canonical_concept_id(description)
            if canonical_id is None:
                field_has_unresolved_meaning[normalized_field] = True
                continue
            field_canonical_targets.setdefault(normalized_field, set()).add(canonical_id)

        registered_field_aliases: set[str] = set()
        for (table, field), description in field_descriptions.items():
            canonical_id = self.resolve_canonical_concept_id(description)
            if canonical_id is None:
                continue
            existing = self._canonical_concepts_by_id.get(canonical_id)
            if existing is None:
                continue
            context = KnowledgeFieldContext(
                system="SAP",
                object_name=table,
                field_name=field,
                category=existing.entity,
                object_description="",
                field_description=description,
                note="",
            )
            self._register_field_context(field, canonical_id, context)
            normalized_field = _normalize_alias(field)
            is_unambiguous_canonical_field = (
                len(field_canonical_targets.get(normalized_field, set())) == 1
                and not field_has_unresolved_meaning.get(normalized_field, False)
            )
            if field not in registered_field_aliases and is_unambiguous_canonical_field:
                self._register_canonical_concept(
                    concept_id=canonical_id,
                    entity=existing.entity,
                    attribute=existing.attribute,
                    display_name=existing.display_name,
                    description=existing.description,
                    data_type=existing.data_type,
                    aliases={field},
                )
                registered_field_aliases.add(field)

    def _load_csv_knowledge_overlay(self, csv_path: Path, source_tag: str = "") -> None:
        """Generic loader: reads a concept_alias CSV and registers aliases into canonical concepts.

        CSV format (same as Admin UI overlay):
            entry_type, canonical_term, canonical_concept_id, alias, domain, source_system, note
        Only rows with entry_type=concept_alias are processed.

        Each alias is registered in both layers:
        - CC layer: as a canonical alias on the CC concept (for direct CC matching)
        - KC layer: as an alias on a KC concept keyed by the canonical_concept_id
                    (so the KC→CC bridge can also reach it)
        """
        if not csv_path.exists():
            return

        import csv as _csv

        with csv_path.open(encoding="utf-8-sig") as fh:
            for row in _csv.DictReader(fh):
                entry_type = (row.get("entry_type") or "").strip()
                if entry_type != "concept_alias":
                    continue
                canonical_concept_id = (row.get("canonical_concept_id") or "").strip()
                alias = _normalize_alias((row.get("alias") or "").strip())
                if not canonical_concept_id or not alias:
                    continue
                concept = self._canonical_concepts_by_id.get(canonical_concept_id)
                if concept is None:
                    continue

                # --- CC layer (direct CC matching) ---
                self._register_canonical_concept(
                    concept_id=canonical_concept_id,
                    entity=concept.entity,
                    attribute=concept.attribute,
                    display_name=concept.display_name,
                    description=concept.description,
                    data_type=concept.data_type,
                    aliases={alias},
                )

                # --- KC layer (dual-registration so KC→CC bridge fires) ---
                # The KC concept uses canonical_concept_id as its concept_id so that
                # _rebuild_kc_cc_bridge() can trivially link it to the CC concept
                # (they share the display_name alias at minimum).
                domain = (row.get("domain") or concept.entity or source_tag or "overlay").strip()
                self._register_concept(
                    canonical_concept_id,
                    domain,
                    concept.display_name,
                    aliases={alias},
                )

    def _load_sap_table_descriptions(self) -> tuple[dict[str, str], dict[tuple[str, str], str]]:
        table_descriptions: dict[str, str] = {}
        field_descriptions: dict[tuple[str, str], str] = {}
        if not self.sap_tables_path.exists():
            return table_descriptions, field_descriptions
        workbook = load_workbook(self.sap_tables_path, read_only=True, data_only=True)
        # Load field-level data from Tbls_Clm using the actual header names.
        if "Tbls_Clm" in workbook.sheetnames:
            worksheet = workbook["Tbls_Clm"]
            header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
            header_index = {
                str(value or "").strip(): index
                for index, value in enumerate(header_row)
                if str(value or "").strip()
            }
            table_index = header_index.get("Table")
            field_index = header_index.get("Field")
            description_index = header_index.get("Description")
            if table_index is None or field_index is None or description_index is None:
                return table_descriptions, field_descriptions
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                table_name = _normalize_alias(str(row[table_index] or ""))
                field_name = _normalize_alias(str(row[field_index] or ""))
                description = str(row[description_index] or "").strip()
                if table_name and field_name and description:
                    field_descriptions[(table_name, field_name)] = description
        # Load table-level descriptions from module sheets (header at row 4, col 0 == "Table")
        for sheet_name in workbook.sheetnames:
            if sheet_name == "Tbls_Clm":
                continue
            worksheet = workbook[sheet_name]
            header = [value for value in next(worksheet.iter_rows(min_row=4, max_row=4, values_only=True), ())]
            if not header or header[0] != "Table":
                continue
            for row in worksheet.iter_rows(min_row=5, values_only=True):
                table_name = _normalize_alias(str(row[0] or ""))
                if not table_name:
                    continue
                description = " ".join(
                    part.strip()
                    for part in (str(row[1] or "").strip(), str(row[4] or "").strip(), sheet_name.replace("_", " "))
                    if part and part.strip()
                )
                table_descriptions[table_name] = description
        return table_descriptions, field_descriptions

    def _load_qad_table_descriptions(self) -> dict[str, str]:
        descriptions: dict[str, str] = {}
        if not self.qad_tables_path.exists():
            return descriptions
        workbook = load_workbook(self.qad_tables_path, read_only=True, data_only=True)
        # Core Tables Catalog: col 0=name, 2=desc, 3=module, 4=sub-area, 5=key fields
        if "Core Tables Catalog" in workbook.sheetnames:
            for row in workbook["Core Tables Catalog"].iter_rows(min_row=2, values_only=True):
                table_name = _normalize_alias(str(row[0] or ""))
                if not table_name:
                    continue
                description = " ".join(
                    part.strip()
                    for part in (
                        str(row[2] or "").strip(),
                        str(row[3] or "").strip(),
                        str(row[4] or "").strip(),
                        str(row[5] or "").strip(),
                    )
                    if part and part.strip()
                )
                descriptions[table_name] = description
        # TablesList: col 0=name, 1=description — adds breadth (930 tables)
        if "TablesList" in workbook.sheetnames:
            for row in workbook["TablesList"].iter_rows(min_row=2, values_only=True):
                table_name = _normalize_alias(str(row[0] or ""))
                if not table_name or table_name in descriptions:
                    continue
                description = str(row[1] or "").strip()
                if description:
                    descriptions[table_name] = description
        return descriptions

    def _load_workday_entity_descriptions(self) -> tuple[dict[str, str], dict[tuple[str, str], str]]:
        entity_descriptions: dict[str, str] = {}
        field_descriptions: dict[tuple[str, str], str] = {}
        if not self.workday_entities_path.exists():
            return entity_descriptions, field_descriptions
        workbook = load_workbook(self.workday_entities_path, read_only=True, data_only=True)

        if "Entities" in workbook.sheetnames:
            worksheet = workbook["Entities"]
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                entity_name = _normalize_alias(str(row[1] or ""))
                if not entity_name:
                    continue
                entity_descriptions[entity_name] = " ".join(
                    part.strip()
                    for part in (
                        str(row[2] or "").strip(),
                        str(row[3] or "").strip(),
                        str(row[7] or "").strip(),
                    )
                    if part and part.strip()
                )

        for sheet_name in workbook.sheetnames:
            if not sheet_name.endswith("_Fields"):
                continue
            entity_name = _normalize_alias(sheet_name.replace("_Fields", "").replace("_", " "))
            worksheet = workbook[sheet_name]
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                field_name = _normalize_alias(str(row[1] or ""))
                if not field_name:
                    continue
                description = str(row[4] or "").strip()
                if description:
                    field_descriptions[(entity_name, field_name)] = description
        return entity_descriptions, field_descriptions

    def _load_metadata_mapping_sheet(
        self,
        sap_descriptions: dict[str, str],
        sap_field_descriptions: dict[tuple[str, str], str],
        qad_descriptions: dict[str, str],
        workday_entities: dict[str, str],
        workday_fields: dict[tuple[str, str], str],
    ) -> None:
        if not self.metadata_workbook_path.exists():
            return
        workbook = load_workbook(self.metadata_workbook_path, read_only=True, data_only=True)
        if "Metadata_Mapping" not in workbook.sheetnames:
            return
        worksheet = workbook["Metadata_Mapping"]
        headers = [str(value or "") for value in next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))]
        for row_values in worksheet.iter_rows(min_row=2, values_only=True):
            row = {headers[index]: row_values[index] for index in range(min(len(headers), len(row_values)))}
            concept_id = _normalize_alias(str(row.get("Naziv (EN)", "")))
            if not concept_id:
                continue
            category = str(row.get("Kategorija") or "general").strip()
            canonical_name = str(row.get("Naziv (EN)") or concept_id).strip()
            aliases = self._extract_aliases(row, ("Naziv (EN)", "Naziv (SR)", "SAP polje", "QAD polje", "Workday polje"), ())
            contexts: list[KnowledgeFieldContext] = []

            sap_table = str(row.get("SAP tabela") or "").strip()
            sap_field = str(row.get("SAP polje") or "").strip()
            if sap_table and sap_field:
                napomena = str(row.get("Napomena") or "").strip()
                sap_auto_desc = sap_field_descriptions.get(
                    (_normalize_alias(sap_table), _normalize_alias(sap_field)), ""
                )
                context = KnowledgeFieldContext(
                    system="SAP",
                    object_name=sap_table,
                    field_name=sap_field,
                    category=category,
                    object_description=sap_descriptions.get(_normalize_alias(sap_table), ""),
                    field_description=napomena or sap_auto_desc,
                    note=napomena,
                )
                contexts.append(context)
                self._register_field_context(sap_field, concept_id, context)
                aliases.add(_normalize_alias(f"{sap_table} {sap_field}"))

            qad_table = str(row.get("QAD tabela") or "").strip()
            qad_field = str(row.get("QAD polje") or "").strip()
            if qad_table and qad_field:
                context = KnowledgeFieldContext(
                    system="QAD",
                    object_name=qad_table,
                    field_name=qad_field,
                    category=category,
                    object_description=qad_descriptions.get(_normalize_alias(qad_table), ""),
                    field_description=str(row.get("Napomena") or "").strip(),
                    note=str(row.get("Napomena") or "").strip(),
                )
                contexts.append(context)
                self._register_field_context(qad_field, concept_id, context)
                aliases.add(_normalize_alias(f"{qad_table} {qad_field}"))

            workday_entity = str(row.get("Workday entitet") or "").strip()
            workday_field = str(row.get("Workday polje") or "").strip()
            if workday_entity and workday_field:
                normalized_entity = _normalize_alias(workday_entity)
                normalized_field = _normalize_alias(workday_field)
                context = KnowledgeFieldContext(
                    system="Workday",
                    object_name=workday_entity,
                    field_name=workday_field,
                    category=category,
                    object_description=workday_entities.get(normalized_entity, ""),
                    field_description=workday_fields.get((normalized_entity, normalized_field), str(row.get("Napomena") or "").strip()),
                    note=str(row.get("Napomena") or "").strip(),
                )
                contexts.append(context)
                self._register_field_context(workday_field, concept_id, context)
                aliases.add(_normalize_alias(f"{workday_entity} {workday_field}"))

            self._register_concept(concept_id, category, canonical_name, aliases=aliases, contexts=contexts)

    def _extract_aliases(
        self,
        row: dict[str, object],
        single_value_fields: Iterable[str],
        multi_value_fields: Iterable[str],
    ) -> set[str]:
        aliases: set[str] = set()
        for field_name in single_value_fields:
            value = str(row.get(field_name) or "").strip()
            normalized = _normalize_alias(value)
            if normalized:
                aliases.add(normalized)
        for field_name in multi_value_fields:
            for value in _split_values(str(row.get(field_name) or "")):
                normalized = _normalize_alias(value)
                if normalized:
                    aliases.add(normalized)
        return aliases

    def _register_field_context(self, alias: str, concept_id: str, context: KnowledgeFieldContext) -> None:
        normalized_alias = _normalize_alias(alias)
        if not normalized_alias:
            return
        self._field_alias_to_contexts.setdefault(normalized_alias, []).append((concept_id, context))

    def _register_concept(
        self,
        concept_id: str,
        domain: str,
        canonical_name: str,
        aliases: Iterable[str] = (),
        contexts: Iterable[KnowledgeFieldContext] = (),
    ) -> None:
        existing = self._concepts_by_id.get(concept_id)
        merged_aliases = set(existing.aliases if existing else ())
        merged_aliases.update(alias for alias in aliases if alias)
        merged_contexts = list(existing.contexts if existing else ())
        merged_contexts.extend(list(contexts))
        unique_contexts = tuple(self._unique_contexts(merged_contexts))
        merged_context_terms = set(existing.context_terms if existing else ())
        merged_context_terms.update(self._context_terms(unique_contexts))
        concept = KnowledgeConcept(
            concept_id=concept_id,
            domain=(existing.domain if existing and existing.domain else domain) or "general",
            canonical_name=(existing.canonical_name if existing and existing.canonical_name else canonical_name) or concept_id,
            aliases=frozenset(merged_aliases),
            contexts=unique_contexts,
            context_terms=frozenset(merged_context_terms),
        )
        self._concepts_by_id[concept_id] = concept
        for alias in concept.aliases:
            self._alias_to_concepts.setdefault(alias, set()).add(concept_id)

    def _register_canonical_concept(
        self,
        concept_id: str,
        entity: str,
        attribute: str,
        display_name: str,
        description: str = "",
        data_type: str = "",
        aliases: Iterable[str] = (),
        privacy: CanonicalPrivacyMetadataRecord | None = None,
    ) -> None:
        existing = self._canonical_concepts_by_id.get(concept_id)
        merged_aliases = set(existing.aliases if existing else ())
        merged_aliases.update(filter_canonical_aliases(aliases))
        concept = CanonicalBusinessConcept(
            concept_id=concept_id,
            entity=existing.entity if existing else entity,
            attribute=existing.attribute if existing else attribute,
            display_name=existing.display_name if existing else display_name,
            description=existing.description if existing else description,
            data_type=existing.data_type if existing else data_type,
            aliases=frozenset(merged_aliases),
            privacy=existing.privacy if existing else (privacy or CanonicalPrivacyMetadataRecord()),
        )
        self._canonical_concepts_by_id[concept_id] = concept
        for alias in concept.aliases:
            self._canonical_alias_to_concepts.setdefault(alias, set()).add(concept_id)

    def _decode_csv_payload(self, payload: bytes) -> str:
        for encoding in ("utf-8-sig", "utf-8", "cp1250", "cp1252", "latin-1"):
            try:
                return payload.decode(encoding)
            except UnicodeDecodeError:
                continue
        raise ValueError("CSV payload could not be decoded with supported encodings.")

    def _context_terms(self, contexts: Iterable[KnowledgeFieldContext]) -> set[str]:
        tokens: set[str] = set()
        for context in contexts:
            for text in (
                context.object_name,
                context.field_name,
                context.category,
                context.object_description,
                context.field_description,
                context.note,
            ):
                normalized = _normalize_alias(text)
                if not normalized:
                    continue
                tokens.update(normalized.split())
        return tokens

    def _unique_contexts(self, contexts: Iterable[KnowledgeFieldContext]) -> list[KnowledgeFieldContext]:
        seen: set[tuple[str, str, str, str, str, str, str]] = set()
        unique_contexts: list[KnowledgeFieldContext] = []
        for context in contexts:
            marker = (
                context.system,
                context.object_name,
                context.field_name,
                context.category,
                context.object_description,
                context.field_description,
                context.note,
            )
            if marker in seen:
                continue
            seen.add(marker)
            unique_contexts.append(context)
        return unique_contexts

    def _format_context(self, context: KnowledgeFieldContext) -> str:
        location = f"{context.object_name}.{context.field_name}" if context.object_name and context.field_name else context.field_name or context.object_name
        if context.object_description:
            return f"{context.system} {location} ({context.object_description})"
        return f"{context.system} {location}"


metadata_knowledge_service = MetadataKnowledgeService()