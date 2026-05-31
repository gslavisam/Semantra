"""Schema-spec upload parsing and normalization for field-oriented metadata files."""

from __future__ import annotations

import csv
from io import BytesIO
import re
from pathlib import Path
from uuid import uuid4

from openpyxl import load_workbook

from app.models.schema import ColumnProfile, SchemaProfile, SpecLayoutHint
from app.services.tabular_upload_service import normalize_rows, parse_tabular_payload
from app.utils.normalization import normalize_name, tokenize_name
from app.utils.tabular import decode_text_payload, is_nullish, normalize_tabular_header


NAME_CANDIDATES = {
    "column",
    "field",
    "field_name",
    "column_name",
    "attribute",
    "name",
    "field name",
    "column name",
}
DESCRIPTION_CANDIDATES = {
    "description",
    "desc",
    "label",
    "comment",
    "remarks",
    "meaning",
}
TYPE_CANDIDATES = {
    "type",
    "data_type",
    "dtype",
    "format",
}
SAMPLE_VALUE_CANDIDATES = {
    "sample",
    "samples",
    "sample value",
    "sample values",
    "example",
    "examples",
    "example value",
    "example values",
}
SPEC_LAYOUT_MAX_COLUMNS = 20
MAX_SPEC_SAMPLE_VALUES = 5
EMBEDDED_SPEC_DELIMITERS = (",", ";", "\t")
EMBEDDED_SPEC_MIN_COLUMNS = 2
EMBEDDED_SPEC_MAX_HEADER_TOKENS = 4
EMBEDDED_SPEC_MAX_HEADER_LENGTH = 48


def build_spec_layout_hint(rows: list[dict[str, object]]) -> SpecLayoutHint | None:
    """Infer a likely schema-spec layout from parsed tabular rows."""

    if not rows:
        return None

    headers = list(rows[0].keys())
    return detect_spec_layout(headers)


def detect_spec_layout(headers: list[str]) -> SpecLayoutHint | None:
    """Detect which headers in a tabular file correspond to name, description, type, and samples."""

    normalized_headers = [normalize_header_candidate(header) for header in headers]
    if not normalized_headers or len(normalized_headers) > SPEC_LAYOUT_MAX_COLUMNS:
        return None

    name_col = find_matching_header(headers, normalized_headers, NAME_CANDIDATES)
    if not name_col:
        return None

    description_col = find_matching_header(headers, normalized_headers, DESCRIPTION_CANDIDATES)
    type_col = find_matching_header(headers, normalized_headers, TYPE_CANDIDATES)
    sample_values_col = find_matching_header(headers, normalized_headers, SAMPLE_VALUE_CANDIDATES)
    if not description_col and not type_col:
        return None

    confidence = 1.0 if description_col and type_col else 0.8
    return SpecLayoutHint(
        name_col=name_col,
        description_col=description_col,
        type_col=type_col,
        sample_values_col=sample_values_col,
        confidence=confidence,
    )


def parse_spec_payload(
    payload: bytes,
    filename: str,
    *,
    header_row_index: int | None = None,
    name_col: str | None = None,
    description_col: str | None = None,
    type_col: str | None = None,
    sample_values_col: str | None = None,
) -> SchemaProfile:
    """Parse an uploaded schema-spec file into the SchemaProfile used by Semantra."""

    rows, _resolved_header_row_index = parse_spec_source_rows(
        payload,
        filename,
        header_row_index=header_row_index,
        name_col=name_col,
        description_col=description_col,
        type_col=type_col,
        sample_values_col=sample_values_col,
    )
    return parse_spec_rows(
        rows,
        dataset_id=str(uuid4()),
        dataset_name=filename,
        name_col=name_col,
        description_col=description_col,
        type_col=type_col,
        sample_values_col=sample_values_col,
    )


def parse_spec_source_rows(
    payload: bytes,
    filename: str,
    *,
    header_row_index: int | None = None,
    name_col: str | None = None,
    description_col: str | None = None,
    type_col: str | None = None,
    sample_values_col: str | None = None,
) -> tuple[list[dict[str, object]], int]:
    """Parse one schema-spec source, allowing a bounded embedded-table fallback for mixed CSV files."""

    requested_header_row_index = max(1, int(header_row_index or 1))
    direct_rows: list[dict[str, object]] | None = None
    direct_error: ValueError | None = None

    if requested_header_row_index == 1:
        try:
            direct_rows = parse_tabular_payload(payload, filename)
        except ValueError as error:
            direct_error = error
        else:
            if _rows_support_requested_spec_layout(
                direct_rows,
                name_col=name_col,
                description_col=description_col,
                type_col=type_col,
                sample_values_col=sample_values_col,
            ):
                return direct_rows, 1

    embedded = _extract_embedded_spec_rows(
        payload,
        filename,
        requested_header_row_index=requested_header_row_index if requested_header_row_index > 1 else None,
    )
    if embedded is not None:
        return embedded

    if direct_rows is not None:
        return direct_rows, 1
    if direct_error is not None:
        raise direct_error
    raise ValueError(f"Could not resolve schema-spec header row {requested_header_row_index} from the uploaded file")


def parse_spec_rows(
    rows: list[dict[str, object]],
    *,
    dataset_id: str,
    dataset_name: str,
    name_col: str | None = None,
    description_col: str | None = None,
    type_col: str | None = None,
    sample_values_col: str | None = None,
) -> SchemaProfile:
    """Convert parsed schema-spec rows into the SchemaProfile used by Semantra."""

    if not rows:
        raise ValueError("Spec upload requires at least one field row")

    hint = resolve_spec_layout(
        rows,
        name_col=name_col,
        description_col=description_col,
        type_col=type_col,
        sample_values_col=sample_values_col,
    )
    columns: list[ColumnProfile] = []

    for row in rows:
        raw_name = row.get(hint.name_col)
        if is_nullish(raw_name):
            continue

        column_name = str(raw_name).strip()
        if not is_plausible_field_name(column_name, hint.name_col):
            continue

        raw_description = row.get(hint.description_col) if hint.description_col else None
        raw_type = row.get(hint.type_col) if hint.type_col else None
        raw_samples = row.get(hint.sample_values_col) if hint.sample_values_col else None
        description = "" if is_nullish(raw_description) else str(raw_description).strip()
        declared_type = "" if is_nullish(raw_type) else str(raw_type).strip()
        sample_values = parse_spec_sample_values(raw_samples)
        mapped_dtype = map_spec_type(raw_type)
        detected_patterns = [map_spec_pattern(mapped_dtype)]

        columns.append(
            ColumnProfile(
                name=column_name,
                normalized_name=normalize_name(column_name),
                description=description,
                declared_type=declared_type,
                dtype=mapped_dtype,
                null_ratio=0.0,
                unique_ratio=0.0,
                avg_length=0.0,
                non_null_count=0,
                sample_values=sample_values,
                distinct_sample_values=list(sample_values),
                detected_patterns=detected_patterns,
                tokenized_name=tokenize_name(column_name),
            )
        )

    if not columns:
        raise ValueError("Spec upload did not produce any usable fields")

    return SchemaProfile(
        dataset_id=dataset_id,
        dataset_name=dataset_name,
        row_count=0,
        columns=columns,
    )


def resolve_spec_layout(
    rows: list[dict[str, object]],
    *,
    name_col: str | None,
    description_col: str | None,
    type_col: str | None,
    sample_values_col: str | None,
) -> SpecLayoutHint:
    """Resolve explicit or detected schema-spec columns into one validated layout hint."""

    headers = list(rows[0].keys())
    available_headers = {normalize_tabular_header(header): header for header in headers}

    if name_col:
        resolved_name_col = available_headers.get(normalize_tabular_header(name_col))
        if not resolved_name_col:
            raise ValueError(f"Unknown spec name column: {name_col}")
        resolved_description_col = resolve_optional_header(description_col, available_headers, "description")
        resolved_type_col = resolve_optional_header(type_col, available_headers, "type")
        resolved_sample_values_col = resolve_optional_header(sample_values_col, available_headers, "sample values")
        confidence = 1.0 if resolved_description_col and resolved_type_col else 0.8 if (resolved_description_col or resolved_type_col) else 0.6
        return SpecLayoutHint(
            name_col=resolved_name_col,
            description_col=resolved_description_col,
            type_col=resolved_type_col,
            sample_values_col=resolved_sample_values_col,
            confidence=confidence,
        )

    detected = build_spec_layout_hint(rows)
    if not detected:
        raise ValueError("Could not detect a schema-spec layout. Provide name_col explicitly or upload a row-data file via /upload.")
    return detected


def _rows_support_requested_spec_layout(
    rows: list[dict[str, object]],
    *,
    name_col: str | None,
    description_col: str | None,
    type_col: str | None,
    sample_values_col: str | None,
) -> bool:
    if not rows:
        return False

    headers = {normalize_tabular_header(header) for header in rows[0].keys()}
    if name_col:
        return normalize_tabular_header(name_col) in headers

    if build_spec_layout_hint(rows) is not None:
        return True

    optional_headers = [description_col, type_col, sample_values_col]
    return any(normalize_tabular_header(value) in headers for value in optional_headers if value)


def resolve_optional_header(
    requested_header: str | None,
    available_headers: dict[str, str],
    header_kind: str,
) -> str | None:
    """Resolve one optional user-specified header against normalized available headers."""

    if not requested_header:
        return None
    resolved_header = available_headers.get(normalize_tabular_header(requested_header))
    if not resolved_header:
        raise ValueError(f"Unknown spec {header_kind} column: {requested_header}")
    return resolved_header


def find_matching_header(
    headers: list[str],
    normalized_headers: list[str],
    candidates: set[str],
) -> str | None:
    """Return the first header whose normalized form matches one of the candidate names."""

    for header, normalized in zip(headers, normalized_headers, strict=False):
        if normalized in candidates:
            return header
    return None


def _extract_embedded_spec_rows(
    payload: bytes,
    filename: str,
    *,
    requested_header_row_index: int | None = None,
) -> tuple[list[dict[str, object]], int] | None:
    candidates = list_embedded_spec_candidates(
        payload,
        filename,
        requested_header_row_index=requested_header_row_index,
    )
    if not candidates:
        return None
    return candidates[0]


def list_embedded_spec_candidates(
    payload: bytes,
    filename: str,
    *,
    requested_header_row_index: int | None = None,
) -> list[tuple[list[dict[str, object]], int]]:
    suffix = Path(filename).suffix.lower()
    if suffix == ".csv":
        return _list_embedded_csv_spec_candidates(
            payload,
            requested_header_row_index=requested_header_row_index,
        )
    if suffix == ".xlsx":
        return _list_embedded_xlsx_spec_candidates(
            payload,
            requested_header_row_index=requested_header_row_index,
        )
    return []


def _list_embedded_csv_spec_candidates(
    payload: bytes,
    *,
    requested_header_row_index: int | None = None,
) -> list[tuple[list[dict[str, object]], int]]:

    decoded = decode_text_payload(payload)
    lines = decoded.splitlines()
    if not lines:
        return []

    candidate_starts = [requested_header_row_index - 1] if requested_header_row_index else list(range(len(lines)))
    candidates: list[tuple[list[dict[str, object]], int]] = []
    seen_signatures: set[tuple[int, tuple[str, ...]]] = set()

    for delimiter in EMBEDDED_SPEC_DELIMITERS:
        for start_index in candidate_starts:
            candidate = _parse_embedded_csv_block(lines, start_index, delimiter)
            if candidate is None:
                continue
            rows, header_row_index = candidate
            signature = (header_row_index, tuple(str(header) for header in rows[0].keys()))
            if signature in seen_signatures:
                continue
            seen_signatures.add(signature)
            candidates.append((rows, header_row_index))

    candidates.sort(key=lambda item: (-_score_embedded_spec_candidate(item[0]), item[1]))
    return candidates


def _list_embedded_xlsx_spec_candidates(
    payload: bytes,
    *,
    requested_header_row_index: int | None = None,
) -> list[tuple[list[dict[str, object]], int]]:
    workbook = load_workbook(filename=BytesIO(payload), read_only=True, data_only=True)
    try:
        sheet = workbook.active
        raw_rows = [tuple(row) for row in sheet.iter_rows(values_only=True)]
    finally:
        workbook.close()

    if not raw_rows:
        return []

    candidate_starts = [requested_header_row_index - 1] if requested_header_row_index else list(range(len(raw_rows)))
    candidates: list[tuple[list[dict[str, object]], int]] = []
    seen_signatures: set[tuple[int, tuple[str, ...]]] = set()

    for start_index in candidate_starts:
        candidate = _parse_embedded_row_block(raw_rows, start_index)
        if candidate is None:
            continue
        rows, header_row_index = candidate
        signature = (header_row_index, tuple(str(header) for header in rows[0].keys()))
        if signature in seen_signatures:
            continue
        seen_signatures.add(signature)
        candidates.append((rows, header_row_index))

    candidates.sort(key=lambda item: (-_score_embedded_spec_candidate(item[0]), item[1]))
    return candidates


def _parse_embedded_csv_block(
    lines: list[str],
    start_index: int,
    delimiter: str,
) -> tuple[list[dict[str, object]], int] | None:
    if start_index < 0 or start_index >= len(lines):
        return None

    header_cells = _parse_csv_cells(lines[start_index], delimiter)
    if not _looks_like_embedded_header_row(header_cells):
        return None
    if start_index > 0:
        previous_cells = _parse_csv_cells(lines[start_index - 1], delimiter)
        if _looks_like_embedded_data_row(previous_cells, len(header_cells)):
            return None

    data_rows: list[list[str]] = []
    expected_columns = len(header_cells)
    for line in lines[start_index + 1 :]:
        if not line.strip():
            break
        row_cells = _parse_csv_cells(line, delimiter)
        if not _looks_like_embedded_data_row(row_cells, expected_columns):
            break
        data_rows.append(row_cells)

    if not data_rows:
        return None

    headers = [str(cell).strip() for cell in header_cells]
    try:
        records = [
            {header: value for header, value in zip(headers, row_cells, strict=False)}
            for row_cells in data_rows
        ]
        rows = normalize_rows(records, header_order=headers)
    except ValueError:
        return None

    return rows, start_index + 1


def _parse_embedded_row_block(
    raw_rows: list[tuple[object, ...]],
    start_index: int,
) -> tuple[list[dict[str, object]], int] | None:
    if start_index < 0 or start_index >= len(raw_rows):
        return None

    header_cells = _normalize_row_cells(raw_rows[start_index])
    if not _looks_like_embedded_header_row(header_cells):
        return None
    if start_index > 0:
        previous_cells = _normalize_row_cells(raw_rows[start_index - 1], expected_columns=len(header_cells))
        if _looks_like_embedded_data_row(previous_cells, len(header_cells)):
            return None

    data_rows: list[list[str]] = []
    expected_columns = len(header_cells)
    for row in raw_rows[start_index + 1 :]:
        row_cells = _normalize_row_cells(row, expected_columns=expected_columns)
        if not any(cell.strip() for cell in row_cells):
            break
        if not _looks_like_embedded_data_row(row_cells, expected_columns):
            break
        data_rows.append(row_cells)

    if not data_rows:
        return None

    headers = [str(cell).strip() for cell in header_cells]
    try:
        records = [
            {header: value for header, value in zip(headers, row_cells, strict=False)}
            for row_cells in data_rows
        ]
        rows = normalize_rows(records, header_order=headers)
    except ValueError:
        return None

    return rows, start_index + 1


def _parse_csv_cells(line: str, delimiter: str) -> list[str]:
    return ["" if value is None else str(value).strip() for value in next(csv.reader([line], delimiter=delimiter))]


def _normalize_row_cells(row: tuple[object, ...], *, expected_columns: int | None = None) -> list[str]:
    cells = ["" if value is None else str(value).strip() for value in row]
    if expected_columns is not None:
        if len(cells) < expected_columns:
            cells = cells + [""] * (expected_columns - len(cells))
        else:
            cells = cells[:expected_columns]
    while cells and not cells[-1]:
        cells.pop()
    return cells


def _looks_like_embedded_header_row(cells: list[str]) -> bool:
    if len(cells) < EMBEDDED_SPEC_MIN_COLUMNS or len(cells) > SPEC_LAYOUT_MAX_COLUMNS:
        return False
    if any(not cell for cell in cells):
        return False
    header_like_count = sum(1 for cell in cells if _looks_like_embedded_header_cell(cell))
    return header_like_count >= min(2, len(cells))


def _looks_like_embedded_header_cell(value: str) -> bool:
    normalized = normalize_header_candidate(value)
    if not normalized:
        return False
    if len(value.strip()) > EMBEDDED_SPEC_MAX_HEADER_LENGTH:
        return False
    return len(normalized.split()) <= EMBEDDED_SPEC_MAX_HEADER_TOKENS


def _looks_like_embedded_data_row(cells: list[str], expected_columns: int) -> bool:
    if len(cells) != expected_columns:
        return False
    return any(cell for cell in cells)


def _score_embedded_spec_candidate(rows: list[dict[str, object]]) -> int:
    headers = list(rows[0].keys()) if rows else []
    score = len(rows) * max(1, len(headers))
    if headers and detect_spec_layout(headers) is not None:
        score += 100
    return score


def normalize_header_candidate(header: str) -> str:
    """Normalize a raw header into the comparable form used for schema-spec detection."""

    normalized = normalize_tabular_header(header).strip().lower().replace("_", " ")
    return re.sub(r"[^a-z0-9]+", " ", normalized).strip()


def is_plausible_field_name(value: str, header_name: str) -> bool:
    """Return whether a cell value looks like a real field name instead of another header."""

    stripped = value.strip()
    if not stripped:
        return False

    normalized_value = normalize_header_candidate(stripped)
    normalized_header = normalize_header_candidate(header_name)
    if normalized_value == normalized_header:
        return False
    if normalized_value in NAME_CANDIDATES | DESCRIPTION_CANDIDATES | TYPE_CANDIDATES:
        return False
    if normalized_value.startswith("table ") or normalized_value.startswith("view "):
        return False
    return True


def map_spec_type(raw_type: object) -> str:
    """Map raw schema-spec type text into Semantra's simplified dtype labels."""

    lowered = "" if is_nullish(raw_type) else str(raw_type).strip().lower()
    if any(token in lowered for token in ("datetime", "timestamp", "dttm")):
        return "datetime"
    if any(token in lowered for token in ("date", "dats")):
        return "date"
    if any(token in lowered for token in ("decimal", "float", "double", "real", "curr", "quan")):
        return "float"
    if any(token in lowered for token in ("bigint", "smallint", "integer", " int", "int(", "number", "numeric", "numc")):
        return "integer"
    if any(token in lowered for token in ("bool", "boolean", "check")):
        return "bool"
    return "string"


def map_spec_pattern(dtype: str) -> str:
    """Map a simplified dtype into the default detected pattern used for spec-only uploads."""

    if dtype in {"date", "datetime"}:
        return "date"
    if dtype == "float":
        return "float"
    if dtype == "integer":
        return "integer"
    return "text"


def parse_spec_sample_values(raw_value: object) -> list[str]:
    """Parse sample values from a schema-spec cell into a bounded list of strings."""

    if is_nullish(raw_value):
        return []

    text = str(raw_value).strip()
    if not text:
        return []

    parts = re.split(r"\s*(?:\||;|\r?\n)\s*", text)
    values: list[str] = []
    seen: set[str] = set()
    for part in parts:
        cleaned = part.strip().strip(",")
        if not cleaned:
            continue
        dedupe_key = cleaned.lower()
        if dedupe_key in seen:
            continue
        seen.add(dedupe_key)
        values.append(cleaned)
        if len(values) >= MAX_SPEC_SAMPLE_VALUES:
            break
    return values