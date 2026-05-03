from __future__ import annotations

import csv
import json
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any
from xml.etree import ElementTree

from openpyxl import load_workbook

from app.utils.tabular import decode_text_payload, is_nullish, normalize_tabular_cell, normalize_tabular_header


SUPPORTED_ROW_FORMATS = (".csv", ".json", ".xml", ".xlsx")


def decode_tabular_payload(payload: bytes) -> str:
    return decode_text_payload(payload)


def parse_tabular_payload(payload: bytes, filename: str) -> list[dict[str, Any]]:
    suffix = Path(filename).suffix.lower()
    if suffix == ".csv":
        return read_csv_payload(payload)
    if suffix == ".json":
        return read_json_payload(payload)
    if suffix == ".xml":
        return read_xml_payload(payload)
    if suffix == ".xlsx":
        return read_xlsx_payload(payload)
    raise ValueError(f"Unsupported row-based upload format: {suffix or 'unknown'}")


def read_csv_payload(payload: bytes) -> list[dict[str, Any]]:
    decoded = decode_tabular_payload(payload)
    reader = csv.DictReader(StringIO(decoded))
    rows = [dict(row) for row in reader]
    if not reader.fieldnames:
        raise ValueError("CSV file has no header row")
    return normalize_rows(rows, header_order=[str(field) for field in reader.fieldnames])


def read_json_payload(payload: bytes) -> list[dict[str, Any]]:
    decoded = decode_tabular_payload(payload)
    parsed = json.loads(decoded)
    records = coerce_json_records(parsed)
    return normalize_rows(records)


def read_xml_payload(payload: bytes) -> list[dict[str, Any]]:
    decoded = decode_tabular_payload(payload)
    root = ElementTree.fromstring(decoded)
    records = coerce_xml_records(root)
    return normalize_rows(records)


def read_xlsx_payload(payload: bytes) -> list[dict[str, Any]]:
    workbook = load_workbook(filename=BytesIO(payload), read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
    finally:
        workbook.close()

    if not rows:
        raise ValueError("XLSX workbook has no rows")

    try:
        headers = [normalize_tabular_header(value) for value in rows[0]]
    except ValueError as error:
        raw_headers = ["" if value is None else str(value).strip() for value in rows[0]]
        if not any(raw_headers):
            raise ValueError("XLSX file has no header row") from error
        raise ValueError("XLSX header row contains empty column names") from error

    records: list[dict[str, Any]] = []
    for row_values in rows[1:]:
        padded = list(row_values) + [None] * max(0, len(headers) - len(row_values))
        if all(is_nullish(value) for value in padded[: len(headers)]):
            continue
        record = {header: normalize_tabular_cell(value) for header, value in zip(headers, padded, strict=False)}
        records.append(record)
    return normalize_rows(records, header_order=headers)


def coerce_json_records(payload: Any) -> list[dict[str, Any]]:
    if isinstance(payload, list):
        if not all(isinstance(item, dict) for item in payload):
            raise ValueError("JSON uploads must be an array of objects")
        return payload

    if isinstance(payload, dict):
        list_candidates = [
            value for value in payload.values() if isinstance(value, list) and all(isinstance(item, dict) for item in value)
        ]
        if len(list_candidates) == 1:
            return list_candidates[0]
        if all(not isinstance(value, (list, dict)) for value in payload.values()):
            return [payload]

    raise ValueError("JSON uploads must be an object, an array of objects, or an object containing one array of objects")


def coerce_xml_records(root: ElementTree.Element) -> list[dict[str, Any]]:
    children = list(root)
    if not children:
        raise ValueError("XML upload must contain at least one record element")

    if all(not list(child) for child in children):
        return [element_to_record(root)]

    if all(list(child) for child in children):
        return [element_to_record(child) for child in children]

    if len(children) == 1 and list(children[0]):
        nested_children = list(children[0])
        if nested_children and all(list(child) for child in nested_children):
            return [element_to_record(child) for child in nested_children]

    raise ValueError(
        "XML uploads must be a single record element or a collection of repeated record elements with child fields"
    )


def element_to_record(element: ElementTree.Element) -> dict[str, Any]:
    record: dict[str, Any] = {}
    for child in list(element):
        key = normalize_tabular_header(child.tag)
        if list(child):
            record[key] = normalize_tabular_cell(element_to_value(child))
        else:
            record[key] = normalize_tabular_cell(child.text)
    if not record:
        raise ValueError(f"XML record '{element.tag}' did not contain any fields")
    return record


def element_to_value(element: ElementTree.Element) -> Any:
    children = list(element)
    if not children:
        return element.text or ""
    if all(not list(child) for child in children):
        return {child.tag: child.text or "" for child in children}
    return {child.tag: element_to_value(child) for child in children}


def normalize_rows(rows: list[dict[str, Any]], header_order: list[str] | None = None) -> list[dict[str, Any]]:
    normalized_header_order: list[str] = []
    seen: set[str] = set()

    for header in header_order or []:
        normalized = normalize_tabular_header(header)
        if normalized not in seen:
            seen.add(normalized)
            normalized_header_order.append(normalized)

    normalized_records: list[dict[str, Any]] = []
    for row in rows:
        if not isinstance(row, dict):
            raise ValueError("Tabular uploads must yield object records")
        normalized_row = {normalize_tabular_header(key): normalize_tabular_cell(value) for key, value in row.items()}
        for key in normalized_row:
            if key not in seen:
                seen.add(key)
                normalized_header_order.append(key)
        normalized_records.append(normalized_row)

    if not normalized_records:
        return []

    return [
        {header: record.get(header, "") for header in normalized_header_order}
        for record in normalized_records
    ]
