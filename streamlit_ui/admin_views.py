"""Admin and Canonical Console UI surfaces for observability and stewardship."""

from __future__ import annotations

import csv
from datetime import datetime
from io import BytesIO, StringIO
from collections.abc import Callable
from typing import Any

import httpx
import streamlit as st
from openpyxl import load_workbook

from streamlit_ui.shared_views import render_status_badge_legend


GOVERNANCE_SECTIONS = ("Canonical", "Knowledge", "Overlays & Runtime", "Stewardship")


def _normalized_text(value: object) -> str:
    return str(value or "").strip()


def _section_label(title: str, detail: str | None = None) -> str:
    detail_text = str(detail or "").strip()
    return f"{title} · {detail_text}" if detail_text else title


def resolve_active_governance_section(session_state: dict) -> str:
    preferred = str(session_state.pop("pending_governance_section", "") or "").strip()
    current = str(session_state.get("active_governance_section", "") or "").strip()
    if preferred in GOVERNANCE_SECTIONS:
        session_state["active_governance_section"] = preferred
    elif current not in GOVERNANCE_SECTIONS:
        session_state["active_governance_section"] = GOVERNANCE_SECTIONS[0]
    return str(session_state.get("active_governance_section") or GOVERNANCE_SECTIONS[0])


def _apply_pending_governance_selectbox_value(
    session_state: dict,
    *,
    pending_key: str,
    target_key: str,
    options: list[str],
) -> bool:
    pending_value = _normalized_text(session_state.get(pending_key))
    if not pending_value or not options:
        return False
    matched = pending_value in {_normalized_text(option) for option in options if _normalized_text(option)}
    if matched:
        session_state[target_key] = pending_value
    session_state.pop(pending_key, None)
    return matched


def _pending_canonical_concept_label(concepts: list[dict] | None, session_state: dict) -> str:
    pending_concept_id = _normalized_text(session_state.get("pending_governance_canonical_concept_id"))
    if not pending_concept_id:
        return ""
    for concept in concepts or []:
        if _normalized_text(concept.get("concept_id")) == pending_concept_id:
            session_state.pop("pending_governance_canonical_concept_id", None)
            return _canonical_concept_option_label(concept)
    return ""


def _governance_focus_sources(session_state: dict) -> list[str]:
    focus_sources: list[str] = []
    for value in session_state.get("governance_focus_sources") or []:
        source = _normalized_text(value)
        if source and source not in focus_sources:
            focus_sources.append(source)
    return focus_sources


def _governance_focus_source_caption(source_filter: str | None, focus_sources: list[str] | None) -> str:
    if _normalized_text(source_filter):
        return ""
    sources = [value for value in focus_sources or [] if _normalized_text(value)]
    if len(sources) == 1:
        return f"Catalog handoff focus keeps Stewardship scoped to source field {sources[0]} while the source filter remains All."
    if sources:
        return f"Catalog handoff focus keeps Stewardship scoped to {len(sources)} unmatched source fields while the source filter remains All."
    return ""


def _filter_canonical_concepts(concepts: list[dict] | None, query: str | None = None) -> list[dict]:
    items = concepts or []
    normalized_query = _normalized_text(query).lower()
    if not normalized_query:
        return list(items)

    filtered: list[dict] = []
    for concept in items:
        haystacks = [
            _normalized_text(concept.get("concept_id")),
            _normalized_text(concept.get("display_name")),
            _normalized_text(concept.get("description")),
            _normalized_text(concept.get("entity")),
            _normalized_text(concept.get("attribute")),
            _normalized_text(concept.get("source")),
            *(str(value or "") for value in concept.get("source_systems", [])),
            *(str(value or "") for value in concept.get("business_domains", [])),
            *(str(alias or "") for alias in concept.get("base_aliases", [])),
            *(str(alias or "") for alias in concept.get("active_overlay_aliases", [])),
        ]
        if any(normalized_query in value.lower() for value in haystacks if value):
            filtered.append(concept)
    return filtered


def _filter_knowledge_concepts(concepts: list[dict] | None, query: str | None = None) -> list[dict]:
    items = concepts or []
    normalized_query = _normalized_text(query).lower()
    if not normalized_query:
        return list(items)

    filtered: list[dict] = []
    for concept in items:
        haystacks = [
            _normalized_text(concept.get("concept_id")),
            _normalized_text(concept.get("canonical_name")),
            _normalized_text(concept.get("domain")),
            _normalized_text(concept.get("source")),
            *(str(value or "") for value in concept.get("source_systems", [])),
            *(str(value or "") for value in concept.get("linked_canonical_concepts", [])),
            *(str(alias or "") for alias in concept.get("aliases", [])),
        ]
        if any(normalized_query in value.lower() for value in haystacks if value):
            filtered.append(concept)
    return filtered


def _filter_knowledge_concepts_by_scope(
    concepts: list[dict] | None,
    source_system: str | None = None,
    source: str | None = None,
) -> list[dict]:
    items = concepts or []
    normalized_source_system = _normalized_text(source_system)
    normalized_source = _normalized_text(source)
    if not normalized_source_system and not normalized_source:
        return list(items)

    filtered: list[dict] = []
    for concept in items:
        concept_source_systems = {
            _normalized_text(value)
            for value in concept.get("source_systems", [])
            if _normalized_text(value)
        }
        concept_source = _normalized_text(concept.get("source"))
        if normalized_source_system and normalized_source_system not in concept_source_systems:
            continue
        if normalized_source and concept_source != normalized_source:
            continue
        filtered.append(concept)
    return filtered


def _filter_knowledge_concepts_by_focus(concepts: list[dict] | None, focus: str | None = None) -> list[dict]:
    items = concepts or []
    normalized_focus = _normalized_text(focus).lower() or "all"
    if normalized_focus == "all":
        return list(items)

    filtered: list[dict] = []
    for concept in items:
        source = _normalized_text(concept.get("source")).lower()
        editable = bool(concept.get("editable"))
        field_context_count = int(concept.get("field_context_count", 0) or 0)
        linked_count = int(concept.get("linked_canonical_concept_count", 0) or 0)
        if normalized_focus == "editable" and editable:
            filtered.append(concept)
        elif normalized_focus == "derived" and source == "derived_runtime":
            filtered.append(concept)
        elif normalized_focus == "generated" and source == "generated_runtime":
            filtered.append(concept)
        elif normalized_focus == "linked" and linked_count > 0:
            filtered.append(concept)
        elif normalized_focus == "with_context" and field_context_count > 0:
            filtered.append(concept)
    return filtered


def _filter_canonical_concepts_by_scope(
    concepts: list[dict] | None,
    source_system: str | None = None,
    business_domain: str | None = None,
) -> list[dict]:
    items = concepts or []
    normalized_source_system = _normalized_text(source_system)
    normalized_business_domain = _normalized_text(business_domain)
    if not normalized_source_system and not normalized_business_domain:
        return list(items)

    filtered: list[dict] = []
    for concept in items:
        concept_source_systems = {
            _normalized_text(value)
            for value in concept.get("source_systems", [])
            if _normalized_text(value)
        }
        concept_business_domains = {
            _normalized_text(value)
            for value in concept.get("business_domains", [])
            if _normalized_text(value)
        }
        if normalized_source_system and normalized_source_system not in concept_source_systems:
            continue
        if normalized_business_domain and normalized_business_domain not in concept_business_domains:
            continue
        filtered.append(concept)
    return filtered


def _filter_canonical_concepts_by_focus(concepts: list[dict] | None, focus: str | None = None) -> list[dict]:
    items = concepts or []
    normalized_focus = _normalized_text(focus).lower() or "all"
    if normalized_focus == "all":
        return list(items)

    filtered: list[dict] = []
    for concept in items:
        source = _normalized_text(concept.get("source")).lower() or "base"
        usage_count = int(concept.get("usage_count", 0) or 0)
        field_context_count = int(concept.get("field_context_count", 0) or 0)
        overlay_count = int(concept.get("active_overlay_entry_count", 0) or 0)
        if normalized_focus == "active_overlay" and overlay_count > 0:
            filtered.append(concept)
        elif normalized_focus == "overlay_only" and source == "overlay_only":
            filtered.append(concept)
        elif normalized_focus == "in_use" and usage_count > 0:
            filtered.append(concept)
        elif normalized_focus == "with_context" and field_context_count > 0:
            filtered.append(concept)
        elif normalized_focus == "base_only" and source == "base":
            filtered.append(concept)
    return filtered


def _canonical_overlay_summary(runtime: dict | None, overlays: list[dict] | None) -> dict[str, object]:
    runtime_payload = runtime or {}
    overlay_rows = overlays or []
    entry_type_counts = runtime_payload.get("entry_type_counts") or {}
    status_counts: dict[str, int] = {}
    for overlay in overlay_rows:
        status = _normalized_text(overlay.get("status")).lower() or "unknown"
        status_counts[status] = status_counts.get(status, 0) + 1
    return {
        "mode": runtime_payload.get("mode") or "base_only",
        "runtime_source": runtime_payload.get("runtime_source") or "unknown",
        "source_hash_state": runtime_payload.get("source_hash_state") or "missing",
        "active_overlay_name": runtime_payload.get("active_overlay_name") or "none",
        "active_entry_count": int(runtime_payload.get("active_entry_count", 0) or 0),
        "concept_alias_entries": int(entry_type_counts.get("concept_alias", 0) or 0),
        "total_versions": len(overlay_rows),
        "active_versions": int(status_counts.get("active", 0) or 0),
        "validated_versions": int(status_counts.get("validated", 0) or 0),
        "archived_versions": int(status_counts.get("archived", 0) or 0),
    }


def _overlay_activation_block_reason(overlay: dict | None) -> str:
    status = _normalized_text((overlay or {}).get("status")).lower() or "draft"
    if status == "validated":
        return ""
    if status == "active":
        return "This knowledge overlay is already active."
    return f"Only validated knowledge overlays can be activated. Current status: {status}."


def _overlay_archive_block_reason(overlay: dict | None) -> str:
    status = _normalized_text((overlay or {}).get("status")).lower() or "draft"
    if status in {"validated", "active"}:
        return ""
    if status == "archived":
        return "This knowledge overlay is already archived."
    return f"Only validated or active knowledge overlays can be archived. Current status: {status}."


def _canonical_concept_registry_rows(concepts: list[dict] | None) -> list[dict]:
    rows: list[dict] = []
    for concept in concepts or []:
        base_aliases = [str(alias).strip() for alias in concept.get("base_aliases", []) if str(alias).strip()]
        overlay_aliases = [str(alias).strip() for alias in concept.get("active_overlay_aliases", []) if str(alias).strip()]
        privacy = concept.get("privacy") or {}
        rows.append(
            {
                "concept_id": concept.get("concept_id"),
                "display_name": concept.get("display_name"),
                "entity": concept.get("entity") or "",
                "attribute": concept.get("attribute") or "",
                "data_type": concept.get("data_type") or "",
                "pii": "yes" if privacy.get("is_pii") else "no",
                "gdpr_special": "yes" if privacy.get("is_gdpr_special_category") else "no",
                "pii_tags": ", ".join(str(value).strip() for value in privacy.get("pii_categories", []) if str(value).strip()),
                "data_subjects": ", ".join(
                    str(value).strip() for value in privacy.get("data_subject_types", []) if str(value).strip()
                ),
                "source": concept.get("source") or "base",
                "usage_count": concept.get("usage_count", 0),
                "field_context_count": concept.get("field_context_count", 0),
                "active_overlay_entry_count": concept.get("active_overlay_entry_count", 0),
                "source_systems": ", ".join(str(value).strip() for value in concept.get("source_systems", []) if str(value).strip()),
                "business_domains": ", ".join(str(value).strip() for value in concept.get("business_domains", []) if str(value).strip()),
                "base_aliases": ", ".join(base_aliases),
                "active_overlay_aliases": ", ".join(overlay_aliases),
            }
        )
    return rows


def _knowledge_concept_registry_rows(concepts: list[dict] | None) -> list[dict]:
    rows: list[dict] = []
    for concept in concepts or []:
        linked_privacy = concept.get("linked_privacy") or {}
        rows.append(
            {
                "concept_id": concept.get("concept_id"),
                "canonical_name": concept.get("canonical_name") or "",
                "domain": concept.get("domain") or "",
                "source": concept.get("source") or "derived_runtime",
                "editable": "yes" if concept.get("editable") else "no",
                "linked_pii": "yes" if linked_privacy.get("is_pii") else "no",
                "linked_gdpr_special": "yes" if linked_privacy.get("is_gdpr_special_category") else "no",
                "linked_pii_tags": ", ".join(
                    str(value).strip() for value in linked_privacy.get("pii_categories", []) if str(value).strip()
                ),
                "linked_data_subjects": ", ".join(
                    str(value).strip() for value in linked_privacy.get("data_subject_types", []) if str(value).strip()
                ),
                "alias_count": concept.get("alias_count", 0),
                "field_context_count": concept.get("field_context_count", 0),
                "linked_canonical_concept_count": concept.get("linked_canonical_concept_count", 0),
                "source_systems": ", ".join(str(value).strip() for value in concept.get("source_systems", []) if str(value).strip()),
                "linked_canonical_concepts": ", ".join(str(value).strip() for value in concept.get("linked_canonical_concepts", []) if str(value).strip()),
                "aliases": ", ".join(str(value).strip() for value in concept.get("aliases", []) if str(value).strip()),
            }
        )
    return rows


def _canonical_concept_option_label(concept: dict) -> str:
    concept_id = _normalized_text(concept.get("concept_id")) or "unknown"
    display_name = _normalized_text(concept.get("display_name")) or concept_id
    source = _normalized_text(concept.get("source")) or "base"
    usage_count = int(concept.get("usage_count", 0) or 0)
    return f"{concept_id} | {display_name} | source={source} | usage={usage_count}"


def _knowledge_concept_option_label(concept: dict) -> str:
    concept_id = _normalized_text(concept.get("concept_id")) or "unknown"
    canonical_name = _normalized_text(concept.get("canonical_name")) or concept_id
    source = _normalized_text(concept.get("source")) or "derived_runtime"
    linked_count = int(concept.get("linked_canonical_concept_count", 0) or 0)
    return f"{concept_id} | {canonical_name} | source={source} | linked={linked_count}"


def _single_overlay_patch_bytes(
    *,
    entry_type: str,
    canonical_term: str,
    alias: str,
    domain: str | None = None,
    source_system: str | None = None,
    note: str | None = None,
) -> bytes:
    buffer = StringIO()
    writer = csv.DictWriter(
        buffer,
        fieldnames=["entry_type", "canonical_term", "alias", "domain", "source_system", "note"],
    )
    writer.writeheader()
    writer.writerow(
        {
            "entry_type": entry_type,
            "canonical_term": canonical_term,
            "alias": alias,
            "domain": domain or "",
            "source_system": source_system or "",
            "note": note or "",
        }
    )
    return buffer.getvalue().encode("utf-8")


def _context_alias_token(object_name: str, field_name: str) -> str:
    object_token = _normalized_text(object_name)
    field_token = _normalized_text(field_name)
    if object_token and field_token:
        return f"{object_token}.{field_token}"
    return field_token or object_token


def _context_note_payload(
    *,
    system: str,
    object_name: str,
    field_name: str,
    object_description: str,
    field_description: str,
    note: str,
) -> str:
    parts = [
        "context_patch=true",
        f"system={_normalized_text(system)}",
        f"object={_normalized_text(object_name)}",
        f"field={_normalized_text(field_name)}",
    ]
    if _normalized_text(object_description):
        parts.append(f"object_description={_normalized_text(object_description)}")
    if _normalized_text(field_description):
        parts.append(f"field_description={_normalized_text(field_description)}")
    if _normalized_text(note):
        parts.append(f"note={_normalized_text(note)}")
    return "; ".join(parts)


def _parse_semicolon_key_values(value: object) -> dict[str, str]:
    parsed: dict[str, str] = {}
    for segment in str(value or "").split(";"):
        token = segment.strip()
        if not token or "=" not in token:
            continue
        key, raw_value = token.split("=", 1)
        normalized_key = _normalized_text(key).lower()
        if not normalized_key:
            continue
        parsed[normalized_key] = _normalized_text(raw_value)
    return parsed


def _context_patch_ingest_rows(canonical_concept_detail: dict | None = None) -> list[dict]:
    detail = canonical_concept_detail or {}
    contexts = detail.get("field_contexts") or []
    entries = detail.get("active_overlay_entries") or []

    context_index = {
        (
            _normalized_text(context.get("system")).lower(),
            _normalized_text(context.get("object_name")).lower(),
            _normalized_text(context.get("field_name")).lower(),
        )
        for context in contexts
    }

    rows: list[dict] = []
    for entry in entries:
        payload = _parse_semicolon_key_values(entry.get("note"))
        if _normalized_text(payload.get("context_patch")).lower() not in {"true", "1", "yes"}:
            continue

        alias = _normalized_text(entry.get("alias"))
        object_name = _normalized_text(payload.get("object"))
        field_name = _normalized_text(payload.get("field"))
        if (not object_name or not field_name) and "." in alias:
            object_part, field_part = alias.split(".", 1)
            object_name = object_name or _normalized_text(object_part)
            field_name = field_name or _normalized_text(field_part)

        system = _normalized_text(payload.get("system")) or _normalized_text(entry.get("source_system"))
        key_exact = (_normalized_text(system).lower(), _normalized_text(object_name).lower(), _normalized_text(field_name).lower())
        key_alias_fallback = (_normalized_text(system).lower(), _normalized_text(object_name).lower(), alias.lower())
        ingested = key_exact in context_index or key_alias_fallback in context_index

        rows.append(
            {
                "entry_id": entry.get("entry_id"),
                "alias": alias,
                "system": system,
                "object": object_name,
                "field": field_name,
                "ingested": "yes" if ingested else "no",
            }
        )
    return rows


def _canonical_gap_candidate_key(index: int, candidate: dict) -> str:
    source = _normalized_text(candidate.get("source"))
    target = _normalized_text(candidate.get("target"))
    return f"canonical_gap_{source or 'unknown'}_{target or 'unknown'}".replace(" ", "_")


def _canonical_gap_console_state(
    candidate_key: str,
    console_states: dict[str, str] | None = None,
    stewardship_item: dict | None = None,
) -> str:
    stewardship_state = _canonical_gap_stewardship_status(stewardship_item)
    if stewardship_state in {"ignored", "approved", "rejected"}:
        return stewardship_state
    if stewardship_state in {"new", "needs_review", "ready_for_approval"}:
        return "active"
    state = _normalized_text((console_states or {}).get(candidate_key)) or "active"
    if state not in {"active", "ignored", "approved", "rejected"}:
        return "active"
    return state


def _canonical_gap_stewardship_status(stewardship_item: dict | None = None) -> str:
    state = _normalized_text((stewardship_item or {}).get("status"))
    if state not in {"new", "needs_review", "ready_for_approval", "approved", "rejected", "ignored", "promoted"}:
        return ""
    return state


def _canonical_gap_proposal_state(
    candidate_key: str,
    proposal_states: dict[str, str] | None = None,
    stewardship_item: dict | None = None,
) -> str:
    stewardship_state = _canonical_gap_stewardship_status(stewardship_item)
    if stewardship_state in {"new", "needs_review", "ready_for_approval"}:
        return stewardship_state
    state = _normalized_text((proposal_states or {}).get(candidate_key)) or "new"
    if state not in {"new", "needs_review", "ready_for_approval"}:
        return "new"
    return state


def _canonical_gap_proposal_state_label(state: str) -> str:
    return {
        "new": "New",
        "needs_review": "Needs review",
        "ready_for_approval": "Ready for approval",
        "approved": "Approved",
        "rejected": "Rejected",
        "ignored": "Ignored",
        "promoted": "Promoted",
        "active": "Active",
        "not_tracked": "Not tracked",
    }.get(state, state)


def _canonical_gap_proposal_state_map(records: list[dict] | None = None) -> dict[str, str]:
    states: dict[str, str] = {}
    for record in records or []:
        candidate_key = _normalized_text(record.get("candidate_key") or record.get("item_key"))
        if not candidate_key:
            continue
        proposal_state = _normalized_text(record.get("proposal_state") or record.get("status"))
        if proposal_state not in {"new", "needs_review", "ready_for_approval"}:
            continue
        states[candidate_key] = _canonical_gap_proposal_state(candidate_key, {candidate_key: proposal_state})
    return states


def _canonical_gap_stewardship_item_map(records: list[dict] | None = None) -> dict[str, dict]:
    items: dict[str, dict] = {}
    for record in records or []:
        item_key = _normalized_text(record.get("item_key") or record.get("candidate_key"))
        item_type = _normalized_text(record.get("item_type")) or "canonical_gap"
        if not item_key or item_type != "canonical_gap":
            continue
        items[item_key] = record
    return items


def _concept_governance_item_key(concept_id: str | None) -> str:
    normalized_concept_id = _normalized_text(concept_id) or "unknown"
    return f"concept_governance_{normalized_concept_id}".replace(" ", "_")


def _concept_governance_item_map(records: list[dict] | None = None) -> dict[str, dict]:
    items: dict[str, dict] = {}
    for record in records or []:
        if _normalized_text(record.get("item_type")) != "concept_governance":
            continue
        concept_id = _normalized_text(record.get("concept_id") or record.get("target"))
        if not concept_id:
            continue
        items[concept_id] = record
    return items


def _concept_governance_item_request(
    concept: dict,
    *,
    owner: str | None = None,
    assignee: str | None = None,
    review_note: str | None = None,
    changed_by: str | None = None,
) -> dict:
    concept_id = _normalized_text(concept.get("concept_id")) or None
    source_systems = [
        _normalized_text(value)
        for value in concept.get("source_systems", [])
        if _normalized_text(value)
    ]
    business_domains = [
        _normalized_text(value)
        for value in concept.get("business_domains", [])
        if _normalized_text(value)
    ]
    has_metadata = any(
        _normalized_text(value)
        for value in (owner, assignee, review_note)
    )
    return {
        "item_type": "concept_governance",
        "item_key": _concept_governance_item_key(concept_id),
        "title": f"Concept governance profile for '{concept_id or 'unknown'}'",
        "status": "approved" if has_metadata else "new",
        "concept_id": concept_id,
        "target": concept_id,
        "source_system": source_systems[0] if len(source_systems) == 1 else None,
        "business_domain": business_domains[0] if len(business_domains) == 1 else None,
        "owner": _normalized_text(owner) or None,
        "assignee": _normalized_text(assignee) or None,
        "review_note": _normalized_text(review_note) or None,
        "created_by": _normalized_text(changed_by) or None,
        "changed_by": _normalized_text(changed_by) or None,
    }


def _concept_governance_rows(records: list[dict] | None = None) -> list[dict]:
    rows: list[dict] = []
    for concept_id, item in _concept_governance_item_map(records).items():
        rows.append(
            {
                "concept_id": concept_id,
                "business_owner": _normalized_text(item.get("owner")),
                "data_steward": _normalized_text(item.get("assignee")),
                "status": _normalized_text(item.get("status")) or "new",
                "governance_note": _normalized_text(item.get("review_note")),
                "updated_at": _normalized_text(item.get("updated_at") or item.get("created_at")),
            }
        )
    return sorted(rows, key=lambda item: (_normalized_text(item.get("concept_id")).lower(),))


def _concept_governance_option_label(item: dict | None = None) -> str:
    concept_id = _normalized_text((item or {}).get("concept_id") or (item or {}).get("target")) or "unknown"
    owner = _normalized_text((item or {}).get("owner")) or "unassigned"
    assignee = _normalized_text((item or {}).get("assignee")) or "unassigned"
    return f"{concept_id} | owner={owner} | steward={assignee}"


def _normalized_governance_import_header(value: object) -> str:
    return _normalized_text(value).lower().replace(" ", "_")


def _concept_governance_import_rows(uploaded_file) -> list[dict[str, str]]:
    if uploaded_file is None:
        return []
    file_name = _normalized_text(getattr(uploaded_file, "name", "")).lower()
    file_bytes = uploaded_file.getvalue()
    rows: list[dict[str, str]] = []
    if file_name.endswith(".csv"):
        decoded = file_bytes.decode("utf-8-sig")
        reader = csv.DictReader(StringIO(decoded))
        for row in reader:
            rows.append({
                _normalized_governance_import_header(key): _normalized_text(value)
                for key, value in (row or {}).items()
                if _normalized_governance_import_header(key)
            })
        return rows
    if file_name.endswith(".xlsx"):
        workbook = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
        worksheet = workbook.active
        header_cells = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_cells:
            return []
        headers = [_normalized_governance_import_header(value) for value in header_cells]
        for row_values in worksheet.iter_rows(min_row=2, values_only=True):
            row: dict[str, str] = {}
            for index, cell_value in enumerate(row_values):
                header = headers[index] if index < len(headers) else ""
                if not header:
                    continue
                row[header] = _normalized_text(cell_value)
            rows.append(row)
        return rows
    raise ValueError("Supported concept governance import formats are CSV and XLSX.")


def _concept_governance_import_payloads(rows: list[dict[str, str]] | None) -> list[dict[str, str]]:
    payloads: list[dict[str, str]] = []
    for row in rows or []:
        concept_id = _normalized_text(
            row.get("concept_id")
            or row.get("canonical_concept_id")
            or row.get("canonical_term")
        )
        owner = _normalized_text(
            row.get("external_business_owner")
            or row.get("business_owner")
            or row.get("owner")
        )
        assignee = _normalized_text(
            row.get("external_data_steward")
            or row.get("data_steward")
            or row.get("assignee")
        )
        review_note = _normalized_text(
            row.get("external_governance_note")
            or row.get("governance_note")
            or row.get("review_note")
            or row.get("note")
        )
        if not concept_id:
            continue
        if not any((owner, assignee, review_note)):
            continue
        payloads.append(
            {
                "concept_id": concept_id,
                "owner": owner,
                "assignee": assignee,
                "review_note": review_note,
            }
        )
    return payloads


def _concept_governance_template_csv_bytes() -> bytes:
    buffer = StringIO()
    writer = csv.DictWriter(
        buffer,
        fieldnames=[
            "concept_id",
            "external_business_owner",
            "external_data_steward",
            "external_governance_note",
        ],
    )
    writer.writeheader()
    writer.writerow(
        {
            "concept_id": "sales.order",
            "external_business_owner": "order-to-cash lead",
            "external_data_steward": "semantic-model steward",
            "external_governance_note": "Mirrored from external governance register.",
        }
    )
    writer.writerow(
        {
            "concept_id": "finance.invoice",
            "external_business_owner": "finance controller",
            "external_data_steward": "finance data steward",
            "external_governance_note": "Owned in ERP master data council.",
        }
    )
    return buffer.getvalue().encode("utf-8")


def _overlay_promotion_item_key(entry: dict, overlay_version: dict | None = None) -> str:
    overlay_id = _normalized_text(entry.get("overlay_id") or entry.get("version_id") or (overlay_version or {}).get("overlay_id"))
    entry_id = _normalized_text(entry.get("entry_id"))
    alias = _normalized_text(entry.get("alias")) or "unknown"
    canonical_term = _normalized_text(entry.get("canonical_concept_id") or entry.get("canonical_term")) or "unknown"
    if overlay_id and entry_id:
        return f"overlay_promotion_{overlay_id}_{entry_id}".replace(" ", "_")
    return f"overlay_promotion_{overlay_id or canonical_term}_{alias}".replace(" ", "_")


def _overlay_promotion_item_map(records: list[dict] | None = None) -> dict[str, dict]:
    items: dict[str, dict] = {}
    for record in records or []:
        item_key = _normalized_text(record.get("item_key"))
        if not item_key or _normalized_text(record.get("item_type")) != "overlay_promotion":
            continue
        items[item_key] = record
    return items


def _overlay_promotion_status(item: dict | None = None) -> str:
    status = _normalized_text((item or {}).get("status"))
    if status in {"new", "needs_review", "ready_for_approval", "promoted", "ignored", "rejected"}:
        return status
    return "not_tracked"


def _overlay_promotion_can_execute(item: dict | None = None) -> bool:
    return bool((item or {}).get("item_id")) and _overlay_promotion_status(item) == "ready_for_approval"


def _overlay_promotion_execution_request(changed_by: str | None, note: str | None = None) -> dict:
    payload = {"changed_by": _normalized_text(changed_by) or None}
    if _normalized_text(note):
        payload["note"] = _normalized_text(note)
    return payload


def _overlay_promotion_entry_rows(
    entries: list[dict] | None,
    promotion_items: dict[str, dict] | None = None,
    overlay_version: dict | None = None,
) -> list[dict]:
    rows: list[dict] = []
    for entry in entries or []:
        if _normalized_text(entry.get("entry_type")) and _normalized_text(entry.get("entry_type")) != "concept_alias":
            continue
        item = (promotion_items or {}).get(_overlay_promotion_item_key(entry, overlay_version)) or {}
        rows.append(
            {
                "alias": _normalized_text(entry.get("alias")),
                "canonical_term": _normalized_text(entry.get("canonical_term")),
                "canonical_concept_id": _normalized_text(entry.get("canonical_concept_id")),
                "source_system": _normalized_text(entry.get("source_system")),
                "promotion_status": _overlay_promotion_status(item),
                "owner": _normalized_text(item.get("owner")),
                "assignee": _normalized_text(item.get("assignee")),
                "review_note": _normalized_text(item.get("review_note")),
            }
        )
    return rows


def _overlay_promotion_items_for_concept(concept: dict | None, records: list[dict] | None = None) -> list[dict]:
    concept_id = _normalized_text((concept or {}).get("concept_id"))
    if not concept_id:
        return []
    items: list[dict] = []
    for record in records or []:
        if _normalized_text(record.get("item_type")) != "overlay_promotion":
            continue
        record_concept_id = _normalized_text(record.get("concept_id") or record.get("target"))
        if record_concept_id != concept_id:
            continue
        items.append(record)
    return items


def _overlay_promotion_rows_for_concept(concept: dict | None, records: list[dict] | None = None) -> list[dict]:
    rows: list[dict] = []
    for item in _overlay_promotion_items_for_concept(concept, records):
        rows.append(
            {
                "alias": _normalized_text(item.get("source")),
                "status": _overlay_promotion_status(item),
                "source_system": _normalized_text(item.get("source_system")),
                "business_domain": _normalized_text(item.get("business_domain")),
                "owner": _normalized_text(item.get("owner")),
                "assignee": _normalized_text(item.get("assignee")),
                "review_note": _normalized_text(item.get("review_note")),
            }
        )
    return rows


def _overlay_promotion_item_record_label(item: dict | None = None) -> str:
    alias = _normalized_text((item or {}).get("source")) or "unknown"
    source_system = _normalized_text((item or {}).get("source_system")) or "n/a"
    status = _overlay_promotion_status(item)
    return f"{alias} | source_system={source_system} | status={status}"


def _overlay_promotion_option_label(entry: dict, item: dict | None = None) -> str:
    alias = _normalized_text(entry.get("alias")) or "unknown"
    concept_id = _normalized_text(entry.get("canonical_concept_id")) or _normalized_text(entry.get("canonical_term")) or "unknown"
    status = _overlay_promotion_status(item)
    return f"{alias} -> {concept_id} | status={status}"


def _overlay_promotion_item_request(
    entry: dict,
    overlay_version: dict | None,
    *,
    status: str,
    owner: str | None = None,
    assignee: str | None = None,
    review_note: str | None = None,
    changed_by: str | None = None,
) -> dict:
    overlay_payload = {
        **entry,
        "overlay_id": entry.get("overlay_id") or entry.get("version_id") or (overlay_version or {}).get("overlay_id"),
        "overlay_name": (overlay_version or {}).get("name"),
        "overlay_status": (overlay_version or {}).get("status"),
    }
    alias = _normalized_text(entry.get("alias")) or None
    concept_id = _normalized_text(entry.get("canonical_concept_id")) or None
    return {
        "item_type": "overlay_promotion",
        "item_key": _overlay_promotion_item_key(entry, overlay_version),
        "title": f"Promote overlay alias '{alias or 'unknown'}'",
        "status": status,
        "concept_id": concept_id,
        "source": alias,
        "target": concept_id or _normalized_text(entry.get("canonical_term")) or None,
        "source_system": _normalized_text(entry.get("source_system")) or None,
        "business_domain": _normalized_text(entry.get("domain")) or None,
        "owner": _normalized_text(owner) or None,
        "assignee": _normalized_text(assignee) or None,
        "review_note": _normalized_text(review_note) or None,
        "overlay_entry_payload": overlay_payload,
        "created_by": _normalized_text(changed_by) or None,
        "changed_by": _normalized_text(changed_by) or None,
    }


def _canonical_gap_effective_status(
    candidate_key: str,
    console_states: dict[str, str] | None = None,
    proposal_states: dict[str, str] | None = None,
    stewardship_item: dict | None = None,
) -> str:
    stewardship_state = _canonical_gap_stewardship_status(stewardship_item)
    if stewardship_state:
        return stewardship_state
    console_state = _canonical_gap_console_state(candidate_key, console_states)
    if console_state != "active":
        return console_state
    return _canonical_gap_proposal_state(candidate_key, proposal_states)


def _canonical_gap_stewardship_item_request(
    candidate_key: str,
    candidate: dict,
    suggestion: dict | None = None,
    *,
    status: str,
    owner: str | None = None,
    assignee: str | None = None,
    review_note: str | None = None,
    changed_by: str | None = None,
) -> dict:
    return {
        "item_type": "canonical_gap",
        "item_key": candidate_key,
        "title": f"{_normalized_text(candidate.get('source')) or 'unknown'} -> {_normalized_text(candidate.get('target')) or 'unknown'}",
        "status": status,
        "concept_id": _normalized_text((suggestion or {}).get("concept_id")) or None,
        "source": _normalized_text(candidate.get("source")) or None,
        "target": _normalized_text(candidate.get("target")) or None,
        "owner": _normalized_text(owner) or None,
        "assignee": _normalized_text(assignee) or None,
        "review_note": _normalized_text(review_note) or None,
        "candidate_payload": candidate,
        "suggestion_payload": suggestion or {},
        "created_by": _normalized_text(changed_by) or None,
        "changed_by": _normalized_text(changed_by) or None,
    }


def _canonical_gap_repeat_summary_rows(
    candidates: list[dict] | None,
    suggestions: dict[str, dict] | None = None,
    console_states: dict[str, str] | None = None,
    proposal_states: dict[str, str] | None = None,
    stewardship_items: dict[str, dict] | None = None,
) -> list[dict]:
    observations: dict[str, dict[str, object]] = {}
    seen_candidate_keys: set[str] = set()

    for index, candidate in enumerate(candidates or []):
        candidate_key = _canonical_gap_candidate_key(index, candidate)
        seen_candidate_keys.add(candidate_key)
        stewardship_item = (stewardship_items or {}).get(candidate_key) or {}
        suggestion = (suggestions or {}).get(candidate_key) or {}
        observations[candidate_key] = {
            "source": _normalized_text(candidate.get("source") or stewardship_item.get("source")) or "unknown",
            "target": _normalized_text(candidate.get("target") or stewardship_item.get("target")) or "unknown",
            "concept_id": _normalized_text(suggestion.get("concept_id") or stewardship_item.get("concept_id")),
            "status": _canonical_gap_effective_status(
                candidate_key,
                console_states,
                proposal_states,
                stewardship_item,
            ),
            "current_queue": True,
            "updated_at": _normalized_text(stewardship_item.get("updated_at") or stewardship_item.get("created_at")),
        }

    for item_key, stewardship_item in (stewardship_items or {}).items():
        if item_key in seen_candidate_keys:
            continue
        observations[item_key] = {
            "source": _normalized_text(stewardship_item.get("source")) or "unknown",
            "target": _normalized_text(stewardship_item.get("target")) or "unknown",
            "concept_id": _normalized_text(stewardship_item.get("concept_id")),
            "status": _canonical_gap_stewardship_status(stewardship_item) or "not_tracked",
            "current_queue": False,
            "updated_at": _normalized_text(stewardship_item.get("updated_at") or stewardship_item.get("created_at")),
        }

    grouped: dict[tuple[str, str], dict[str, object]] = {}
    for observation in observations.values():
        target = _normalized_text(observation.get("target")) or "unknown"
        concept_id = _normalized_text(observation.get("concept_id"))
        group_key = (target.lower(), concept_id.lower())
        group = grouped.setdefault(
            group_key,
            {
                "target": target,
                "suggested_concept_id": concept_id,
                "observations": 0,
                "distinct_source_count": 0,
                "current_queue_count": 0,
                "new": 0,
                "needs_review": 0,
                "ready_for_approval": 0,
                "approved": 0,
                "ignored": 0,
                "rejected": 0,
                "promoted": 0,
                "source_examples": [],
                "latest_observed_at": "",
                "_source_keys": set(),
            },
        )
        group["observations"] = int(group["observations"]) + 1
        source = _normalized_text(observation.get("source")) or "unknown"
        source_key = source.lower()
        source_keys = group["_source_keys"]
        if isinstance(source_keys, set) and source_key not in source_keys:
            source_keys.add(source_key)
            group["distinct_source_count"] = int(group["distinct_source_count"]) + 1
            source_examples = group["source_examples"]
            if isinstance(source_examples, list) and len(source_examples) < 4:
                source_examples.append(source)
        if observation.get("current_queue"):
            group["current_queue_count"] = int(group["current_queue_count"]) + 1
        status = _normalized_text(observation.get("status"))
        if status in {"new", "needs_review", "ready_for_approval", "approved", "ignored", "rejected", "promoted"}:
            group[status] = int(group[status]) + 1
        updated_at = _normalized_text(observation.get("updated_at"))
        if updated_at and updated_at > _normalized_text(group.get("latest_observed_at")):
            group["latest_observed_at"] = updated_at

    rows: list[dict] = []
    for group in grouped.values():
        if int(group["observations"]) < 2:
            continue
        rows.append(
            {
                "target": group["target"],
                "suggested_concept_id": group["suggested_concept_id"],
                "observations": group["observations"],
                "distinct_source_count": group["distinct_source_count"],
                "current_queue_count": group["current_queue_count"],
                "ready_for_approval": group["ready_for_approval"],
                "needs_review": group["needs_review"],
                "new": group["new"],
                "ignored": group["ignored"],
                "rejected": group["rejected"],
                "approved": group["approved"],
                "promoted": group["promoted"],
                "source_examples": ", ".join(group["source_examples"]),
                "latest_observed_at": group["latest_observed_at"],
            }
        )

    return sorted(
        rows,
        key=lambda item: (
            -int(item.get("observations", 0) or 0),
            -int(item.get("distinct_source_count", 0) or 0),
            _normalized_text(item.get("target")).lower(),
            _normalized_text(item.get("suggested_concept_id")).lower(),
        ),
    )


def _knowledge_stewardship_status_update_request(
    status: str,
    changed_by: str | None,
    *,
    owner: str | None = None,
    assignee: str | None = None,
    review_note: str | None = None,
    note: str | None = None,
) -> dict:
    payload = {
        "status": status,
        "changed_by": _normalized_text(changed_by) or None,
        "owner": _normalized_text(owner) or None,
        "assignee": _normalized_text(assignee) or None,
        "review_note": _normalized_text(review_note) or None,
    }
    if _normalized_text(note):
        payload["note"] = _normalized_text(note)
    return payload


def _canonical_gap_queue_rows(
    candidates: list[dict] | None,
    suggestions: dict[str, dict] | None = None,
    console_states: dict[str, str] | None = None,
    proposal_states: dict[str, str] | None = None,
    stewardship_items: dict[str, dict] | None = None,
) -> list[dict]:
    rows: list[dict] = []
    suggestion_map = suggestions or {}
    for index, candidate in enumerate(candidates or []):
        candidate_key = _canonical_gap_candidate_key(index, candidate)
        suggestion = suggestion_map.get(candidate_key) or {}
        stewardship_item = (stewardship_items or {}).get(candidate_key) or {}
        aliases = [str(alias).strip() for alias in suggestion.get("aliases", []) if str(alias).strip()]
        rows.append(
            {
                "source": candidate.get("source") or "",
                "target": candidate.get("target") or "",
                "confidence_pct": int(float(candidate.get("confidence", 0.0) or 0.0) * 100),
                "confidence_label": candidate.get("confidence_label") or "",
                "status": candidate.get("status") or "",
                "method": candidate.get("method") or "",
                "reason": candidate.get("reason") or "",
                "suggested_action": suggestion.get("action") or "pending",
                "suggested_concept": suggestion.get("concept_id") or "",
                "suggested_display_name": suggestion.get("display_name") or "",
                "alias_count": len(aliases),
                "reasoning_count": len(suggestion.get("reasoning") or []),
                "risk_count": len(suggestion.get("risk_notes") or []),
                "console_state": _canonical_gap_console_state(candidate_key, console_states, stewardship_item),
                "proposal_state": _canonical_gap_proposal_state(candidate_key, proposal_states, stewardship_item),
                "stewardship_status": _canonical_gap_effective_status(candidate_key, console_states, proposal_states, stewardship_item),
                "owner": stewardship_item.get("owner") or "",
                "assignee": stewardship_item.get("assignee") or "",
            }
        )
    return rows


def _canonical_gap_option_label(index: int, candidate: dict, suggestion: dict | None = None) -> str:
    source = _normalized_text(candidate.get("source")) or "unknown"
    target = _normalized_text(candidate.get("target")) or "unknown"
    confidence_pct = int(float(candidate.get("confidence", 0.0) or 0.0) * 100)
    action = _normalized_text((suggestion or {}).get("action")) or "pending"
    return f"{source} -> {target} | confidence={confidence_pct}% | action={action}"


def _canonical_gap_can_approve(
    suggestion: dict | None,
    console_state: str = "active",
    proposal_state: str = "new",
) -> bool:
    action = _normalized_text((suggestion or {}).get("action"))
    return bool(
        action
        and action != "no_action"
        and console_state == "active"
        and proposal_state == "ready_for_approval"
    )


def _canonical_gap_can_ignore(console_state: str) -> bool:
    return console_state == "active"


def _canonical_gap_can_restore(console_state: str) -> bool:
    return console_state == "ignored"


def _canonical_gap_can_reject(console_state: str) -> bool:
    return console_state == "active"


def _canonical_gap_rejection_request(
    candidate: dict,
    suggestion: dict | None,
    rejected_by: str | None,
    note: str | None,
    disposition: str = "rejected",
) -> dict:
    payload: dict[str, object] = {
        "candidate": candidate,
        "disposition": disposition,
        "rejected_by": _normalized_text(rejected_by) or "streamlit-admin-debug",
    }
    if suggestion:
        payload["suggestion"] = suggestion
    if _normalized_text(note):
        payload["note"] = _normalized_text(note)
    return payload


def _canonical_gap_related_audit_entries(audit_entries: list[dict] | None, candidate: dict | None) -> list[dict]:
    source = _normalized_text((candidate or {}).get("source"))
    target = _normalized_text((candidate or {}).get("target"))
    if not source or not target:
        return []

    needle = f"{source} -> {target}"
    rows: list[dict] = []
    for entry in audit_entries or []:
        message = _normalized_text(entry.get("message"))
        if needle not in message:
            continue
        rows.append(
            {
                "action": entry.get("action") or "",
                "overlay_name": entry.get("overlay_name") or "",
                "created_at": entry.get("created_at") or "",
                "message": message,
            }
        )
    return rows


def _canonical_gap_approval_request(candidate: dict, suggestion: dict, approved_by: str | None) -> dict:
    return {
        "candidate": candidate,
        "suggestion": suggestion,
        "approved_by": _normalized_text(approved_by) or "streamlit-admin-debug",
    }


def _canonical_gap_impact_preview_rows(
    selected_index: int,
    selected_candidate: dict,
    selected_suggestion: dict | None,
    candidates: list[dict] | None,
    suggestions: dict[str, dict] | None = None,
    console_states: dict[str, str] | None = None,
) -> list[dict]:
    action = _normalized_text((selected_suggestion or {}).get("action"))
    if action not in {"existing_concept_alias", "new_canonical_concept"}:
        return []

    selected_source = _normalized_text(selected_candidate.get("source")).lower()
    selected_target = _normalized_text(selected_candidate.get("target")).lower()
    suggested_concept_id = _normalized_text((selected_suggestion or {}).get("concept_id")).lower()
    alias_set = {
        _normalized_text(alias).lower()
        for alias in (selected_suggestion or {}).get("aliases", [])
        if _normalized_text(alias)
    }
    if selected_source:
        alias_set.add(selected_source)

    rows: list[dict] = []
    suggestion_map = suggestions or {}
    for index, candidate in enumerate(candidates or []):
        candidate_source = _normalized_text(candidate.get("source"))
        candidate_target = _normalized_text(candidate.get("target"))
        candidate_source_normalized = candidate_source.lower()
        candidate_target_normalized = candidate_target.lower()
        candidate_key = _canonical_gap_candidate_key(index, candidate)
        candidate_suggestion = suggestion_map.get(candidate_key) or {}

        impact_reasons: list[str] = []
        if index == selected_index:
            impact_reasons.append("selected gap under review")
        else:
            if selected_source and candidate_source_normalized == selected_source:
                impact_reasons.append("same source column")
            if alias_set and candidate_source_normalized in alias_set and candidate_source_normalized != selected_source:
                impact_reasons.append("source matches proposed alias")
            if selected_target and candidate_target_normalized == selected_target:
                impact_reasons.append("same target field")
            if suggested_concept_id and _normalized_text(candidate_suggestion.get("concept_id")).lower() == suggested_concept_id:
                impact_reasons.append("same suggested concept")

        if not impact_reasons:
            continue

        rows.append(
            {
                "source": candidate_source,
                "target": candidate_target,
                "confidence_pct": int(float(candidate.get("confidence", 0.0) or 0.0) * 100),
                "console_state": _canonical_gap_console_state(candidate_key, console_states),
                "impact_reason": " | ".join(dict.fromkeys(impact_reasons)),
            }
        )

    return rows


def _canonical_gap_pending_rows_for_concept(
    concept: dict | None,
    candidates: list[dict] | None,
    suggestions: dict[str, dict] | None = None,
    console_states: dict[str, str] | None = None,
    proposal_states: dict[str, str] | None = None,
    stewardship_items: dict[str, dict] | None = None,
) -> list[dict]:
    concept_id = _normalized_text((concept or {}).get("concept_id"))
    if not concept_id:
        return []

    rows: list[dict] = []
    suggestion_map = suggestions or {}
    for index, candidate in enumerate(candidates or []):
        candidate_key = _canonical_gap_candidate_key(index, candidate)
        suggestion = suggestion_map.get(candidate_key) or {}
        stewardship_item = (stewardship_items or {}).get(candidate_key) or {}
        if _normalized_text(suggestion.get("action")) != "existing_concept_alias":
            continue
        if _normalized_text(suggestion.get("concept_id")) != concept_id:
            continue

        console_state = _canonical_gap_console_state(candidate_key, console_states, stewardship_item)
        if console_state != "active":
            continue
        proposal_state = _canonical_gap_proposal_state(candidate_key, proposal_states, stewardship_item)

        reasoning = [str(item).strip() for item in suggestion.get("reasoning") or [] if str(item).strip()]
        risk_notes = [str(item).strip() for item in suggestion.get("risk_notes") or [] if str(item).strip()]
        aliases = [str(item).strip() for item in suggestion.get("aliases") or [] if str(item).strip()]
        rows.append(
            {
                "source": _normalized_text(candidate.get("source")),
                "target": _normalized_text(candidate.get("target")),
                "confidence_pct": int(float(candidate.get("confidence", 0.0) or 0.0) * 100),
                "suggested_action": _normalized_text(suggestion.get("action")) or "pending",
                "proposal_state": proposal_state,
                "aliases": ", ".join(aliases),
                "reasoning": " | ".join(reasoning),
                "risk_notes": " | ".join(risk_notes),
            }
        )

    order = {"ready_for_approval": 0, "needs_review": 1, "new": 2}
    rows.sort(
        key=lambda item: (
            order.get(_normalized_text(item.get("proposal_state")), 99),
            -int(item.get("confidence_pct", 0) or 0),
            item.get("source") or "",
        )
    )
    return rows


def _refresh_canonical_console_knowledge_state(*, api_request: Callable[..., Any]) -> None:
    st.session_state["debug_knowledge_runtime"] = api_request("POST", "/knowledge/reload")
    st.session_state["debug_knowledge_overlays"] = api_request("GET", "/knowledge/overlays")
    st.session_state["debug_knowledge_audit_logs"] = api_request("GET", "/knowledge/audit")
    st.session_state["debug_knowledge_stewardship_items"] = api_request("GET", "/knowledge/stewardship-items")
    st.session_state["debug_canonical_gap_proposal_states"] = _canonical_gap_proposal_state_map(
        st.session_state.get("debug_knowledge_stewardship_items")
    )


def _bootstrap_active_overlay_detail(*, api_request: Callable[..., Any]) -> None:
    if st.session_state.get("debug_selected_knowledge_overlay") is not None:
        return

    runtime = st.session_state.get("debug_knowledge_runtime") or {}
    active_overlay_id = runtime.get("active_overlay_id")
    if active_overlay_id is None:
        return

    overlays = st.session_state.get("debug_knowledge_overlays") or []
    active_overlay = next(
        (item for item in overlays if item.get("overlay_id") == active_overlay_id),
        None,
    )
    if active_overlay is None:
        return

    st.session_state["debug_selected_overlay_version"] = (
        f"#{active_overlay['overlay_id']} | {active_overlay['name']} | {active_overlay['status']}"
    )
    st.session_state["debug_selected_knowledge_overlay"] = api_request("GET", f"/knowledge/overlays/{active_overlay_id}")


def _ensure_canonical_concept_detail_loaded(*, api_request: Callable[..., Any], selected_concept_id: str | None) -> bool:
    normalized_selected_concept_id = _normalized_text(selected_concept_id)
    if not normalized_selected_concept_id:
        return False

    current_concept_id = _normalized_text(
        ((st.session_state.get("debug_canonical_concept_detail") or {}).get("concept") or {}).get("concept_id")
    )
    if current_concept_id == normalized_selected_concept_id:
        return False

    st.session_state["debug_canonical_concept_detail"] = api_request(
        "GET",
        f"/knowledge/canonical-concepts/{normalized_selected_concept_id}",
    )
    return True


def _ensure_knowledge_concept_detail_loaded(*, api_request: Callable[..., Any], selected_concept_id: str | None) -> bool:
    normalized_selected_concept_id = _normalized_text(selected_concept_id)
    if not normalized_selected_concept_id:
        return False

    current_concept_id = _normalized_text(
        ((st.session_state.get("debug_knowledge_concept_detail") or {}).get("concept") or {}).get("concept_id")
    )
    if current_concept_id == normalized_selected_concept_id:
        return False

    st.session_state["debug_knowledge_concept_detail"] = api_request(
        "GET",
        f"/knowledge/concepts/{normalized_selected_concept_id}",
    )
    return True


def _preferred_canonical_concept_label(
    concepts: list[dict] | None,
    records: list[dict] | None = None,
    *,
    active_overlay_id: object | None = None,
    current_label: str | None = None,
) -> str | None:
    concept_items = concepts or []
    if not concept_items:
        return None

    labels_by_concept_id = {
        _normalized_text(item.get("concept_id")): _canonical_concept_option_label(item)
        for item in concept_items
        if _normalized_text(item.get("concept_id"))
    }
    if current_label in labels_by_concept_id.values():
        return current_label

    normalized_active_overlay_id = _normalized_text(active_overlay_id)
    promotion_concept_ids: set[str] = set()
    active_overlay_promotion_concept_ids: set[str] = set()
    for record in records or []:
        if _normalized_text(record.get("item_type")) != "overlay_promotion":
            continue
        record_concept_id = _normalized_text(record.get("concept_id") or record.get("target"))
        if not record_concept_id:
            continue
        promotion_concept_ids.add(record_concept_id)
        record_overlay_id = _normalized_text(((record.get("overlay_entry_payload") or {}).get("overlay_id")))
        if normalized_active_overlay_id and record_overlay_id == normalized_active_overlay_id:
            active_overlay_promotion_concept_ids.add(record_concept_id)

    def concept_rank(item: dict) -> tuple[int, int, int, str]:
        concept_id = _normalized_text(item.get("concept_id"))
        source = _normalized_text(item.get("source")).lower() or "base"
        active_overlay_entry_count = int(item.get("active_overlay_entry_count", 0) or 0)
        usage_count = int(item.get("usage_count", 0) or 0)
        return (
            0 if concept_id in active_overlay_promotion_concept_ids else 1,
            0 if concept_id in promotion_concept_ids else 1,
            0 if active_overlay_entry_count > 0 or source == "overlay_only" else 1,
            -usage_count,
        )

    preferred_concept = min(concept_items, key=concept_rank)
    return labels_by_concept_id.get(_normalized_text(preferred_concept.get("concept_id")))


def _bootstrap_canonical_console_state(*, api_request: Callable[..., Any]) -> None:
    if st.session_state.get("debug_canonical_console_manual_clear"):
        return
    if st.session_state.get("debug_canonical_console_bootstrapped"):
        return

    _refresh_canonical_console_knowledge_state(api_request=api_request)
    _bootstrap_active_overlay_detail(api_request=api_request)
    st.session_state["debug_canonical_concepts"] = api_request("GET", "/knowledge/canonical-concepts")
    st.session_state["debug_knowledge_concepts"] = api_request("GET", "/knowledge/concepts")
    st.session_state["debug_canonical_console_bootstrapped"] = True


def _canonical_console_action_label(loaded: bool, noun: str) -> str:
    prefix = "Refresh" if loaded else "Load"
    return f"{prefix} {noun}"


def render_canonical_console_panel(
    *,
    api_request: Callable[..., Any],
    api_request_content: Callable[..., bytes],
    upload_file_to_request_files: Callable[[Any], dict | None],
) -> None:
    """Render the Canonical Console for registry browsing, overlay lifecycle, and stewardship review."""

    st.header("Governance")
    st.caption(
        "Governance console for canonical, knowledge, overlay runtime, and stewardship workflows without changing the underlying authoring logic."
    )
    render_status_badge_legend(title="Governance Status Legend")

    try:
        _bootstrap_canonical_console_state(api_request=api_request)
    except httpx.HTTPError as error:
        st.session_state["last_action"] = {
            "level": "error",
            "message": f"Bootstrapping canonical console failed: {error}",
        }

    if "debug_knowledge_audit_logs" not in st.session_state:
        try:
            st.session_state["debug_knowledge_audit_logs"] = api_request("GET", "/knowledge/audit")
        except httpx.HTTPError:
            st.session_state["debug_knowledge_audit_logs"] = []

    if "debug_knowledge_stewardship_items" not in st.session_state:
        try:
            st.session_state["debug_knowledge_stewardship_items"] = api_request("GET", "/knowledge/stewardship-items")
        except httpx.HTTPError:
            st.session_state["debug_knowledge_stewardship_items"] = []

    if "debug_canonical_gap_proposal_states" not in st.session_state:
        try:
            st.session_state["debug_canonical_gap_proposal_states"] = _canonical_gap_proposal_state_map(
                st.session_state.get("debug_knowledge_stewardship_items")
            )
        except httpx.HTTPError:
            st.session_state["debug_canonical_gap_proposal_states"] = {}

    resolve_active_governance_section(st.session_state)
    active_governance_section = st.radio(
        "Governance section",
        GOVERNANCE_SECTIONS,
        key="active_governance_section",
        horizontal=True,
    )

    if active_governance_section == "Overlays & Runtime":
        overlay_header_actions = st.columns(2)
        if overlay_header_actions[0].button(
            "Refresh overlay summary",
            width="stretch",
            key="canonical_console_refresh_overlay_summary",
        ):
            try:
                st.session_state.pop("debug_canonical_console_manual_clear", None)
                _refresh_canonical_console_knowledge_state(api_request=api_request)
                _bootstrap_active_overlay_detail(api_request=api_request)
                if st.session_state.get("debug_canonical_concepts") is None:
                    st.session_state["debug_canonical_concepts"] = api_request("GET", "/knowledge/canonical-concepts")
                if st.session_state.get("debug_knowledge_concepts") is None:
                    st.session_state["debug_knowledge_concepts"] = api_request("GET", "/knowledge/concepts")
                st.session_state["debug_canonical_console_bootstrapped"] = True
                st.session_state["last_action"] = {
                    "level": "success",
                    "message": "Refreshed overlay summary and knowledge audit state.",
                }
            except httpx.HTTPError as error:
                st.session_state["last_action"] = {
                    "level": "error",
                    "message": f"Refreshing overlay summary failed: {error}",
                }
            st.rerun()

        overlay_summary = _canonical_overlay_summary(
            st.session_state.get("debug_knowledge_runtime"),
            st.session_state.get("debug_knowledge_overlays"),
        )
        with st.expander(
            _section_label("Overlay Summary", str(overlay_summary.get("active_overlay_name") or "none")),
            expanded=False,
        ):
            overlay_summary_columns = st.columns(5)
            overlay_summary_columns[0].metric("Active overlay", str(overlay_summary.get("active_overlay_name") or "none"))
            overlay_summary_columns[1].metric("Active entries", int(overlay_summary.get("active_entry_count", 0) or 0))
            overlay_summary_columns[2].metric("Concept aliases", int(overlay_summary.get("concept_alias_entries", 0) or 0))
            overlay_summary_columns[3].metric("Versions", int(overlay_summary.get("total_versions", 0) or 0))
            overlay_summary_columns[4].metric("Validated", int(overlay_summary.get("validated_versions", 0) or 0))
            st.caption(
                f"Mode={overlay_summary.get('mode') or 'base_only'} | "
                f"runtime_source={overlay_summary.get('runtime_source') or 'unknown'} | "
                f"source_hash_state={overlay_summary.get('source_hash_state') or 'missing'} | "
                f"active_versions={overlay_summary.get('active_versions', 0)} | "
                f"archived_versions={overlay_summary.get('archived_versions', 0)}"
            )

        st.subheader("Overlay Management")
        overlay_action_columns = st.columns(4)
        if overlay_action_columns[0].button(
            _canonical_console_action_label(st.session_state.get("debug_knowledge_overlays") is not None, "knowledge overlays"),
            width="stretch",
            key="canonical_console_load_knowledge_overlays",
        ):
            try:
                st.session_state["debug_knowledge_overlays"] = api_request("GET", "/knowledge/overlays")
                _bootstrap_active_overlay_detail(api_request=api_request)
                st.session_state["last_action"] = {"level": "success", "message": "Loaded knowledge overlay versions."}
            except httpx.HTTPError as error:
                st.session_state["last_action"] = {"level": "error", "message": f"Loading knowledge overlays failed: {error}"}
            st.rerun()

        if overlay_action_columns[1].button(
            "Reload knowledge",
            width="stretch",
            key="canonical_console_reload_knowledge",
        ):
            try:
                _refresh_canonical_console_knowledge_state(api_request=api_request)
                st.session_state["last_action"] = {"level": "success", "message": "Reloaded active knowledge overlay into runtime."}
            except httpx.HTTPError as error:
                st.session_state["last_action"] = {"level": "error", "message": f"Knowledge reload failed: {error}"}
            st.rerun()

        if overlay_action_columns[2].button(
            _canonical_console_action_label(st.session_state.get("debug_knowledge_runtime") is not None, "active knowledge status"),
            width="stretch",
            key="canonical_console_load_knowledge_runtime",
        ):
            try:
                st.session_state["debug_knowledge_runtime"] = api_request("POST", "/knowledge/reload")
                st.session_state["last_action"] = {"level": "success", "message": "Loaded active knowledge runtime status."}
            except httpx.HTTPError as error:
                st.session_state["last_action"] = {"level": "error", "message": f"Loading knowledge runtime status failed: {error}"}
            st.rerun()

        if overlay_action_columns[3].button(
            "Load overlay audit log",
            width="stretch",
            key="canonical_console_load_overlay_audit",
        ):
            try:
                st.session_state["debug_knowledge_audit_logs"] = api_request("GET", "/knowledge/audit")
                st.session_state["last_action"] = {"level": "success", "message": "Loaded knowledge audit log."}
            except httpx.HTTPError as error:
                st.session_state["last_action"] = {"level": "error", "message": f"Loading knowledge audit log failed: {error}"}
            st.rerun()

        knowledge_upload = st.file_uploader(
            "Knowledge overlay CSV",
            type=["csv"],
            key="knowledge_overlay_file",
            help="Upload CSV entries for abbreviations, synonyms, field aliases, or concept aliases.",
        )
        knowledge_overlay_name = st.text_input(
            "Overlay version name",
            value="",
            key="knowledge_overlay_name",
            placeholder="Example: customer-domain-overlay-v1",
        )
        knowledge_overlay_created_by = st.text_input(
            "Created by",
            value="",
            key="knowledge_overlay_created_by",
            placeholder="Example: data-governance-team",
        )
        knowledge_upload_columns = st.columns(2)
        if knowledge_upload_columns[0].button(
            "Validate knowledge CSV",
            width="stretch",
            key="canonical_console_validate_knowledge_overlay",
            disabled=knowledge_upload is None,
        ):
            try:
                st.session_state["debug_knowledge_validation"] = api_request(
                    "POST",
                    "/knowledge/overlays/validate",
                    files=upload_file_to_request_files(knowledge_upload),
                )
                st.session_state["last_action"] = {"level": "success", "message": "Validated knowledge overlay CSV."}
            except httpx.HTTPError as error:
                st.session_state["last_action"] = {"level": "error", "message": f"Knowledge CSV validation failed: {error}"}
            st.rerun()

        if knowledge_upload_columns[1].button(
            "Save overlay version",
            width="stretch",
            key="canonical_console_save_knowledge_overlay",
            disabled=knowledge_upload is None,
        ):
            try:
                created = api_request(
                    "POST",
                    "/knowledge/overlays",
                    files=upload_file_to_request_files(knowledge_upload),
                    data={
                        key: value
                        for key, value in {
                            "name": knowledge_overlay_name.strip(),
                            "created_by": knowledge_overlay_created_by.strip(),
                        }.items()
                        if value
                    }
                    or None,
                )
                st.session_state["debug_knowledge_created"] = created
                st.session_state["debug_knowledge_validation"] = created.get("validation")
                _refresh_canonical_console_knowledge_state(api_request=api_request)
                st.session_state["last_action"] = {
                    "level": "success",
                    "message": f"Saved knowledge overlay version '{created['version']['name']}' with {created['saved_entry_count']} entries.",
                }
            except httpx.HTTPError as error:
                st.session_state["last_action"] = {"level": "error", "message": f"Saving knowledge overlay failed: {error}"}
            st.rerun()

        knowledge_validation = st.session_state.get("debug_knowledge_validation")
        if knowledge_validation:
            st.caption(
                f"Validation summary: total={knowledge_validation.get('total_rows', 0)} | "
                f"valid={knowledge_validation.get('valid_rows', 0)} | invalid={knowledge_validation.get('invalid_rows', 0)} | "
                f"duplicates={knowledge_validation.get('duplicate_rows', 0)} | conflicts={knowledge_validation.get('conflicts', 0)}"
            )

        knowledge_overlays = st.session_state.get("debug_knowledge_overlays")
        if knowledge_overlays:
            st.dataframe(knowledge_overlays, width="stretch", hide_index=True)
            overlay_options = {
                f"#{item['overlay_id']} | {item['name']} | {item['status']}": item["overlay_id"]
                for item in knowledge_overlays
            }
            if overlay_options:
                selected_overlay_label = st.selectbox(
                    "Overlay version",
                    list(overlay_options.keys()),
                    key="debug_selected_overlay_version",
                )
                selected_overlay_id = overlay_options[selected_overlay_label]
                selected_overlay = next(
                    (item for item in knowledge_overlays if item.get("overlay_id") == selected_overlay_id),
                    None,
                )
                activation_block_reason = _overlay_activation_block_reason(selected_overlay)
                archive_block_reason = _overlay_archive_block_reason(selected_overlay)
                overlay_columns = st.columns(4)
                if overlay_columns[0].button("Load details", width="stretch", key="canonical_console_load_overlay_details"):
                    try:
                        st.session_state["debug_selected_knowledge_overlay"] = api_request("GET", f"/knowledge/overlays/{selected_overlay_id}")
                        st.session_state["last_action"] = {
                            "level": "success",
                            "message": f"Loaded knowledge overlay details for version #{selected_overlay_id}.",
                        }
                    except httpx.HTTPError as error:
                        st.session_state["last_action"] = {"level": "error", "message": f"Loading overlay details failed: {error}"}
                    st.rerun()

                if overlay_columns[1].button(
                    "Activate selected overlay",
                    width="stretch",
                    key="canonical_console_activate_overlay",
                    disabled=bool(activation_block_reason),
                ):
                    try:
                        activated = api_request("POST", f"/knowledge/overlays/{selected_overlay_id}/activate")
                        st.session_state["debug_selected_knowledge_overlay"] = api_request("GET", f"/knowledge/overlays/{selected_overlay_id}")
                        _refresh_canonical_console_knowledge_state(api_request=api_request)
                        st.session_state["last_action"] = {
                            "level": "success",
                            "message": f"Activated knowledge overlay '{activated['name']}'.",
                        }
                    except httpx.HTTPError as error:
                        st.session_state["last_action"] = {"level": "error", "message": f"Overlay activation failed: {error}"}
                    st.rerun()
                if activation_block_reason:
                    st.caption(activation_block_reason)

                if overlay_columns[2].button("Deactivate selected overlay", width="stretch", key="canonical_console_deactivate_overlay"):
                    try:
                        deactivated = api_request("POST", f"/knowledge/overlays/{selected_overlay_id}/deactivate")
                        st.session_state["debug_selected_knowledge_overlay"] = api_request("GET", f"/knowledge/overlays/{selected_overlay_id}")
                        _refresh_canonical_console_knowledge_state(api_request=api_request)
                        st.session_state["last_action"] = {
                            "level": "success",
                            "message": f"Deactivated knowledge overlay '{deactivated['name']}'.",
                        }
                    except httpx.HTTPError as error:
                        st.session_state["last_action"] = {"level": "error", "message": f"Overlay deactivation failed: {error}"}
                    st.rerun()

                if overlay_columns[3].button(
                    "Archive selected overlay",
                    width="stretch",
                    key="canonical_console_archive_overlay",
                    disabled=bool(archive_block_reason),
                ):
                    try:
                        archived = api_request("POST", f"/knowledge/overlays/{selected_overlay_id}/archive")
                        st.session_state["debug_selected_knowledge_overlay"] = api_request("GET", f"/knowledge/overlays/{selected_overlay_id}")
                        _refresh_canonical_console_knowledge_state(api_request=api_request)
                        st.session_state["last_action"] = {
                            "level": "success",
                            "message": f"Archived knowledge overlay '{archived['name']}'.",
                        }
                    except httpx.HTTPError as error:
                        st.session_state["last_action"] = {"level": "error", "message": f"Overlay archive failed: {error}"}
                    st.rerun()
                if archive_block_reason:
                    st.caption(archive_block_reason)

                if st.button("Rollback active overlay", width="stretch", key="canonical_console_rollback_overlay"):
                    try:
                        runtime = api_request("POST", "/knowledge/overlays/rollback")
                        st.session_state["debug_knowledge_runtime"] = runtime
                        st.session_state["debug_knowledge_overlays"] = api_request("GET", "/knowledge/overlays")
                        st.session_state["debug_knowledge_audit_logs"] = api_request("GET", "/knowledge/audit")
                        active_overlay_id = runtime.get("active_overlay_id")
                        st.session_state["debug_selected_knowledge_overlay"] = (
                            api_request("GET", f"/knowledge/overlays/{active_overlay_id}") if active_overlay_id else None
                        )
                        st.session_state["last_action"] = {
                            "level": "success",
                            "message": (
                                f"Rolled back to knowledge overlay '{runtime['active_overlay_name']}'."
                                if active_overlay_id
                                else "Rolled back to base-only knowledge mode."
                            ),
                        }
                    except httpx.HTTPError as error:
                        st.session_state["last_action"] = {"level": "error", "message": f"Overlay rollback failed: {error}"}
                    st.rerun()

        selected_overlay = st.session_state.get("debug_selected_knowledge_overlay")
        if selected_overlay:
            version = selected_overlay.get("version", {})
            overlay_entry_counts: dict[str, int] = {}
            entries = selected_overlay.get("entries", [])
            for entry in entries:
                entry_type = str(entry.get("entry_type") or "")
                if not entry_type:
                    continue
                overlay_entry_counts[entry_type] = overlay_entry_counts.get(entry_type, 0) + 1
            overlay_promotion_items = _overlay_promotion_item_map(st.session_state.get("debug_knowledge_stewardship_items"))
            promotable_entries = [
                entry
                for entry in entries
                if _normalized_text(entry.get("entry_type")) == "concept_alias" and entry.get("entry_id") is not None
            ]
            with st.expander(
                _section_label("Selected Overlay Detail", _normalized_text(version.get("name")) or "selected"),
                expanded=False,
            ):
                st.caption(
                    f"Overlay detail: #{version.get('overlay_id')} | {version.get('name')} | status={version.get('status')} | created_by={version.get('created_by') or 'n/a'} | source={version.get('source_filename') or 'n/a'}"
                )
                if overlay_entry_counts:
                    st.caption(
                        "Overlay entry summary: "
                        + " | ".join(f"{entry_type}={count}" for entry_type, count in sorted(overlay_entry_counts.items()))
                    )
                if entries:
                    st.dataframe(entries, width="stretch", hide_index=True)
                else:
                    st.info("This overlay version does not contain any saved entries.")

            if promotable_entries:
                promotion_rows = _overlay_promotion_entry_rows(promotable_entries, overlay_promotion_items, version)
                with st.expander(
                    _section_label("Overlay Promotion Candidates", f"{len(promotable_entries)} aliases"),
                    expanded=False,
                ):
                    st.caption(
                        "Track which active overlay aliases should be promoted into the stable canonical glossary later. "
                        "This does not write to the base glossary; it creates a durable stewardship candidate."
                    )
                    promotion_metrics = st.columns(4)
                    promotion_metrics[0].metric("Promotable aliases", len(promotable_entries))
                    promotion_metrics[1].metric(
                        "Tracked",
                        sum(1 for row in promotion_rows if row.get("promotion_status") != "not_tracked"),
                    )
                    promotion_metrics[2].metric(
                        "Ready",
                        sum(1 for row in promotion_rows if row.get("promotion_status") == "ready_for_approval"),
                    )
                    promotion_metrics[3].metric(
                        "Promoted",
                        sum(1 for row in promotion_rows if row.get("promotion_status") == "promoted"),
                    )
                    st.dataframe(promotion_rows, width="stretch", hide_index=True)

                    promotion_options = {
                        _overlay_promotion_option_label(
                            entry,
                            overlay_promotion_items.get(_overlay_promotion_item_key(entry, version)),
                        ): entry
                        for entry in promotable_entries
                    }
                    selected_promotion_label = st.selectbox(
                        "Overlay promotion detail",
                        list(promotion_options.keys()),
                        key="debug_selected_overlay_promotion_label",
                    )
                    selected_promotion_entry = promotion_options[selected_promotion_label]
                    selected_promotion_key = _overlay_promotion_item_key(selected_promotion_entry, version)
                    selected_promotion_item = overlay_promotion_items.get(selected_promotion_key) or {}
                    selected_promotion_item_id = selected_promotion_item.get("item_id")
                    selected_promotion_status = _overlay_promotion_status(selected_promotion_item)

                    st.caption(
                        f"Promotion status: {_canonical_gap_proposal_state_label(selected_promotion_status)} | "
                        f"alias={_normalized_text(selected_promotion_entry.get('alias')) or 'n/a'} | "
                        f"concept={_normalized_text(selected_promotion_entry.get('canonical_concept_id')) or _normalized_text(selected_promotion_entry.get('canonical_term')) or 'n/a'}"
                    )
                    if selected_promotion_entry.get("note"):
                        st.caption(f"Overlay note: {_normalized_text(selected_promotion_entry.get('note'))}")

                    promotion_owner_columns = st.columns(2)
                    promotion_owner = promotion_owner_columns[0].text_input(
                        "Promotion owner",
                        value=_normalized_text(selected_promotion_item.get("owner")),
                        key=f"debug_overlay_promotion_owner_{selected_promotion_key}",
                        placeholder="Example: master-data-governance",
                    )
                    promotion_assignee = promotion_owner_columns[1].text_input(
                        "Promotion assignee",
                        value=_normalized_text(selected_promotion_item.get("assignee")),
                        key=f"debug_overlay_promotion_assignee_{selected_promotion_key}",
                        placeholder="Example: canonical-model-owner",
                    )
                    promotion_review_note = st.text_input(
                        "Promotion review note",
                        value=_normalized_text(selected_promotion_item.get("review_note")),
                        key=f"debug_overlay_promotion_review_note_{selected_promotion_key}",
                        placeholder="Why should this overlay alias be promoted to the stable glossary?",
                    )
                    promotion_status_options = ["new", "needs_review", "ready_for_approval", "ignored", "rejected"]
                    if selected_promotion_status == "promoted":
                        promotion_status_options.append("promoted")
                    promotion_status = st.selectbox(
                        "Promotion status",
                        options=promotion_status_options,
                        index=promotion_status_options.index(selected_promotion_status) if selected_promotion_status in promotion_status_options else 0,
                        key=f"debug_overlay_promotion_status_{selected_promotion_key}",
                        format_func=_canonical_gap_proposal_state_label,
                    )
                    promotion_save_disabled = _normalized_text(version.get("status")) != "active" and not selected_promotion_item_id
                    promotion_action_columns = st.columns(2)
                    if promotion_action_columns[0].button(
                        "Save promotion candidate",
                        width="stretch",
                        key=f"debug_save_overlay_promotion_{selected_promotion_key}",
                        disabled=promotion_save_disabled or selected_promotion_status == "promoted",
                    ):
                        try:
                            api_request(
                                "POST",
                                "/knowledge/stewardship-items",
                                json=_overlay_promotion_item_request(
                                    selected_promotion_entry,
                                    version,
                                    status=promotion_status,
                                    owner=promotion_owner,
                                    assignee=promotion_assignee,
                                    review_note=promotion_review_note,
                                    changed_by=st.session_state.get("admin_token"),
                                ),
                            )
                            st.session_state["debug_knowledge_stewardship_items"] = api_request("GET", "/knowledge/stewardship-items")
                            st.session_state["debug_canonical_gap_proposal_states"] = _canonical_gap_proposal_state_map(
                                st.session_state.get("debug_knowledge_stewardship_items")
                            )
                            st.session_state["debug_knowledge_audit_logs"] = api_request("GET", "/knowledge/audit")
                            st.session_state["last_action"] = {
                                "level": "info",
                                "message": "Saved durable overlay promotion candidate for the selected alias.",
                            }
                        except httpx.HTTPError as error:
                            st.session_state["last_action"] = {
                                "level": "error",
                                "message": f"Saving overlay promotion candidate failed: {error}",
                            }
                        st.rerun()
                    if promotion_action_columns[1].button(
                        "Promote to stable glossary",
                        width="stretch",
                        key=f"debug_execute_overlay_promotion_{selected_promotion_key}",
                        disabled=not _overlay_promotion_can_execute(selected_promotion_item),
                    ):
                        try:
                            promotion_result = api_request(
                                "POST",
                                f"/knowledge/stewardship-items/{selected_promotion_item_id}/promote-to-glossary",
                                json=_overlay_promotion_execution_request(
                                    st.session_state.get("admin_token"),
                                    note="Promoted from Canonical Console overlay detail.",
                                ),
                            )
                            _refresh_canonical_console_knowledge_state(api_request=api_request)
                            st.session_state["debug_selected_knowledge_overlay"] = api_request(
                                "GET",
                                f"/knowledge/overlays/{version.get('overlay_id')}",
                            )
                            if st.session_state.get("debug_canonical_concepts") is not None:
                                st.session_state["debug_canonical_concepts"] = api_request("GET", "/knowledge/canonical-concepts")
                            promoted_concept_id = _normalized_text((promotion_result.get("glossary_entry") or {}).get("concept_id"))
                            if promoted_concept_id and promoted_concept_id == _normalized_text(
                                ((st.session_state.get("debug_canonical_concept_detail") or {}).get("concept") or {}).get("concept_id")
                            ):
                                st.session_state["debug_canonical_concept_detail"] = api_request(
                                    "GET",
                                    f"/knowledge/canonical-concepts/{promoted_concept_id}",
                                )
                            st.session_state["canonical_glossary_export_bytes"] = api_request_content(
                                "GET",
                                "/knowledge/canonical-glossary/export",
                            )
                            st.session_state["last_action"] = {
                                "level": "success",
                                "message": (
                                    "Promoted selected overlay alias into the stable canonical glossary."
                                    if promotion_result.get("alias_added", True)
                                    else "Selected overlay alias was already present in the stable canonical glossary; stewardship item marked as promoted."
                                ),
                            }
                        except httpx.HTTPError as error:
                            st.session_state["last_action"] = {
                                "level": "error",
                                "message": f"Stable glossary promotion failed: {error}",
                            }
                        st.rerun()
                    if selected_promotion_status == "ready_for_approval":
                        st.caption("Promotion execution writes the alias into the base canonical glossary and then marks the stewardship item as promoted.")
                    elif selected_promotion_status == "promoted":
                        st.caption("This alias has already been promoted to the stable canonical glossary.")
                    if _normalized_text(version.get("status")) != "active" and not selected_promotion_item_id:
                        st.caption("Activate this overlay before creating a new promotion candidate. Existing candidates can still be reviewed.")

    if active_governance_section == "Knowledge":
        st.caption("Knowledge registry stewardship with concept counts, linked canonical paths, and promotion readiness.")
        knowledge_header_actions = st.columns(2)
        if knowledge_header_actions[0].button(
            _canonical_console_action_label(st.session_state.get("debug_knowledge_concepts") is not None, "knowledge concept registry"),
            width="stretch",
            key="debug_load_knowledge_concepts",
        ):
            try:
                st.session_state.pop("debug_canonical_console_manual_clear", None)
                st.session_state["debug_knowledge_concepts"] = api_request("GET", "/knowledge/concepts")
                _refresh_canonical_console_knowledge_state(api_request=api_request)
                st.session_state["debug_canonical_console_bootstrapped"] = True
                st.session_state["last_action"] = {"level": "success", "message": "Loaded knowledge concept registry."}
            except httpx.HTTPError as error:
                st.session_state["last_action"] = {
                    "level": "error",
                    "message": f"Loading knowledge concept registry failed: {error}",
                }
            st.rerun()

        if knowledge_header_actions[1].button(
            _canonical_console_action_label(st.session_state.get("debug_knowledge_audit_logs") is not None, "knowledge audit log"),
            width="stretch",
            key="canonical_console_load_audit",
        ):
            try:
                st.session_state.pop("debug_canonical_console_manual_clear", None)
                st.session_state["debug_knowledge_audit_logs"] = api_request("GET", "/knowledge/audit")
                st.session_state["last_action"] = {"level": "success", "message": "Loaded knowledge audit log."}
            except httpx.HTTPError as error:
                st.session_state["last_action"] = {
                    "level": "error",
                    "message": f"Loading knowledge audit log failed: {error}",
                }
            st.rerun()

        with st.expander("Knowledge Registry", expanded=False):
            knowledge_registry_upload = st.file_uploader(
            "Base knowledge registry CSV",
            type=["csv"],
            key="knowledge_registry_file",
            help="Import or export the base metadata-driven knowledge registry as CSV.",
            )
            knowledge_registry_columns = st.columns(2)
            if knowledge_registry_columns[0].button(
                "Load knowledge registry export",
                width="stretch",
                key="canonical_console_export_knowledge_registry",
            ):
                try:
                    st.session_state["knowledge_registry_export_bytes"] = api_request_content("GET", "/knowledge/base-registry/export")
                    st.session_state["last_action"] = {"level": "success", "message": "Loaded knowledge registry export."}
                except httpx.HTTPError as error:
                    st.session_state["last_action"] = {"level": "error", "message": f"Knowledge registry export failed: {error}"}
                st.rerun()

            if knowledge_registry_columns[1].button(
                "Import knowledge registry",
                width="stretch",
                key="canonical_console_import_knowledge_registry",
                disabled=knowledge_registry_upload is None,
            ):
                try:
                    st.session_state["debug_knowledge_registry_import"] = api_request(
                        "POST",
                        "/knowledge/base-registry/import",
                        files=upload_file_to_request_files(knowledge_registry_upload),
                    )
                    _refresh_canonical_console_knowledge_state(api_request=api_request)
                    st.session_state["debug_knowledge_concepts"] = api_request("GET", "/knowledge/concepts")
                    selected_knowledge_concept_id = _normalized_text(
                        ((st.session_state.get("debug_knowledge_concept_detail") or {}).get("concept") or {}).get("concept_id")
                    )
                    if selected_knowledge_concept_id:
                        try:
                            st.session_state["debug_knowledge_concept_detail"] = api_request(
                                "GET",
                                f"/knowledge/concepts/{selected_knowledge_concept_id}",
                            )
                        except httpx.HTTPError:
                            st.session_state.pop("debug_knowledge_concept_detail", None)
                    st.session_state["last_action"] = {"level": "success", "message": "Imported knowledge registry CSV."}
                except httpx.HTTPError as error:
                    st.session_state["last_action"] = {"level": "error", "message": f"Knowledge registry import failed: {error}"}
                st.rerun()

            knowledge_registry_export_bytes = st.session_state.get("knowledge_registry_export_bytes")
            if knowledge_registry_export_bytes:
                st.download_button(
                    "Download knowledge registry CSV",
                    data=knowledge_registry_export_bytes,
                    file_name="metadata_dict.csv",
                    mime="text/csv",
                    width="stretch",
                )

            knowledge_registry_import = st.session_state.get("debug_knowledge_registry_import")
            if knowledge_registry_import:
                st.caption(
                    "Knowledge registry import: "
                    f"rows={knowledge_registry_import.get('imported_row_count', 0)}, "
                    f"concepts={knowledge_registry_import.get('knowledge_concept_count', 0)}."
                )
            st.caption("Editable concepts come from the base metadata registry. Derived runtime concepts remain visible here but are not directly editable through this CSV import.")

        knowledge_concepts = st.session_state.get("debug_knowledge_concepts") or []
        if knowledge_concepts:
            available_knowledge_source_systems = sorted(
            {
                _normalized_text(value)
                for concept in knowledge_concepts
                for value in concept.get("source_systems", [])
                if _normalized_text(value)
            }
            )
            if "debug_knowledge_concept_focus" not in st.session_state:
                st.session_state["debug_knowledge_concept_focus"] = "editable"
            with st.expander("Knowledge Concept Registry", expanded=True):
                filter_columns = st.columns(4)
                concept_query = filter_columns[0].text_input(
                    "Knowledge concept search",
                    value=st.session_state.get("debug_knowledge_concept_query", ""),
                    key="debug_knowledge_concept_query",
                    placeholder="Search by concept id, name, alias, linked canonical concept, domain, or source",
                )
                concept_focus = filter_columns[1].selectbox(
                    "Knowledge focus",
                    options=["all", "editable", "derived", "generated", "linked", "with_context"],
                    key="debug_knowledge_concept_focus",
                    format_func=lambda value: {
                        "all": "All concepts",
                        "editable": "Editable base concepts",
                        "derived": "Derived runtime concepts",
                        "generated": "Generated runtime concepts",
                        "linked": "Linked to canonical",
                        "with_context": "With field contexts",
                    }.get(value, value),
                )
                concept_source_system = filter_columns[2].selectbox(
                    "Source system",
                    options=[""] + available_knowledge_source_systems,
                    key="debug_knowledge_concept_source_system",
                    format_func=lambda value: value or "All source systems",
                )
                concept_source = filter_columns[3].selectbox(
                    "Registry source",
                    options=["", "base_registry", "derived_runtime", "generated_runtime"],
                    key="debug_knowledge_concept_source",
                    format_func=lambda value: value or "All sources",
                )
                filtered_knowledge_concepts = _filter_knowledge_concepts_by_scope(
                    _filter_knowledge_concepts_by_focus(
                        _filter_knowledge_concepts(knowledge_concepts, concept_query),
                        concept_focus,
                    ),
                    concept_source_system,
                    concept_source,
                )
                summary_columns = st.columns(6)
                summary_columns[0].metric("Filtered", len(filtered_knowledge_concepts))
                summary_columns[1].metric(
                    "Editable",
                    sum(1 for item in filtered_knowledge_concepts if item.get("editable")),
                )
                summary_columns[2].metric(
                    "Promotable",
                    sum(1 for item in filtered_knowledge_concepts if len(item.get("linked_canonical_concepts") or []) == 1),
                )
                summary_columns[3].metric(
                    "With context",
                    sum(1 for item in filtered_knowledge_concepts if int(item.get("field_context_count", 0) or 0) > 0),
                )
                summary_columns[4].metric(
                    "Linked PII",
                    sum(1 for item in filtered_knowledge_concepts if bool((item.get("linked_privacy") or {}).get("is_pii"))),
                )
                summary_columns[5].metric(
                    "Linked GDPR special",
                    sum(
                        1
                        for item in filtered_knowledge_concepts
                        if bool((item.get("linked_privacy") or {}).get("is_gdpr_special_category"))
                    ),
                )
                st.dataframe(_knowledge_concept_registry_rows(filtered_knowledge_concepts), width="stretch", hide_index=True)

                promotable_knowledge_concepts = [
                    item
                    for item in filtered_knowledge_concepts
                    if len(item.get("linked_canonical_concepts") or []) == 1
                ]
                if promotable_knowledge_concepts:
                    promotion_options = {
                        (
                            _knowledge_concept_option_label(item)
                            + f" -> {((item.get('linked_canonical_concepts') or [''])[0])}"
                        ): item["concept_id"]
                        for item in promotable_knowledge_concepts
                    }
                    selected_promotion_labels = st.multiselect(
                        "Bulk promote knowledge concepts to stable glossary",
                        list(promotion_options.keys()),
                        key="debug_selected_bulk_knowledge_promotions",
                        help="Only concepts with exactly one linked canonical concept are available for bulk promotion.",
                    )
                    if st.button(
                        "Promote selected knowledge concepts",
                        width="stretch",
                        key="debug_promote_selected_knowledge_concepts",
                        disabled=not selected_promotion_labels,
                    ):
                        selected_concept_ids = [promotion_options[label] for label in selected_promotion_labels]
                        try:
                            st.session_state["debug_knowledge_promotion_result"] = api_request(
                                "POST",
                                "/knowledge/concepts/promote-to-glossary",
                                json={
                                    "concept_ids": selected_concept_ids,
                                    "changed_by": st.session_state.get("admin_token"),
                                    "note": "Bulk-promoted from Knowledge Concept Registry.",
                                },
                            )
                            st.session_state["debug_knowledge_concepts"] = api_request("GET", "/knowledge/concepts")
                            if st.session_state.get("debug_canonical_concepts") is not None:
                                st.session_state["debug_canonical_concepts"] = api_request("GET", "/knowledge/canonical-concepts")
                            selected_knowledge_concept_id = _normalized_text(
                                ((st.session_state.get("debug_knowledge_concept_detail") or {}).get("concept") or {}).get("concept_id")
                            )
                            if selected_knowledge_concept_id:
                                st.session_state["debug_knowledge_concept_detail"] = api_request(
                                    "GET",
                                    f"/knowledge/concepts/{selected_knowledge_concept_id}",
                                )
                            st.session_state["canonical_glossary_export_bytes"] = api_request_content(
                                "GET",
                                "/knowledge/canonical-glossary/export",
                            )
                            st.session_state["last_action"] = {
                                "level": "success",
                                "message": f"Promoted {len(selected_concept_ids)} selected knowledge concepts into the stable glossary workflow.",
                            }
                        except httpx.HTTPError as error:
                            st.session_state["last_action"] = {
                                "level": "error",
                                "message": f"Bulk knowledge promotion failed: {error}",
                            }
                        st.rerun()

                knowledge_promotion_result = st.session_state.get("debug_knowledge_promotion_result") or {}
                if knowledge_promotion_result.get("results"):
                    st.caption(
                        "Latest knowledge promotion batch: "
                        f"promoted={knowledge_promotion_result.get('promoted_count', 0)}, "
                        f"skipped={knowledge_promotion_result.get('skipped_count', 0)}."
                    )
                    st.dataframe(knowledge_promotion_result.get("results") or [], width="stretch", hide_index=True)

                if filtered_knowledge_concepts:
                    concept_options = {
                        _knowledge_concept_option_label(item): item["concept_id"]
                        for item in filtered_knowledge_concepts
                    }
                    selected_concept_label = st.selectbox(
                        "Knowledge concept detail",
                        list(concept_options.keys()),
                        key="debug_selected_knowledge_concept_label",
                    )
                    selected_concept_id = concept_options[selected_concept_label]
                    try:
                        _ensure_knowledge_concept_detail_loaded(
                            api_request=api_request,
                            selected_concept_id=selected_concept_id,
                        )
                    except httpx.HTTPError as error:
                        st.session_state["last_action"] = {
                            "level": "error",
                            "message": f"Loading knowledge concept detail failed: {error}",
                        }

                    if st.button(
                        _canonical_console_action_label(
                            _normalized_text(
                                ((st.session_state.get("debug_knowledge_concept_detail") or {}).get("concept") or {}).get("concept_id")
                            )
                            == _normalized_text(selected_concept_id),
                            "knowledge concept detail",
                        ),
                        width="stretch",
                        key="debug_load_knowledge_concept_detail",
                    ):
                        try:
                            st.session_state["debug_knowledge_concept_detail"] = api_request(
                                "GET",
                                f"/knowledge/concepts/{selected_concept_id}",
                            )
                            st.session_state["last_action"] = {
                                "level": "success",
                                "message": f"Loaded knowledge concept detail for {selected_concept_id}.",
                            }
                        except httpx.HTTPError as error:
                            st.session_state["last_action"] = {
                                "level": "error",
                                "message": f"Loading knowledge concept detail failed: {error}",
                            }
                        st.rerun()
                else:
                    st.info("No knowledge concepts match the current search.")

        knowledge_concept_detail = st.session_state.get("debug_knowledge_concept_detail")
        if knowledge_concept_detail:
            concept = knowledge_concept_detail.get("concept") or {}
            with st.expander(
                _section_label("Knowledge Concept Detail", _normalized_text(concept.get("concept_id")) or "selected"),
                expanded=False,
            ):
                summary_columns = st.columns(4)
                summary_columns[0].metric("Aliases", int(concept.get("alias_count", 0) or 0))
                summary_columns[1].metric("Field contexts", int(concept.get("field_context_count", 0) or 0))
                summary_columns[2].metric("Linked canonical", int(concept.get("linked_canonical_concept_count", 0) or 0))
                summary_columns[3].metric("Editable", 1 if concept.get("editable") else 0)
                st.caption(
                    f"{concept.get('concept_id') or 'unknown'} | {concept.get('canonical_name') or 'n/a'} | "
                    f"domain={concept.get('domain') or '-'} | source={concept.get('source') or 'derived_runtime'} | "
                    f"editable={'yes' if concept.get('editable') else 'no'}"
                )
                linked_privacy = concept.get("linked_privacy") or {}
                linked_privacy_parts = [
                    f"linked PII={'yes' if linked_privacy.get('is_pii') else 'no'}",
                    f"linked GDPR special={'yes' if linked_privacy.get('is_gdpr_special_category') else 'no'}",
                ]
                if linked_privacy.get("pii_categories"):
                    linked_privacy_parts.append("linked PII tags=" + ", ".join(linked_privacy.get("pii_categories") or []))
                if linked_privacy.get("data_subject_types"):
                    linked_privacy_parts.append(
                        "linked data subjects=" + ", ".join(linked_privacy.get("data_subject_types") or [])
                    )
                st.caption("Linked privacy: " + " | ".join(linked_privacy_parts))
                aliases = concept.get("aliases") or []
            if aliases:
                st.caption("Aliases: " + ", ".join(str(value).strip() for value in aliases if str(value).strip()))
            linked_canonical_concepts = concept.get("linked_canonical_concepts") or []
            if linked_canonical_concepts:
                st.caption("Linked canonical concepts: " + ", ".join(linked_canonical_concepts))
            if concept.get("editable") and knowledge_concept_detail.get("base_record"):
                base_record = knowledge_concept_detail.get("base_record") or {}
                with st.form(key=f"knowledge_base_record_form_{_normalized_text(concept.get('concept_id'))}"):
                    edit_columns = st.columns(2)
                    edit_columns[0].text_input(
                        "English name",
                        value=_normalized_text(base_record.get("english_name")),
                        disabled=True,
                    )
                    domain_value = edit_columns[1].text_input(
                        "Domain",
                        value=_normalized_text(base_record.get("domain")),
                    )
                    serbian_name_value = st.text_input(
                        "Serbian name",
                        value=_normalized_text(base_record.get("serbian_name")),
                    )
                    abbreviations_value = st.text_area(
                        "Abbreviations",
                        value=_normalized_text(base_record.get("abbreviations")),
                        height=80,
                    )
                    alternative_names_value = st.text_area(
                        "Alternative names",
                        value=_normalized_text(base_record.get("alternative_names")),
                        height=100,
                    )
                    value_columns = st.columns(3)
                    data_type_value = value_columns[0].text_input(
                        "Data type",
                        value=_normalized_text(base_record.get("data_type")),
                    )
                    typical_length_value = value_columns[1].text_input(
                        "Typical length",
                        value=_normalized_text(base_record.get("typical_length")),
                    )
                    example_value = value_columns[2].text_input(
                        "Example value",
                        value=_normalized_text(base_record.get("example_value")),
                    )
                    if st.form_submit_button("Save knowledge concept", width="stretch"):
                        try:
                            st.session_state["debug_knowledge_concept_detail"] = api_request(
                                "PUT",
                                f"/knowledge/concepts/{concept.get('concept_id')}/base-record",
                                json={
                                    "domain": domain_value,
                                    "serbian_name": serbian_name_value,
                                    "abbreviations": abbreviations_value,
                                    "alternative_names": alternative_names_value,
                                    "data_type": data_type_value,
                                    "typical_length": typical_length_value,
                                    "example_value": example_value,
                                    "changed_by": st.session_state.get("admin_token"),
                                },
                            )
                            st.session_state["debug_knowledge_concepts"] = api_request("GET", "/knowledge/concepts")
                            st.session_state["knowledge_registry_export_bytes"] = api_request_content(
                                "GET",
                                "/knowledge/base-registry/export",
                            )
                            st.session_state["last_action"] = {
                                "level": "success",
                                "message": f"Saved base knowledge concept {concept.get('concept_id')}.",
                            }
                        except httpx.HTTPError as error:
                            st.session_state["last_action"] = {
                                "level": "error",
                                "message": f"Saving knowledge concept failed: {error}",
                            }
                        st.rerun()
            elif concept.get("editable"):
                st.info("Base knowledge record is not available for this concept yet.")

            single_promotion_columns = st.columns(2)
            selected_target_concept = None
            if linked_canonical_concepts:
                selected_target_concept = single_promotion_columns[0].selectbox(
                    "Promotion target",
                    options=linked_canonical_concepts,
                    key=f"debug_knowledge_concept_promotion_target_{_normalized_text(concept.get('concept_id'))}",
                )
            if single_promotion_columns[1].button(
                "Promote this knowledge concept",
                width="stretch",
                key=f"debug_promote_knowledge_concept_{_normalized_text(concept.get('concept_id'))}",
                disabled=not linked_canonical_concepts,
            ):
                try:
                    st.session_state["debug_knowledge_promotion_result"] = api_request(
                        "POST",
                        "/knowledge/concepts/promote-to-glossary",
                        json={
                            "concept_ids": [concept.get("concept_id")],
                            "target_concept_id": selected_target_concept,
                            "changed_by": st.session_state.get("admin_token"),
                            "note": "Single-concept promotion from Knowledge Concept Detail.",
                        },
                    )
                    st.session_state["debug_knowledge_concepts"] = api_request("GET", "/knowledge/concepts")
                    if st.session_state.get("debug_canonical_concepts") is not None:
                        st.session_state["debug_canonical_concepts"] = api_request("GET", "/knowledge/canonical-concepts")
                    st.session_state["debug_knowledge_concept_detail"] = api_request(
                        "GET",
                        f"/knowledge/concepts/{concept.get('concept_id')}",
                    )
                    st.session_state["canonical_glossary_export_bytes"] = api_request_content(
                        "GET",
                        "/knowledge/canonical-glossary/export",
                    )
                    st.session_state["last_action"] = {
                        "level": "success",
                        "message": f"Submitted knowledge concept {concept.get('concept_id')} for stable glossary promotion.",
                    }
                except httpx.HTTPError as error:
                    st.session_state["last_action"] = {
                        "level": "error",
                        "message": f"Knowledge concept promotion failed: {error}",
                    }
                st.rerun()
            if not linked_canonical_concepts:
                st.caption("This knowledge concept has no linked canonical concept yet, so direct promotion is disabled.")
            field_contexts = knowledge_concept_detail.get("field_contexts") or []
            if field_contexts:
                st.dataframe(field_contexts, width="stretch", hide_index=True)
            else:
                st.info("No field contexts are attached to this knowledge concept.")

    if active_governance_section == "Canonical":
        st.caption("Canonical glossary stewardship with filtered/total concept counts and context coverage.")
        canonical_header_actions = st.columns(1)
        if canonical_header_actions[0].button(
            _canonical_console_action_label(st.session_state.get("debug_canonical_concepts") is not None, "canonical concept registry"),
            width="stretch",
            key="debug_load_canonical_concepts",
        ):
            try:
                st.session_state.pop("debug_canonical_console_manual_clear", None)
                st.session_state["debug_canonical_concepts"] = api_request("GET", "/knowledge/canonical-concepts")
                _refresh_canonical_console_knowledge_state(api_request=api_request)
                _bootstrap_active_overlay_detail(api_request=api_request)
                st.session_state["debug_canonical_console_bootstrapped"] = True
                st.session_state["last_action"] = {"level": "success", "message": "Loaded canonical concept registry."}
            except httpx.HTTPError as error:
                st.session_state["last_action"] = {
                    "level": "error",
                    "message": f"Loading canonical concept registry failed: {error}",
                }
            st.rerun()

        with st.expander("Canonical Glossary", expanded=False):
            canonical_glossary_upload = st.file_uploader(
                "Canonical glossary CSV",
                type=["csv"],
                key="canonical_glossary_file",
                help="Import or export the canonical business concept glossary as CSV.",
            )
            canonical_glossary_columns = st.columns(2)
            if canonical_glossary_columns[0].button(
                "Load canonical glossary export",
                width="stretch",
                key="canonical_console_export_canonical_glossary",
            ):
                try:
                    st.session_state["canonical_glossary_export_bytes"] = api_request_content("GET", "/knowledge/canonical-glossary/export")
                    st.session_state["last_action"] = {"level": "success", "message": "Loaded canonical glossary export."}
                except httpx.HTTPError as error:
                    st.session_state["last_action"] = {"level": "error", "message": f"Canonical glossary export failed: {error}"}
                st.rerun()

            if canonical_glossary_columns[1].button(
                "Import canonical glossary",
                width="stretch",
                key="canonical_console_import_canonical_glossary",
                disabled=canonical_glossary_upload is None,
            ):
                try:
                    st.session_state["debug_canonical_glossary_import"] = api_request(
                        "POST",
                        "/knowledge/canonical-glossary/import",
                        files=upload_file_to_request_files(canonical_glossary_upload),
                    )
                    _refresh_canonical_console_knowledge_state(api_request=api_request)
                    if st.session_state.get("debug_canonical_concepts") is not None:
                        st.session_state["debug_canonical_concepts"] = api_request("GET", "/knowledge/canonical-concepts")
                    st.session_state["last_action"] = {"level": "success", "message": "Imported canonical glossary CSV."}
                except httpx.HTTPError as error:
                    st.session_state["last_action"] = {"level": "error", "message": f"Canonical glossary import failed: {error}"}
                st.rerun()

            canonical_glossary_export_bytes = st.session_state.get("canonical_glossary_export_bytes")
            if canonical_glossary_export_bytes:
                st.download_button(
                    "Download canonical glossary CSV",
                    data=canonical_glossary_export_bytes,
                    file_name="canonical_glossary.csv",
                    mime="text/csv",
                    width="stretch",
                )

            canonical_glossary_import = st.session_state.get("debug_canonical_glossary_import")
            if canonical_glossary_import:
                st.caption(
                    "Canonical glossary import: "
                    f"rows={canonical_glossary_import.get('imported_row_count', 0)}, "
                    f"concepts={canonical_glossary_import.get('canonical_concept_count', 0)}."
                )

        canonical_concepts = st.session_state.get("debug_canonical_concepts") or []
        if canonical_concepts:
            available_source_systems = sorted(
                {
                    _normalized_text(value)
                    for concept in canonical_concepts
                    for value in concept.get("source_systems", [])
                    if _normalized_text(value)
                }
            )
            available_business_domains = sorted(
                {
                    _normalized_text(value)
                    for concept in canonical_concepts
                    for value in concept.get("business_domains", [])
                    if _normalized_text(value)
                }
            )
            _apply_pending_governance_selectbox_value(
                st.session_state,
                pending_key="pending_governance_canonical_source_system",
                target_key="debug_canonical_concept_source_system",
                options=available_source_systems,
            )
            _apply_pending_governance_selectbox_value(
                st.session_state,
                pending_key="pending_governance_canonical_business_domain",
                target_key="debug_canonical_concept_business_domain",
                options=available_business_domains,
            )
            filter_columns = st.columns(4)
            concept_query = filter_columns[0].text_input(
                "Canonical concept search",
                value=st.session_state.get("debug_canonical_concept_query", ""),
                key="debug_canonical_concept_query",
                placeholder="Search by concept id, display name, alias, source system, domain, entity, or source",
            )
            concept_focus = filter_columns[1].selectbox(
                "Concept focus",
                options=["all", "active_overlay", "overlay_only", "in_use", "with_context", "base_only"],
                index=0,
                key="debug_canonical_concept_focus",
                format_func=lambda value: {
                    "all": "All concepts",
                    "active_overlay": "With active overlay aliases",
                    "overlay_only": "Overlay-only concepts",
                    "in_use": "Used in catalog",
                    "with_context": "With field contexts",
                    "base_only": "Base-only concepts",
                }.get(value, value),
            )
            concept_source_system = filter_columns[2].selectbox(
                "Source system",
                options=[""] + available_source_systems,
                key="debug_canonical_concept_source_system",
                format_func=lambda value: value or "All source systems",
            )
            concept_business_domain = filter_columns[3].selectbox(
                "Business domain",
                options=[""] + available_business_domains,
                key="debug_canonical_concept_business_domain",
                format_func=lambda value: value or "All business domains",
            )
            filtered_concepts = _filter_canonical_concepts_by_scope(
                _filter_canonical_concepts_by_focus(
                    _filter_canonical_concepts(canonical_concepts, concept_query),
                    concept_focus,
                ),
                concept_source_system,
                concept_business_domain,
            )
            summary_columns = st.columns(6)
            summary_columns[0].metric("Filtered", len(filtered_concepts))
            summary_columns[1].metric("Total", len(canonical_concepts))
            summary_columns[2].metric(
                "With active overlay",
                sum(1 for item in filtered_concepts if int(item.get("active_overlay_entry_count", 0) or 0) > 0),
            )
            summary_columns[3].metric(
                "With context",
                sum(1 for item in filtered_concepts if int(item.get("field_context_count", 0) or 0) > 0),
            )
            summary_columns[4].metric(
                "PII",
                sum(1 for item in filtered_concepts if bool((item.get("privacy") or {}).get("is_pii"))),
            )
            summary_columns[5].metric(
                "GDPR special",
                sum(1 for item in filtered_concepts if bool((item.get("privacy") or {}).get("is_gdpr_special_category"))),
            )
            st.dataframe(_canonical_concept_registry_rows(filtered_concepts), width="stretch", hide_index=True)

            if filtered_concepts:
                concept_options = {_canonical_concept_option_label(item): item["concept_id"] for item in filtered_concepts}
                pending_concept_label = _pending_canonical_concept_label(filtered_concepts, st.session_state)
                if pending_concept_label and st.session_state.get("debug_selected_canonical_concept_label") != pending_concept_label:
                    st.session_state["debug_selected_canonical_concept_label"] = pending_concept_label
                preferred_concept_label = _preferred_canonical_concept_label(
                    filtered_concepts,
                    st.session_state.get("debug_knowledge_stewardship_items"),
                    active_overlay_id=(st.session_state.get("debug_knowledge_runtime") or {}).get("active_overlay_id"),
                    current_label=st.session_state.get("debug_selected_canonical_concept_label"),
                )
                if preferred_concept_label and st.session_state.get("debug_selected_canonical_concept_label") != preferred_concept_label:
                    st.session_state["debug_selected_canonical_concept_label"] = preferred_concept_label
                selected_concept_label = st.selectbox(
                    "Canonical concept detail",
                    list(concept_options.keys()),
                    key="debug_selected_canonical_concept_label",
                )
                selected_concept_id = concept_options[selected_concept_label]
                try:
                    _ensure_canonical_concept_detail_loaded(
                        api_request=api_request,
                        selected_concept_id=selected_concept_id,
                    )
                except httpx.HTTPError as error:
                    st.session_state["last_action"] = {
                        "level": "error",
                        "message": f"Loading canonical concept detail failed: {error}",
                    }

                if st.button(
                    _canonical_console_action_label(
                        _normalized_text(
                            ((st.session_state.get("debug_canonical_concept_detail") or {}).get("concept") or {}).get("concept_id")
                        )
                        == _normalized_text(selected_concept_id),
                        "canonical concept detail",
                    ),
                    width="stretch",
                    key="debug_load_canonical_concept_detail",
                ):
                    try:
                        st.session_state["debug_canonical_concept_detail"] = api_request(
                            "GET",
                            f"/knowledge/canonical-concepts/{selected_concept_id}",
                        )
                        st.session_state["debug_knowledge_audit_logs"] = api_request("GET", "/knowledge/audit")
                        st.session_state["last_action"] = {
                            "level": "success",
                            "message": f"Loaded canonical concept detail for {selected_concept_id}.",
                        }
                    except httpx.HTTPError as error:
                        st.session_state["last_action"] = {
                            "level": "error",
                            "message": f"Loading canonical concept detail failed: {error}",
                        }
                    st.rerun()
            else:
                st.info("No canonical concepts match the current search.")

        canonical_concept_detail = st.session_state.get("debug_canonical_concept_detail")
        if canonical_concept_detail:
            concept = canonical_concept_detail.get("concept") or {}
            concept_pending_rows = _canonical_gap_pending_rows_for_concept(
                concept,
                st.session_state.get("canonical_gap_candidates"),
                st.session_state.get("canonical_gap_suggestions"),
                st.session_state.get("debug_canonical_gap_console_states"),
                st.session_state.get("debug_canonical_gap_proposal_states"),
                _canonical_gap_stewardship_item_map(st.session_state.get("debug_knowledge_stewardship_items")),
            )
            concept_overlay_promotion_items = _overlay_promotion_items_for_concept(
                concept,
                st.session_state.get("debug_knowledge_stewardship_items"),
            )
            concept_overlay_promotion_rows = _overlay_promotion_rows_for_concept(
                concept,
                st.session_state.get("debug_knowledge_stewardship_items"),
            )
            with st.expander(
                _section_label("Canonical Concept Detail", _normalized_text(concept.get("concept_id")) or "selected"),
                expanded=False,
            ):
                summary_columns = st.columns(5)
                summary_columns[0].metric("Usage", int(concept.get("usage_count", 0) or 0))
                summary_columns[1].metric("Field contexts", int(concept.get("field_context_count", 0) or 0))
                summary_columns[2].metric("Active overlay aliases", int(concept.get("active_overlay_entry_count", 0) or 0))
                summary_columns[3].metric("Pending proposals", len(concept_pending_rows))
                summary_columns[4].metric("Overlay promotions", len(concept_overlay_promotion_items))
                st.caption(
                    f"{concept.get('concept_id') or 'unknown'} | {concept.get('display_name') or 'n/a'} | "
                    f"entity={concept.get('entity') or '-'} | attribute={concept.get('attribute') or '-'} | "
                    f"source={concept.get('source') or 'base'}"
                )
                if concept.get("description"):
                    st.write(concept.get("description"))
                privacy = concept.get("privacy") or {}
                privacy_parts = [
                    f"PII={'yes' if privacy.get('is_pii') else 'no'}",
                    f"GDPR special={'yes' if privacy.get('is_gdpr_special_category') else 'no'}",
                ]
                if privacy.get("pii_categories"):
                    privacy_parts.append("PII tags=" + ", ".join(privacy.get("pii_categories") or []))
                if privacy.get("data_subject_types"):
                    privacy_parts.append("Data subjects=" + ", ".join(privacy.get("data_subject_types") or []))
                st.caption("Privacy: " + " | ".join(privacy_parts))
                if concept.get("base_aliases") or concept.get("active_overlay_aliases"):
                    st.caption(
                        "Aliases: "
                        + " | ".join(
                            part
                            for part in [
                                (
                                    "base=" + ", ".join(concept.get("base_aliases") or [])
                                    if concept.get("base_aliases")
                                    else ""
                                ),
                                (
                                    "active_overlay=" + ", ".join(concept.get("active_overlay_aliases") or [])
                                    if concept.get("active_overlay_aliases")
                                    else ""
                                ),
                            ]
                            if part
                        )
                    )
                if concept.get("source_systems") or concept.get("business_domains"):
                    st.caption(
                        "Discovery facets: "
                        + " | ".join(
                            part
                            for part in [
                                (
                                    "source_systems=" + ", ".join(concept.get("source_systems") or [])
                                    if concept.get("source_systems")
                                    else ""
                                ),
                                (
                                    "business_domains=" + ", ".join(concept.get("business_domains") or [])
                                    if concept.get("business_domains")
                                    else ""
                                ),
                            ]
                            if part
                        )
                    )
                st.caption(f"Alias count: {int(concept.get('alias_count', 0) or 0)}")

                concept_id = _normalized_text(concept.get("concept_id")) or "unknown"
                concept_governance_item = _concept_governance_item_map(
                    st.session_state.get("debug_knowledge_stewardship_items")
                ).get(concept_id) or {}
                concept_governance_item_id = concept_governance_item.get("item_id")
                st.write("**Concept governance profile**")
                st.caption(
                    "Thin ownership metadata attached to the canonical concept. External governance systems remain the process source of truth; Semantra stores the current business owner and data steward for reference."
                )
                governance_columns = st.columns(2)
                governance_owner = governance_columns[0].text_input(
                    "External business owner",
                    value=_normalized_text(concept_governance_item.get("owner")),
                    key=f"debug_concept_governance_owner_{concept_id}",
                    placeholder="Example: order-to-cash lead",
                )
                governance_assignee = governance_columns[1].text_input(
                    "External data steward",
                    value=_normalized_text(concept_governance_item.get("assignee")),
                    key=f"debug_concept_governance_assignee_{concept_id}",
                    placeholder="Example: semantic-model steward",
                )
                governance_note = st.text_input(
                    "External governance note",
                    value=_normalized_text(concept_governance_item.get("review_note")),
                    key=f"debug_concept_governance_note_{concept_id}",
                    placeholder="Optional note copied from the external governance process.",
                )
                desired_governance_status = (
                    "approved"
                    if any(_normalized_text(value) for value in (governance_owner, governance_assignee, governance_note))
                    else "new"
                )
                save_governance_disabled = (
                    not concept_governance_item_id
                    and not any(_normalized_text(value) for value in (governance_owner, governance_assignee, governance_note))
                ) or (
                    _normalized_text(concept_governance_item.get("owner")) == _normalized_text(governance_owner)
                    and _normalized_text(concept_governance_item.get("assignee")) == _normalized_text(governance_assignee)
                    and _normalized_text(concept_governance_item.get("review_note")) == _normalized_text(governance_note)
                    and (_normalized_text(concept_governance_item.get("status")) or "new") == desired_governance_status
                )
                if st.button(
                    "Save governance metadata",
                    width="stretch",
                    key=f"debug_save_concept_governance_{concept_id}",
                    disabled=save_governance_disabled,
                ):
                    try:
                        api_request(
                            "POST",
                            "/knowledge/stewardship-items",
                            json=_concept_governance_item_request(
                                concept,
                                owner=governance_owner,
                                assignee=governance_assignee,
                                review_note=governance_note,
                                changed_by=st.session_state.get("admin_token"),
                            ),
                        )
                        st.session_state["debug_knowledge_stewardship_items"] = api_request("GET", "/knowledge/stewardship-items")
                        st.session_state["debug_knowledge_audit_logs"] = api_request("GET", "/knowledge/audit")
                        st.session_state["last_action"] = {
                            "level": "info",
                            "message": (
                                f"Saved concept governance metadata for {_normalized_text(concept.get('concept_id')) or 'selected concept'} "
                                "using the existing stewardship store."
                            ),
                        }
                    except httpx.HTTPError as error:
                        st.session_state["last_action"] = {
                            "level": "error",
                            "message": f"Saving concept governance metadata failed: {error}",
                        }
                    st.rerun()
                if concept_governance_item_id:
                    st.caption(
                        "Current governance record: "
                        f"status={_normalized_text(concept_governance_item.get('status')) or 'new'} | "
                        f"updated_at={_normalized_text(concept_governance_item.get('updated_at') or concept_governance_item.get('created_at')) or 'n/a'}"
                    )

                st.write("**Quick link correction patch (overlay)**")
                st.caption(
                    "Create a one-row overlay patch from this concept detail to correct knowledge/vendor links without modifying the base glossary directly."
                )
                concept_id = _normalized_text(concept.get("concept_id"))
                concept_systems = [
                    _normalized_text(value)
                    for value in concept.get("source_systems", [])
                    if _normalized_text(value)
                ]
                concept_domains = [
                    _normalized_text(value)
                    for value in concept.get("business_domains", [])
                    if _normalized_text(value)
                ]
                patch_columns = st.columns(3)
                patch_entry_type = patch_columns[0].selectbox(
                    "Patch type",
                    options=["concept_alias", "field_alias", "synonym"],
                    key=f"debug_patch_entry_type_{concept_id}",
                    help="Use concept_alias for canonical/vendor linking, field_alias for source field naming, synonym for generic terminology normalization.",
                )
                patch_source_system = patch_columns[1].selectbox(
                    "Source system",
                    options=[""] + sorted(set(concept_systems + ["SAP", "QuickBooks"])),
                    key=f"debug_patch_source_system_{concept_id}",
                    format_func=lambda value: value or "(optional)",
                )
                patch_domain = patch_columns[2].selectbox(
                    "Business domain",
                    options=[""] + sorted(set(concept_domains)),
                    key=f"debug_patch_domain_{concept_id}",
                    format_func=lambda value: value or "(optional)",
                )
                patch_alias = st.text_input(
                    "Alias / field token",
                    value="",
                    key=f"debug_patch_alias_{concept_id}",
                    placeholder="Examples: ZWELS, payment office, LIKP.VSART",
                    help="For vendor-specific field links, prefer a precise token such as TABLE.FIELD or field code.",
                )
                patch_note = st.text_input(
                    "Patch note",
                    value="",
                    key=f"debug_patch_note_{concept_id}",
                    placeholder="Why this link is valid (module/table semantics).",
                )
                patch_overlay_name = st.text_input(
                    "Overlay patch version name",
                    value=f"canonical-console-{concept_id}-patch-{datetime.utcnow().strftime('%Y%m%d%H%M%S')}",
                    key=f"debug_patch_overlay_name_{concept_id}",
                )
                create_patch_disabled = not _normalized_text(patch_alias)
                if st.button(
                    "Create overlay patch",
                    width="stretch",
                    key=f"debug_create_overlay_patch_{concept_id}",
                    disabled=create_patch_disabled,
                ):
                    try:
                        patch_bytes = _single_overlay_patch_bytes(
                            entry_type=patch_entry_type,
                            canonical_term=concept_id,
                            alias=_normalized_text(patch_alias),
                            domain=_normalized_text(patch_domain) or None,
                            source_system=_normalized_text(patch_source_system) or None,
                            note=_normalized_text(patch_note) or None,
                        )
                        created_patch = api_request(
                            "POST",
                            "/knowledge/overlays",
                            files={
                                "file": (
                                    f"{_normalized_text(patch_overlay_name) or 'canonical_console_patch'}.csv",
                                    patch_bytes,
                                    "text/csv",
                                )
                            },
                            data={
                                "name": _normalized_text(patch_overlay_name) or f"canonical-console-{concept_id}-patch",
                                "created_by": _normalized_text(st.session_state.get("admin_token")) or "canonical-console",
                            },
                        )
                        _refresh_canonical_console_knowledge_state(api_request=api_request)
                        st.session_state["debug_knowledge_overlays"] = api_request("GET", "/knowledge/overlays")
                        overlay_version = created_patch.get("version") or {}
                        overlay_id = overlay_version.get("overlay_id")
                        if overlay_id is not None:
                            st.session_state["debug_selected_overlay_version"] = (
                                f"#{overlay_id} | {overlay_version.get('name')} | {overlay_version.get('status')}"
                            )
                            st.session_state["debug_selected_knowledge_overlay"] = api_request(
                                "GET",
                                f"/knowledge/overlays/{overlay_id}",
                            )
                        st.session_state["last_action"] = {
                            "level": "success",
                            "message": (
                                f"Created overlay patch '{overlay_version.get('name')}' with one {patch_entry_type} entry. "
                                "Activate it in Overlay Management when ready."
                            ),
                        }
                    except httpx.HTTPError as error:
                        st.session_state["last_action"] = {
                            "level": "error",
                            "message": f"Creating overlay patch failed: {error}",
                        }
                    st.rerun()

                st.write("**Vendor context patch (overlay)**")
                st.caption(
                    "Capture system/object/field context in a dedicated patch row. This creates a field_alias overlay entry with an explicit context payload in the note."
                )
                context_columns = st.columns(3)
                context_system = context_columns[0].selectbox(
                    "Context system",
                    options=sorted(set(["SAP", "QuickBooks"] + concept_systems)),
                    key=f"debug_context_patch_system_{concept_id}",
                )
                context_object = context_columns[1].text_input(
                    "Object / table",
                    value="",
                    key=f"debug_context_patch_object_{concept_id}",
                    placeholder="Examples: LIKP, VBAK, Invoice",
                )
                context_field = context_columns[2].text_input(
                    "Field",
                    value="",
                    key=f"debug_context_patch_field_{concept_id}",
                    placeholder="Examples: VSART, TAXM1, DueDate",
                )
                context_desc_columns = st.columns(2)
                context_object_description = context_desc_columns[0].text_input(
                    "Object description",
                    value="",
                    key=f"debug_context_patch_object_desc_{concept_id}",
                    placeholder="Optional object/table description",
                )
                context_field_description = context_desc_columns[1].text_input(
                    "Field description",
                    value="",
                    key=f"debug_context_patch_field_desc_{concept_id}",
                    placeholder="Optional field business description",
                )
                context_domain = st.selectbox(
                    "Context business domain",
                    options=[""] + sorted(set(concept_domains)),
                    key=f"debug_context_patch_domain_{concept_id}",
                    format_func=lambda value: value or "(optional)",
                )
                context_note = st.text_input(
                    "Context patch note",
                    value="",
                    key=f"debug_context_patch_note_{concept_id}",
                    placeholder="Why this vendor context should be linked to the concept.",
                )
                context_overlay_name = st.text_input(
                    "Context overlay patch version name",
                    value=f"canonical-console-{concept_id}-context-{datetime.utcnow().strftime('%Y%m%d%H%M%S')}",
                    key=f"debug_context_patch_overlay_name_{concept_id}",
                )
                context_alias = _context_alias_token(context_object, context_field)
                create_context_patch_disabled = not _normalized_text(context_alias)
                if st.button(
                    "Create context patch",
                    width="stretch",
                    key=f"debug_create_context_patch_{concept_id}",
                    disabled=create_context_patch_disabled,
                ):
                    try:
                        context_patch_bytes = _single_overlay_patch_bytes(
                            entry_type="field_alias",
                            canonical_term=concept_id,
                            alias=_normalized_text(context_alias),
                            domain=_normalized_text(context_domain) or None,
                            source_system=_normalized_text(context_system) or None,
                            note=_context_note_payload(
                                system=context_system,
                                object_name=context_object,
                                field_name=context_field,
                                object_description=context_object_description,
                                field_description=context_field_description,
                                note=context_note,
                            ),
                        )
                        created_context_patch = api_request(
                            "POST",
                            "/knowledge/overlays",
                            files={
                                "file": (
                                    f"{_normalized_text(context_overlay_name) or 'canonical_console_context_patch'}.csv",
                                    context_patch_bytes,
                                    "text/csv",
                                )
                            },
                            data={
                                "name": _normalized_text(context_overlay_name) or f"canonical-console-{concept_id}-context",
                                "created_by": _normalized_text(st.session_state.get("admin_token")) or "canonical-console",
                            },
                        )
                        _refresh_canonical_console_knowledge_state(api_request=api_request)
                        st.session_state["debug_knowledge_overlays"] = api_request("GET", "/knowledge/overlays")
                        context_overlay_version = created_context_patch.get("version") or {}
                        context_overlay_id = context_overlay_version.get("overlay_id")
                        if context_overlay_id is not None:
                            st.session_state["debug_selected_overlay_version"] = (
                                f"#{context_overlay_id} | {context_overlay_version.get('name')} | {context_overlay_version.get('status')}"
                            )
                            st.session_state["debug_selected_knowledge_overlay"] = api_request(
                                "GET",
                                f"/knowledge/overlays/{context_overlay_id}",
                            )
                        st.session_state["last_action"] = {
                            "level": "success",
                            "message": (
                                f"Created context overlay patch '{context_overlay_version.get('name')}' "
                                f"for {_normalized_text(context_system)} {context_alias}. Activate it in Overlay Management when ready."
                            ),
                        }
                    except httpx.HTTPError as error:
                        st.session_state["last_action"] = {
                            "level": "error",
                            "message": f"Creating context patch failed: {error}",
                        }
                    st.rerun()

                if concept_pending_rows:
                    st.write("**Pending queue proposals for this concept**")
                    st.caption(
                        "Current active gap suggestions that would extend this concept through the console approve flow."
                    )
                    st.dataframe(concept_pending_rows, width="stretch", hide_index=True)
                else:
                    st.caption("No active queue proposals currently target this concept.")

                if concept_overlay_promotion_items:
                    st.write("**Overlay promotion stewardship**")
                    st.caption(
                        "Durable overlay alias candidates already linked to this canonical concept, including execution-ready items."
                    )
                    st.dataframe(concept_overlay_promotion_rows, width="stretch", hide_index=True)
                    concept_promotion_options = {
                        _overlay_promotion_item_record_label(item): item for item in concept_overlay_promotion_items
                    }
                    selected_concept_promotion_label = st.selectbox(
                        "Concept overlay promotion detail",
                        list(concept_promotion_options.keys()),
                        key=f"debug_concept_overlay_promotion_{_normalized_text(concept.get('concept_id')) or 'unknown'}",
                    )
                    selected_concept_promotion_item = concept_promotion_options[selected_concept_promotion_label]
                    st.caption(
                        f"Promotion status: {_canonical_gap_proposal_state_label(_overlay_promotion_status(selected_concept_promotion_item))} | "
                        f"alias={_normalized_text(selected_concept_promotion_item.get('source')) or 'n/a'} | "
                        f"source_system={_normalized_text(selected_concept_promotion_item.get('source_system')) or 'n/a'} | "
                        f"business_domain={_normalized_text(selected_concept_promotion_item.get('business_domain')) or 'n/a'}"
                    )
                    if _normalized_text(selected_concept_promotion_item.get("review_note")):
                        st.caption(f"Review note: {_normalized_text(selected_concept_promotion_item.get('review_note'))}")
                    if st.button(
                        "Promote selected concept candidate",
                        width="stretch",
                        key=f"debug_execute_concept_overlay_promotion_{selected_concept_promotion_item.get('item_id')}",
                        disabled=not _overlay_promotion_can_execute(selected_concept_promotion_item),
                    ):
                        try:
                            promotion_result = api_request(
                                "POST",
                                f"/knowledge/stewardship-items/{selected_concept_promotion_item.get('item_id')}/promote-to-glossary",
                                json=_overlay_promotion_execution_request(
                                    st.session_state.get("admin_token"),
                                    note="Promoted from Canonical Console concept detail.",
                                ),
                            )
                            _refresh_canonical_console_knowledge_state(api_request=api_request)
                            st.session_state["debug_canonical_concept_detail"] = api_request(
                                "GET",
                                f"/knowledge/canonical-concepts/{_normalized_text(concept.get('concept_id'))}",
                            )
                            st.session_state["canonical_glossary_export_bytes"] = api_request_content(
                                "GET",
                                "/knowledge/canonical-glossary/export",
                            )
                            st.session_state["last_action"] = {
                                "level": "success",
                                "message": (
                                    "Promoted selected concept-linked overlay alias into the stable canonical glossary."
                                    if promotion_result.get("alias_added", True) or promotion_result.get("concept_created", False)
                                    else "Selected concept-linked overlay alias was already present in the stable canonical glossary; stewardship item marked as promoted."
                                ),
                            }
                        except httpx.HTTPError as error:
                            st.session_state["last_action"] = {
                                "level": "error",
                                "message": f"Concept-linked stable glossary promotion failed: {error}",
                            }
                        st.rerun()
                    if _overlay_promotion_status(selected_concept_promotion_item) == "ready_for_approval":
                        st.caption("This concept-linked promotion item can now write into the stable canonical glossary.")
                    elif _overlay_promotion_status(selected_concept_promotion_item) == "promoted":
                        st.caption("This concept-linked overlay alias has already been promoted.")
                else:
                    st.caption("No overlay promotion stewardship items currently target this concept.")

                context_patch_rows = _context_patch_ingest_rows(canonical_concept_detail)
                if context_patch_rows:
                    st.write("**Context patch ingest status**")
                    context_patch_columns = st.columns(3)
                    context_patch_columns[0].metric("Context patch entries", len(context_patch_rows))
                    context_patch_columns[1].metric(
                        "Ingested",
                        sum(1 for row in context_patch_rows if row.get("ingested") == "yes"),
                    )
                    context_patch_columns[2].metric(
                        "Not ingested",
                        sum(1 for row in context_patch_rows if row.get("ingested") == "no"),
                    )
                    st.caption(
                        "Shows whether each active overlay context patch note is currently materialized into runtime field context."
                    )
                    st.dataframe(context_patch_rows, width="stretch", hide_index=True)

                if canonical_concept_detail.get("field_contexts"):
                    st.write("**Field contexts**")
                    st.dataframe(canonical_concept_detail.get("field_contexts"), width="stretch", hide_index=True)
                if canonical_concept_detail.get("active_overlay_entries"):
                    st.write("**Active overlay entries**")
                    st.dataframe(canonical_concept_detail.get("active_overlay_entries"), width="stretch", hide_index=True)
                if canonical_concept_detail.get("integrations"):
                    st.write("**Catalog usage**")
                    st.dataframe(canonical_concept_detail.get("integrations"), width="stretch", hide_index=True)
                if canonical_concept_detail.get("audit_entries"):
                    st.write("**Knowledge audit references**")
                    st.dataframe(canonical_concept_detail.get("audit_entries"), width="stretch", hide_index=True)

    if active_governance_section == "Stewardship":
        stewardship_header_actions = st.columns(1)
        if stewardship_header_actions[0].button(
            "Clear governance state",
            width="stretch",
            key="debug_clear_canonical_concepts",
        ):
            for key in (
                "debug_canonical_concepts",
                "debug_selected_canonical_concept_label",
                "debug_canonical_concept_detail",
                "debug_knowledge_concepts",
                "debug_selected_knowledge_concept_label",
                "debug_knowledge_concept_detail",
                "debug_knowledge_promotion_result",
                "debug_selected_canonical_gap_label",
                "debug_canonical_gap_console_states",
                "debug_knowledge_audit_logs",
                "debug_knowledge_runtime",
                "debug_knowledge_overlays",
                "debug_selected_knowledge_overlay",
                "debug_knowledge_validation",
                "debug_knowledge_created",
                "debug_knowledge_registry_import",
                "knowledge_registry_export_bytes",
                "debug_canonical_glossary_import",
                "canonical_glossary_export_bytes",
                "debug_canonical_console_bootstrapped",
                "pending_governance_canonical_concept_id",
                "pending_governance_canonical_source_system",
                "pending_governance_canonical_business_domain",
                "pending_governance_gap_source_filter",
                "governance_focus_sources",
            ):
                st.session_state.pop(key, None)
            st.session_state["debug_canonical_console_manual_clear"] = True
            st.session_state["last_action"] = {"level": "info", "message": "Cleared canonical console state."}
            st.rerun()

        concept_governance_items = _concept_governance_item_map(st.session_state.get("debug_knowledge_stewardship_items"))
        st.write("**Concept governance assignments**")
        st.caption(
            "Reference-only ownership metadata attached to canonical concepts. These records mirror who owns and stewards a concept without moving approval workflows into Semantra."
        )
        with st.expander("Bulk import concept governance", expanded=False):
            st.download_button(
                "Download concept governance CSV template",
                data=_concept_governance_template_csv_bytes(),
                file_name="concept_governance_template.csv",
                mime="text/csv",
                width="stretch",
            )
            governance_import_file = st.file_uploader(
                "Concept governance CSV or XLSX",
                type=["csv", "xlsx"],
                key="concept_governance_import_file",
                help=(
                    "Accepted columns: concept_id plus any of business_owner/external_business_owner, "
                    "data_steward/external_data_steward, governance_note/external_governance_note."
                ),
            )
            st.caption(
                "Use this to mirror core ownership metadata from an external governance system. Rows without concept_id or without any ownership metadata are skipped."
            )
            if governance_import_file is not None:
                try:
                    governance_import_rows = _concept_governance_import_payloads(
                        _concept_governance_import_rows(governance_import_file)
                    )
                    st.caption(f"Parsed {len(governance_import_rows)} import-ready governance rows.")
                    if governance_import_rows:
                        st.dataframe(governance_import_rows[:25], width="stretch", hide_index=True)
                        if len(governance_import_rows) > 25:
                            st.caption(f"Showing first 25 of {len(governance_import_rows)} rows.")
                    import_disabled = not governance_import_rows
                except Exception as error:
                    governance_import_rows = []
                    import_disabled = True
                    st.error(f"Parsing concept governance import failed: {error}")
            else:
                governance_import_rows = []
                import_disabled = True
            if st.button(
                "Import concept governance metadata",
                width="stretch",
                key="debug_import_concept_governance_metadata",
                disabled=import_disabled,
            ):
                saved_count = 0
                failed_rows: list[str] = []
                for row in governance_import_rows:
                    try:
                        api_request(
                            "POST",
                            "/knowledge/stewardship-items",
                            json=_concept_governance_item_request(
                                {"concept_id": row.get("concept_id")},
                                owner=row.get("owner"),
                                assignee=row.get("assignee"),
                                review_note=row.get("review_note"),
                                changed_by=st.session_state.get("admin_token"),
                            ),
                        )
                        saved_count += 1
                    except httpx.HTTPError as error:
                        failed_rows.append(f"{row.get('concept_id')}: {error}")
                try:
                    st.session_state["debug_knowledge_stewardship_items"] = api_request("GET", "/knowledge/stewardship-items")
                    st.session_state["debug_knowledge_audit_logs"] = api_request("GET", "/knowledge/audit")
                except httpx.HTTPError:
                    pass
                if failed_rows:
                    st.session_state["last_action"] = {
                        "level": "warning",
                        "message": (
                            f"Imported {saved_count} concept governance rows; {len(failed_rows)} failed. "
                            f"First failure: {failed_rows[0]}"
                        ),
                    }
                else:
                    st.session_state["last_action"] = {
                        "level": "success",
                        "message": f"Imported {saved_count} concept governance rows from external metadata file.",
                    }
                st.rerun()
        if concept_governance_items:
            governance_rows = _concept_governance_rows(st.session_state.get("debug_knowledge_stewardship_items"))
            governance_metric_columns = st.columns(4)
            governance_metric_columns[0].metric("Governed concepts", len(governance_rows))
            governance_metric_columns[1].metric(
                "With external owner",
                sum(1 for row in governance_rows if _normalized_text(row.get("business_owner"))),
            )
            governance_metric_columns[2].metric(
                "With external steward",
                sum(1 for row in governance_rows if _normalized_text(row.get("data_steward"))),
            )
            governance_metric_columns[3].metric(
                "With note",
                sum(1 for row in governance_rows if _normalized_text(row.get("governance_note"))),
            )
            st.dataframe(governance_rows, width="stretch", hide_index=True)
            governance_options = {
                _concept_governance_option_label(item): item for item in concept_governance_items.values()
            }
            selected_governance_label = st.selectbox(
                "Concept governance detail",
                list(governance_options.keys()),
                key="debug_selected_concept_governance_label",
            )
            selected_governance_item = governance_options[selected_governance_label]
            st.caption(
                f"Selected concept={_normalized_text(selected_governance_item.get('concept_id')) or 'unknown'} | "
                f"status={_normalized_text(selected_governance_item.get('status')) or 'new'}"
            )
            if _normalized_text(selected_governance_item.get("review_note")):
                st.caption(f"Governance note: {_normalized_text(selected_governance_item.get('review_note'))}")
            if st.button(
                "Open concept in Canonical",
                width="stretch",
                key=f"debug_open_concept_governance_{_normalized_text(selected_governance_item.get('concept_id')) or 'unknown'}",
            ):
                st.session_state["pending_governance_section"] = "Canonical"
                st.session_state["pending_governance_canonical_concept_id"] = _normalized_text(
                    selected_governance_item.get("concept_id")
                )
                st.rerun()
        else:
            st.caption("No concept governance assignments have been recorded yet.")

        st.write("**Canonical gap review queue**")
        st.caption(
            "Mirror of the Review tab canonical gap state. Console can approve, ignore, or reject cached suggestions while showing direct audit context for the selected gap."
        )
        canonical_gap_candidates = st.session_state.get("canonical_gap_candidates") or []
        canonical_gap_suggestions = st.session_state.get("canonical_gap_suggestions") or {}
        canonical_gap_console_states = st.session_state.setdefault("debug_canonical_gap_console_states", {})
        canonical_gap_proposal_states = st.session_state.setdefault("debug_canonical_gap_proposal_states", {})
        canonical_gap_stewardship_items = _canonical_gap_stewardship_item_map(
            st.session_state.get("debug_knowledge_stewardship_items")
        )
        if canonical_gap_candidates:
            repeated_gap_rows = _canonical_gap_repeat_summary_rows(
                canonical_gap_candidates,
                canonical_gap_suggestions,
                canonical_gap_console_states,
                canonical_gap_proposal_states,
                canonical_gap_stewardship_items,
            )
            queue_columns = st.columns(5)
            queue_columns[0].metric("Candidates", len(canonical_gap_candidates))
            queue_columns[1].metric(
                "Ignored",
                sum(
                    1
                    for index, candidate in enumerate(canonical_gap_candidates)
                    if _canonical_gap_console_state(
                        _canonical_gap_candidate_key(index, candidate),
                        canonical_gap_console_states,
                        canonical_gap_stewardship_items.get(_canonical_gap_candidate_key(index, candidate)),
                    )
                    == "ignored"
                ),
            )
            queue_columns[2].metric(
                "Rejected",
                sum(
                    1
                    for index, candidate in enumerate(canonical_gap_candidates)
                    if _canonical_gap_console_state(
                        _canonical_gap_candidate_key(index, candidate),
                        canonical_gap_console_states,
                        canonical_gap_stewardship_items.get(_canonical_gap_candidate_key(index, candidate)),
                    )
                    == "rejected"
                ),
            )
            queue_columns[3].metric(
                "With suggestion",
                sum(
                    1
                    for index, candidate in enumerate(canonical_gap_candidates)
                    if canonical_gap_suggestions.get(_canonical_gap_candidate_key(index, candidate))
                ),
            )
            queue_columns[4].metric(
                "Ready for approval",
                sum(
                    1
                    for index, candidate in enumerate(canonical_gap_candidates)
                    if _canonical_gap_can_approve(
                        canonical_gap_suggestions.get(_canonical_gap_candidate_key(index, candidate)),
                        _canonical_gap_console_state(
                            _canonical_gap_candidate_key(index, candidate),
                            canonical_gap_console_states,
                            canonical_gap_stewardship_items.get(_canonical_gap_candidate_key(index, candidate)),
                        ),
                        _canonical_gap_proposal_state(
                            _canonical_gap_candidate_key(index, candidate),
                            canonical_gap_proposal_states,
                            canonical_gap_stewardship_items.get(_canonical_gap_candidate_key(index, candidate)),
                        ),
                    )
                ),
            )
            if repeated_gap_rows:
                with st.expander(
                    _section_label("Repeated gap signals", f"{len(repeated_gap_rows)} patterns"),
                    expanded=False,
                ):
                    st.caption(
                        "Aggregated target and suggested-concept patterns seen across the current queue and durable stewardship history. "
                        "Use this to spot glossary or overlay work that can eliminate multiple gap reviews at once."
                    )
                    st.dataframe(repeated_gap_rows, width="stretch", hide_index=True)
            available_stewardship_statuses = sorted(
                {
                    _canonical_gap_effective_status(
                        _canonical_gap_candidate_key(index, candidate),
                        canonical_gap_console_states,
                        canonical_gap_proposal_states,
                        canonical_gap_stewardship_items.get(_canonical_gap_candidate_key(index, candidate)),
                    )
                    for index, candidate in enumerate(canonical_gap_candidates)
                },
                key=lambda value: {
                    "new": 0,
                    "needs_review": 1,
                    "ready_for_approval": 2,
                    "approved": 3,
                    "rejected": 4,
                    "ignored": 5,
                    "promoted": 6,
                    "active": 7,
                }.get(value, 99),
            )
            available_owners = sorted(
                {
                    _normalized_text(item.get("owner"))
                    for item in canonical_gap_stewardship_items.values()
                    if _normalized_text(item.get("owner"))
                }
            )
            available_assignees = sorted(
                {
                    _normalized_text(item.get("assignee"))
                    for item in canonical_gap_stewardship_items.values()
                    if _normalized_text(item.get("assignee"))
                }
            )
            available_sources = sorted(
                {
                    _normalized_text(candidate.get("source"))
                    for candidate in canonical_gap_candidates
                    if _normalized_text(candidate.get("source"))
                }
            )
            _apply_pending_governance_selectbox_value(
                st.session_state,
                pending_key="pending_governance_gap_source_filter",
                target_key="debug_canonical_gap_source_filter",
                options=available_sources,
            )
            queue_filter_columns = st.columns(4)
            queue_status_filter = queue_filter_columns[0].selectbox(
                "Stewardship status",
                options=[""] + available_stewardship_statuses,
                key="debug_canonical_gap_status_filter",
                format_func=lambda value: _canonical_gap_proposal_state_label(value) if value else "All statuses",
            )
            queue_owner_filter = queue_filter_columns[1].selectbox(
                "Owner",
                options=[""] + available_owners,
                key="debug_canonical_gap_owner_filter",
                format_func=lambda value: value or "All owners",
            )
            queue_assignee_filter = queue_filter_columns[2].selectbox(
                "Assignee",
                options=[""] + available_assignees,
                key="debug_canonical_gap_assignee_filter",
                format_func=lambda value: value or "All assignees",
            )
            queue_source_filter = queue_filter_columns[3].selectbox(
                "Source field",
                options=[""] + available_sources,
                key="debug_canonical_gap_source_filter",
                format_func=lambda value: value or "All source fields",
            )
            focus_sources = _governance_focus_sources(st.session_state)
            focus_caption = _governance_focus_source_caption(queue_source_filter, focus_sources)
            if focus_caption:
                st.caption(focus_caption)
            visible_gap_indices = [
                index
                for index, candidate in enumerate(canonical_gap_candidates)
                if (
                    (not queue_status_filter)
                    or _canonical_gap_effective_status(
                        _canonical_gap_candidate_key(index, candidate),
                        canonical_gap_console_states,
                        canonical_gap_proposal_states,
                        canonical_gap_stewardship_items.get(_canonical_gap_candidate_key(index, candidate)),
                    )
                    == queue_status_filter
                )
                and (
                    (not queue_owner_filter)
                    or _normalized_text(
                        (canonical_gap_stewardship_items.get(_canonical_gap_candidate_key(index, candidate)) or {}).get("owner")
                    )
                    == queue_owner_filter
                )
                and (
                    (not queue_assignee_filter)
                    or _normalized_text(
                        (canonical_gap_stewardship_items.get(_canonical_gap_candidate_key(index, candidate)) or {}).get("assignee")
                    )
                    == queue_assignee_filter
                )
                and (
                    (not queue_source_filter and not focus_sources)
                    or (queue_source_filter and _normalized_text(candidate.get("source")) == queue_source_filter)
                    or ((not queue_source_filter) and focus_sources and _normalized_text(candidate.get("source")) in focus_sources)
                )
            ]
            st.dataframe(
                _canonical_gap_queue_rows(
                    [canonical_gap_candidates[index] for index in visible_gap_indices],
                    canonical_gap_suggestions,
                    canonical_gap_console_states,
                    canonical_gap_proposal_states,
                    canonical_gap_stewardship_items,
                ),
                width="stretch",
                hide_index=True,
            )

            gap_options = {
                _canonical_gap_option_label(index, canonical_gap_candidates[index], canonical_gap_suggestions.get(_canonical_gap_candidate_key(index, canonical_gap_candidates[index]))): index
                for index in visible_gap_indices
            }
            if not gap_options:
                st.info("No canonical gaps match current stewardship filters.")
                return
            selected_gap_label = st.selectbox(
                "Canonical gap queue detail",
                list(gap_options.keys()),
                key="debug_selected_canonical_gap_label",
            )
            selected_gap_index = gap_options[selected_gap_label]
            selected_candidate = canonical_gap_candidates[selected_gap_index]
            selected_candidate_key = _canonical_gap_candidate_key(selected_gap_index, selected_candidate)
            selected_suggestion = canonical_gap_suggestions.get(selected_candidate_key) or {}
            selected_stewardship_item = canonical_gap_stewardship_items.get(selected_candidate_key) or {}
            selected_stewardship_item_id = selected_stewardship_item.get("item_id")
            selected_console_state = _canonical_gap_console_state(
                selected_candidate_key,
                canonical_gap_console_states,
                selected_stewardship_item,
            )
            selected_proposal_state = _canonical_gap_proposal_state(
                selected_candidate_key,
                canonical_gap_proposal_states,
                selected_stewardship_item,
            )
            selected_effective_status = _canonical_gap_effective_status(
                selected_candidate_key,
                canonical_gap_console_states,
                canonical_gap_proposal_states,
                selected_stewardship_item,
            )

            with st.expander(
                _section_label("Selected Gap Detail", _canonical_gap_proposal_state_label(selected_effective_status)),
                expanded=True,
            ):
                detail_columns = st.columns([3, 3, 2])
                detail_columns[0].markdown(f"**Source:** {_normalized_text(selected_candidate.get('source')) or 'n/a'}")
                detail_columns[1].markdown(f"**Target:** {_normalized_text(selected_candidate.get('target')) or 'n/a'}")
                detail_columns[2].metric("Confidence", f"{int(float(selected_candidate.get('confidence', 0.0) or 0.0) * 100)}%")
                st.caption(f"Console state: {selected_console_state}")
                st.caption(f"Proposal triage: {_canonical_gap_proposal_state_label(selected_proposal_state)}")
                st.caption(f"Stewardship status: {_canonical_gap_proposal_state_label(selected_effective_status)}")
                if selected_candidate.get("reason"):
                    st.caption(selected_candidate.get("reason"))
                if selected_candidate.get("explanation"):
                    st.caption("Signals: " + " | ".join(selected_candidate.get("explanation") or []))

                stewardship_columns = st.columns(2)
                stewardship_owner = stewardship_columns[0].text_input(
                    "Owner",
                    value=_normalized_text(selected_stewardship_item.get("owner")),
                    key=f"debug_stewardship_owner_{selected_gap_index}",
                    placeholder="Example: data-governance",
                )
                stewardship_assignee = stewardship_columns[1].text_input(
                    "Assignee",
                    value=_normalized_text(selected_stewardship_item.get("assignee")),
                    key=f"debug_stewardship_assignee_{selected_gap_index}",
                    placeholder="Example: analyst-on-duty",
                )

                review_note = st.text_input(
                    "Review note",
                    value=_normalized_text(selected_stewardship_item.get("review_note")),
                    key=f"debug_review_note_{selected_gap_index}",
                    placeholder="Why is this gap being ignored, rejected, or ready for approval?",
                )

                save_stewardship_disabled = (
                    bool(selected_stewardship_item_id)
                    and _normalized_text(selected_stewardship_item.get("owner")) == _normalized_text(stewardship_owner)
                    and _normalized_text(selected_stewardship_item.get("assignee")) == _normalized_text(stewardship_assignee)
                    and _normalized_text(selected_stewardship_item.get("review_note")) == _normalized_text(review_note)
                    and _canonical_gap_stewardship_status(selected_stewardship_item) == selected_effective_status
                )
                if st.button(
                    "Save stewardship fields",
                    width="stretch",
                    key=f"debug_save_stewardship_item_{selected_gap_index}",
                    disabled=save_stewardship_disabled,
                ):
                    try:
                        api_request(
                            "POST",
                            "/knowledge/stewardship-items",
                            json=_canonical_gap_stewardship_item_request(
                                selected_candidate_key,
                                selected_candidate,
                                selected_suggestion or None,
                                status=selected_effective_status,
                                owner=stewardship_owner,
                                assignee=stewardship_assignee,
                                review_note=review_note,
                                changed_by=st.session_state.get("admin_token"),
                            ),
                        )
                        st.session_state["debug_knowledge_stewardship_items"] = api_request("GET", "/knowledge/stewardship-items")
                        st.session_state["debug_canonical_gap_proposal_states"] = _canonical_gap_proposal_state_map(
                            st.session_state.get("debug_knowledge_stewardship_items")
                        )
                        st.session_state["debug_knowledge_audit_logs"] = api_request("GET", "/knowledge/audit")
                        st.session_state["last_action"] = {
                            "level": "info",
                            "message": "Saved durable stewardship owner, assignee, and review note for the selected canonical gap.",
                        }
                    except httpx.HTTPError as error:
                        st.session_state["last_action"] = {
                            "level": "error",
                            "message": f"Saving stewardship fields failed: {error}",
                        }
                    st.rerun()

                triage_columns = st.columns([2, 1])
                triage_selection = triage_columns[0].selectbox(
                    "Proposal triage",
                    options=["new", "needs_review", "ready_for_approval"],
                    index=["new", "needs_review", "ready_for_approval"].index(selected_proposal_state),
                    key=f"debug_proposal_state_{selected_gap_index}",
                    format_func=_canonical_gap_proposal_state_label,
                )
                if triage_columns[1].button(
                    "Update triage",
                    width="stretch",
                    key=f"debug_update_proposal_state_{selected_gap_index}",
                    disabled=triage_selection == selected_proposal_state,
                ):
                    try:
                        if selected_stewardship_item_id:
                            api_request(
                                "POST",
                                f"/knowledge/stewardship-items/{selected_stewardship_item_id}/status",
                                json=_knowledge_stewardship_status_update_request(
                                    triage_selection,
                                    st.session_state.get("admin_token"),
                                    owner=stewardship_owner,
                                    assignee=stewardship_assignee,
                                    review_note=review_note,
                                ),
                            )
                        else:
                            api_request(
                                "POST",
                                "/knowledge/stewardship-items",
                                json=_canonical_gap_stewardship_item_request(
                                    selected_candidate_key,
                                    selected_candidate,
                                    selected_suggestion or None,
                                    status=triage_selection,
                                    owner=stewardship_owner,
                                    assignee=stewardship_assignee,
                                    review_note=review_note,
                                    changed_by=st.session_state.get("admin_token"),
                                ),
                            )
                        st.session_state["debug_knowledge_stewardship_items"] = api_request("GET", "/knowledge/stewardship-items")
                        st.session_state["debug_canonical_gap_proposal_states"] = _canonical_gap_proposal_state_map(
                            st.session_state.get("debug_knowledge_stewardship_items")
                        )
                        st.session_state["debug_knowledge_audit_logs"] = api_request("GET", "/knowledge/audit")
                        st.session_state["last_action"] = {
                            "level": "info",
                            "message": (
                                f"Updated proposal triage to '{_canonical_gap_proposal_state_label(triage_selection)}' for "
                                f"{_normalized_text(selected_candidate.get('source')) or 'selected gap'}."
                            ),
                        }
                    except httpx.HTTPError as error:
                        st.session_state["last_action"] = {
                            "level": "error",
                            "message": f"Updating proposal triage failed: {error}",
                        }
                    st.rerun()

                state_action_columns = st.columns(3)
                if state_action_columns[0].button(
                    "Ignore with audit",
                    width="stretch",
                    key=f"debug_ignore_canonical_gap_{selected_gap_index}",
                    disabled=not _canonical_gap_can_ignore(selected_console_state),
                ):
                    try:
                        api_request(
                            "POST",
                            "/knowledge/canonical-gaps/reject",
                            json=_canonical_gap_rejection_request(
                                selected_candidate,
                                selected_suggestion or None,
                                st.session_state.get("admin_token"),
                                review_note,
                                disposition="ignored",
                            ),
                        )
                        canonical_gap_console_states[selected_candidate_key] = "ignored"
                        if selected_stewardship_item_id:
                            api_request(
                                "POST",
                                f"/knowledge/stewardship-items/{selected_stewardship_item_id}/status",
                                json=_knowledge_stewardship_status_update_request(
                                    "ignored",
                                    st.session_state.get("admin_token"),
                                    owner=stewardship_owner,
                                    assignee=stewardship_assignee,
                                    review_note=review_note,
                                    note=review_note,
                                ),
                            )
                        else:
                            api_request(
                                "POST",
                                "/knowledge/stewardship-items",
                                json=_canonical_gap_stewardship_item_request(
                                    selected_candidate_key,
                                    selected_candidate,
                                    selected_suggestion or None,
                                    status="ignored",
                                    owner=stewardship_owner,
                                    assignee=stewardship_assignee,
                                    review_note=review_note,
                                    changed_by=st.session_state.get("admin_token"),
                                ),
                            )
                        st.session_state["debug_knowledge_stewardship_items"] = api_request("GET", "/knowledge/stewardship-items")
                        st.session_state["debug_canonical_gap_proposal_states"] = _canonical_gap_proposal_state_map(
                            st.session_state.get("debug_knowledge_stewardship_items")
                        )
                        st.session_state["debug_knowledge_audit_logs"] = api_request("GET", "/knowledge/audit")
                        st.session_state["last_action"] = {
                            "level": "info",
                            "message": "Ignored canonical gap suggestion and persisted an audit entry.",
                        }
                    except httpx.HTTPError as error:
                        st.session_state["last_action"] = {
                            "level": "error",
                            "message": f"Canonical gap ignore failed: {error}",
                        }
                    st.rerun()

                if state_action_columns[1].button(
                    "Restore to queue",
                    width="stretch",
                    key=f"debug_restore_canonical_gap_{selected_gap_index}",
                    disabled=not _canonical_gap_can_restore(selected_console_state),
                ):
                    canonical_gap_console_states.pop(selected_candidate_key, None)
                    if selected_stewardship_item_id:
                        try:
                            api_request(
                                "POST",
                                f"/knowledge/stewardship-items/{selected_stewardship_item_id}/status",
                                json=_knowledge_stewardship_status_update_request(
                                    "needs_review",
                                    st.session_state.get("admin_token"),
                                    owner=stewardship_owner,
                                    assignee=stewardship_assignee,
                                    review_note=review_note,
                                ),
                            )
                            st.session_state["debug_knowledge_stewardship_items"] = api_request("GET", "/knowledge/stewardship-items")
                            st.session_state["debug_canonical_gap_proposal_states"] = _canonical_gap_proposal_state_map(
                                st.session_state.get("debug_knowledge_stewardship_items")
                            )
                        except httpx.HTTPError as error:
                            st.session_state["last_action"] = {
                                "level": "error",
                                "message": f"Restoring stewardship status failed: {error}",
                            }
                            st.rerun()
                    st.session_state["last_action"] = {
                        "level": "info",
                        "message": "Restored canonical gap to the active console queue. Existing ignore audit entries remain in history.",
                    }
                    st.rerun()

                if state_action_columns[2].button(
                    "Reject with audit",
                    width="stretch",
                    key=f"debug_reject_canonical_gap_{selected_gap_index}",
                    disabled=not _canonical_gap_can_reject(selected_console_state),
                ):
                    try:
                        api_request(
                            "POST",
                            "/knowledge/canonical-gaps/reject",
                            json=_canonical_gap_rejection_request(
                                selected_candidate,
                                selected_suggestion or None,
                                st.session_state.get("admin_token"),
                                review_note,
                            ),
                        )
                        canonical_gap_console_states[selected_candidate_key] = "rejected"
                        if selected_stewardship_item_id:
                            api_request(
                                "POST",
                                f"/knowledge/stewardship-items/{selected_stewardship_item_id}/status",
                                json=_knowledge_stewardship_status_update_request(
                                    "rejected",
                                    st.session_state.get("admin_token"),
                                    owner=stewardship_owner,
                                    assignee=stewardship_assignee,
                                    review_note=review_note,
                                    note=review_note,
                                ),
                            )
                        else:
                            api_request(
                                "POST",
                                "/knowledge/stewardship-items",
                                json=_canonical_gap_stewardship_item_request(
                                    selected_candidate_key,
                                    selected_candidate,
                                    selected_suggestion or None,
                                    status="rejected",
                                    owner=stewardship_owner,
                                    assignee=stewardship_assignee,
                                    review_note=review_note,
                                    changed_by=st.session_state.get("admin_token"),
                                ),
                            )
                        st.session_state["debug_knowledge_stewardship_items"] = api_request("GET", "/knowledge/stewardship-items")
                        st.session_state["debug_canonical_gap_proposal_states"] = _canonical_gap_proposal_state_map(
                            st.session_state.get("debug_knowledge_stewardship_items")
                        )
                        st.session_state["debug_knowledge_audit_logs"] = api_request("GET", "/knowledge/audit")
                        st.session_state["last_action"] = {
                            "level": "info",
                            "message": "Rejected canonical gap suggestion and persisted an audit entry.",
                        }
                    except httpx.HTTPError as error:
                        st.session_state["last_action"] = {
                            "level": "error",
                            "message": f"Canonical gap rejection failed: {error}",
                        }
                    st.rerun()

                related_gap_audit_entries = _canonical_gap_related_audit_entries(
                    st.session_state.get("debug_knowledge_audit_logs"),
                    selected_candidate,
                )
                if related_gap_audit_entries:
                    st.write("**Gap audit references**")
                    st.dataframe(related_gap_audit_entries, width="stretch", hide_index=True)

                if selected_suggestion:
                    st.write(
                        f"Action: **{selected_suggestion.get('action', 'no_action')}** | "
                        f"Concept: **{selected_suggestion.get('concept_id') or 'n/a'}** - {selected_suggestion.get('display_name') or 'n/a'}"
                    )
                    if selected_suggestion.get("aliases"):
                        st.caption("Aliases: " + ", ".join(selected_suggestion.get("aliases") or []))
                    for line in selected_suggestion.get("reasoning") or []:
                        st.caption(f"Reason: {line}")
                    for line in selected_suggestion.get("risk_notes") or []:
                        st.caption(f"Risk: {line}")

                    impact_preview_rows = _canonical_gap_impact_preview_rows(
                        selected_gap_index,
                        selected_candidate,
                        selected_suggestion,
                        canonical_gap_candidates,
                        canonical_gap_suggestions,
                        canonical_gap_console_states,
                    )
                    suggested_concept_id = _normalized_text(selected_suggestion.get("concept_id"))
                    suggested_concept_summary = next(
                        (
                            concept
                            for concept in st.session_state.get("debug_canonical_concepts") or []
                            if _normalized_text(concept.get("concept_id")) == suggested_concept_id
                        ),
                        None,
                    )
                    if impact_preview_rows or suggested_concept_summary:
                        st.write("**Impact preview**")
                        impact_metrics = st.columns(3)
                        impact_metrics[0].metric("Potential queue rows", len(impact_preview_rows))
                        impact_metrics[1].metric("Additional rows", max(0, len(impact_preview_rows) - 1))
                        impact_metrics[2].metric(
                            "Saved usages",
                            int((suggested_concept_summary or {}).get("usage_count", 0) or 0),
                        )
                        if suggested_concept_summary and (
                            suggested_concept_summary.get("source_systems") or suggested_concept_summary.get("business_domains")
                        ):
                            st.caption(
                                "Concept facets: "
                                + " | ".join(
                                    part
                                    for part in [
                                        (
                                            "source_systems=" + ", ".join(suggested_concept_summary.get("source_systems") or [])
                                            if suggested_concept_summary.get("source_systems")
                                            else ""
                                        ),
                                        (
                                            "business_domains=" + ", ".join(suggested_concept_summary.get("business_domains") or [])
                                            if suggested_concept_summary.get("business_domains")
                                            else ""
                                        ),
                                    ]
                                    if part
                                )
                            )
                        if impact_preview_rows:
                            st.dataframe(impact_preview_rows, width="stretch", hide_index=True)
                        else:
                            st.caption("No additional current queue rows match this proposed concept/alias pattern yet.")

                    approve_ready = _canonical_gap_can_approve(
                        selected_suggestion,
                        selected_console_state,
                        selected_proposal_state,
                    )
                    if st.button(
                        "Approve from console",
                        width="stretch",
                        key=f"debug_approve_canonical_gap_{selected_gap_index}",
                        disabled=not approve_ready,
                    ):
                        try:
                            response = api_request(
                                "POST",
                                "/knowledge/canonical-gaps/approve",
                                json=_canonical_gap_approval_request(
                                    selected_candidate,
                                    selected_suggestion,
                                    st.session_state.get("admin_token"),
                                ),
                            )
                            canonical_gap_console_states[selected_candidate_key] = "approved"
                            if selected_stewardship_item_id:
                                api_request(
                                    "POST",
                                    f"/knowledge/stewardship-items/{selected_stewardship_item_id}/status",
                                    json=_knowledge_stewardship_status_update_request(
                                        "approved",
                                        st.session_state.get("admin_token"),
                                        owner=stewardship_owner,
                                        assignee=stewardship_assignee,
                                        review_note=review_note,
                                    ),
                                )
                            else:
                                api_request(
                                    "POST",
                                    "/knowledge/stewardship-items",
                                    json=_canonical_gap_stewardship_item_request(
                                        selected_candidate_key,
                                        selected_candidate,
                                        selected_suggestion,
                                        status="approved",
                                        owner=stewardship_owner,
                                        assignee=stewardship_assignee,
                                        review_note=review_note,
                                        changed_by=st.session_state.get("admin_token"),
                                    ),
                                )
                            st.session_state["debug_knowledge_runtime"] = api_request("POST", "/knowledge/reload")
                            st.session_state["debug_knowledge_audit_logs"] = api_request("GET", "/knowledge/audit")
                            st.session_state["debug_knowledge_stewardship_items"] = api_request("GET", "/knowledge/stewardship-items")
                            st.session_state["debug_canonical_gap_proposal_states"] = _canonical_gap_proposal_state_map(
                                st.session_state.get("debug_knowledge_stewardship_items")
                            )
                            if st.session_state.get("debug_canonical_concepts") is not None:
                                st.session_state["debug_canonical_concepts"] = api_request("GET", "/knowledge/canonical-concepts")
                            approved_concept_id = _normalized_text(selected_suggestion.get("concept_id"))
                            if approved_concept_id:
                                st.session_state["debug_canonical_concept_detail"] = api_request(
                                    "GET",
                                    f"/knowledge/canonical-concepts/{approved_concept_id}",
                                )
                            if st.session_state.get("debug_knowledge_overlays") is not None:
                                st.session_state["debug_knowledge_overlays"] = api_request("GET", "/knowledge/overlays")
                            st.session_state["last_action"] = {
                                "level": "success",
                                "message": (
                                    f"Approved canonical gap into overlay '{response.get('overlay_name')}'. "
                                    "Regenerate mapping to see the canonical path filled."
                                ),
                            }
                        except httpx.HTTPError as error:
                            st.session_state["last_action"] = {
                                "level": "error",
                                "message": f"Canonical gap approval failed: {error}",
                            }
                        st.rerun()
                    elif selected_console_state == "ignored":
                        st.caption("This gap is currently ignored in the console. Restore it to the active queue before approving.")
                    elif selected_console_state == "approved":
                        st.caption("This gap was already approved from the console in this session.")
                    elif selected_console_state == "rejected":
                        st.caption("This gap was rejected from the console and the decision was persisted to the audit log.")
                    elif selected_proposal_state != "ready_for_approval":
                        st.caption(
                            "Move proposal triage to 'Ready for approval' before approving from the console."
                        )
                    elif not approve_ready:
                        st.caption("This suggestion is not approve-ready. Generate a usable non-`no_action` suggestion from the Review tab first.")
                else:
                    if selected_console_state == "rejected":
                        st.caption("This gap was rejected without a cached suggestion payload. The audit decision is persisted.")
                    elif selected_console_state == "ignored":
                        st.caption("This gap was ignored with audit. Restore it locally if you want to reconsider it in the queue.")
                    else:
                        st.info("No cached LLM suggestion for this gap yet. Generate it from the Review tab.")
        else:
            st.info("Canonical gap review queue is empty. Use the Review tab to find high-confidence gaps first.")


def render_admin_debug_tab(
    *,
    admin_token_required: Callable[[], bool],
    api_request: Callable[..., Any],
    api_request_content: Callable[..., bytes],
    upload_file_to_request_files: Callable[[Any], dict | None],
    current_mapping_rows: Callable[[dict], list[dict]],
    knowledge_debug_rows: Callable[[dict], list[dict]],
) -> None:
    """Render the system surface for admin controls and debug observability."""

    st.header("System")
    admin_token = st.session_state.get("admin_token", "").strip()
    token_required = admin_token_required()
    if token_required and not admin_token:
        st.warning("Admin token is required for system endpoints.")
        return
    if not token_required:
        st.info("Backend currently exposes these system endpoints without an admin token.")

    knowledge_runtime = st.session_state.get("debug_knowledge_runtime")
    knowledge_audit_logs = st.session_state.get("debug_knowledge_audit_logs")
    runtime_config = st.session_state.get("debug_runtime_config")
    decision_logs = st.session_state.get("debug_decision_logs")
    corrections = st.session_state.get("debug_corrections")
    runs = st.session_state.get("debug_runs")
    mapping_response = st.session_state.get("mapping_response")

    admin_tab, debug_tab = st.tabs(["Admin", "Debug"])

    with admin_tab:
        action_columns = st.columns(3)
        if action_columns[0].button("Load runtime config", width="stretch", key="debug_load_runtime_config"):
            try:
                loaded_runtime = api_request("GET", "/observability/config")
                st.session_state["debug_runtime_config"] = loaded_runtime
                st.session_state["runtime_config_snapshot"] = loaded_runtime
                st.session_state["last_action"] = {"level": "success", "message": "Loaded runtime config snapshot."}
            except httpx.HTTPError as error:
                st.session_state["last_action"] = {"level": "error", "message": f"Loading runtime config failed: {error}"}
            st.rerun()

        if action_columns[1].button("Load saved corrections", width="stretch", key="debug_load_corrections"):
            try:
                st.session_state["debug_corrections"] = api_request("GET", "/observability/corrections")
                st.session_state["last_action"] = {"level": "success", "message": "Loaded saved corrections."}
            except httpx.HTTPError as error:
                st.session_state["last_action"] = {"level": "error", "message": f"Loading corrections failed: {error}"}
            st.rerun()

        if action_columns[2].button("Load benchmark runs", width="stretch", key="debug_load_benchmark_runs"):
            try:
                st.session_state["debug_runs"] = api_request("GET", "/evaluation/runs")
                st.session_state["last_action"] = {"level": "success", "message": "Loaded evaluation runs."}
            except httpx.HTTPError as error:
                st.session_state["last_action"] = {"level": "error", "message": f"Loading evaluation runs failed: {error}"}
            st.rerun()

        runtime_source = runtime_config or st.session_state.get("runtime_config_snapshot") or {}
        available_scoring_profiles = list(runtime_source.get("available_scoring_profiles") or [])
        if not available_scoring_profiles:
            available_scoring_profiles = [
                "balanced",
                "schema_only",
                "data_rich",
                "canonical_first",
                "description_priority",
            ]
        current_scoring_profile = str(runtime_source.get("scoring_profile") or "balanced").strip() or "balanced"
        if current_scoring_profile not in available_scoring_profiles:
            available_scoring_profiles = [current_scoring_profile, *available_scoring_profiles]
        current_backend_build = str(runtime_source.get("backend_build") or "n/a").strip() or "n/a"
        current_app_version = str(runtime_source.get("app_version") or "n/a").strip() or "n/a"

        st.subheader("Scoring Runtime")
        st.caption(f"Current build: {current_backend_build} | app_version={current_app_version}")

        if st.button("Refresh runtime snapshot", width="stretch", key="debug_refresh_runtime_snapshot"):
            try:
                refreshed_runtime = api_request("GET", "/observability/config")
                st.session_state["debug_runtime_config"] = refreshed_runtime
                st.session_state["runtime_config_snapshot"] = refreshed_runtime
                st.session_state["last_action"] = {"level": "success", "message": "Refreshed runtime snapshot."}
            except httpx.HTTPError as error:
                st.session_state["last_action"] = {"level": "error", "message": f"Refreshing runtime snapshot failed: {error}"}
            st.rerun()

        selected_scoring_profile = st.selectbox(
            "Active scoring profile",
            available_scoring_profiles,
            index=available_scoring_profiles.index(current_scoring_profile),
            key="debug_active_scoring_profile",
            help="Applies to new mapping runs after this update. Existing mapping results keep their original runtime fingerprint.",
        )
        if st.button("Apply scoring profile", width="stretch", key="debug_apply_scoring_profile"):
            try:
                updated_runtime = api_request(
                    "POST",
                    "/observability/config/scoring-profile",
                    json={"scoring_profile": selected_scoring_profile},
                )
                st.session_state["debug_runtime_config"] = updated_runtime
                st.session_state["runtime_config_snapshot"] = updated_runtime
                st.session_state["last_action"] = {
                    "level": "success",
                    "message": f"Updated active scoring profile to {updated_runtime.get('scoring_profile', selected_scoring_profile)}.",
                }
            except httpx.HTTPError as error:
                st.session_state["last_action"] = {
                    "level": "error",
                    "message": f"Updating scoring profile failed: {error}",
                }
            st.rerun()

        if runtime_config:
            st.subheader("Runtime Config")
            st.json(runtime_config)

        if corrections:
            st.subheader("Saved Corrections")
            st.dataframe(corrections, width="stretch", hide_index=True)

        if runs:
            st.subheader("Evaluation Runs")
            st.dataframe(runs, width="stretch", hide_index=True)

    with debug_tab:
        debug_action_columns = st.columns(3)
        if debug_action_columns[0].button("Load decision logs", width="stretch", key="debug_load_decision_logs"):
            try:
                st.session_state["debug_decision_logs"] = api_request("GET", "/observability/decision-logs")
                st.session_state["last_action"] = {"level": "success", "message": "Loaded decision logs."}
            except httpx.HTTPError as error:
                st.session_state["last_action"] = {"level": "error", "message": f"Loading decision logs failed: {error}"}
            st.rerun()

        if debug_action_columns[1].button("Load active knowledge status", width="stretch", key="debug_load_knowledge_runtime"):
            try:
                st.session_state["debug_knowledge_runtime"] = api_request("POST", "/knowledge/reload")
                st.session_state["last_action"] = {"level": "success", "message": "Loaded active knowledge runtime status."}
            except httpx.HTTPError as error:
                st.session_state["last_action"] = {"level": "error", "message": f"Loading knowledge runtime status failed: {error}"}
            st.rerun()

        if debug_action_columns[2].button("Load knowledge audit log", width="stretch", key="debug_load_knowledge_audit"):
            try:
                st.session_state["debug_knowledge_audit_logs"] = api_request("GET", "/knowledge/audit")
                st.session_state["last_action"] = {"level": "success", "message": "Loaded knowledge audit log."}
            except httpx.HTTPError as error:
                st.session_state["last_action"] = {"level": "error", "message": f"Loading knowledge audit log failed: {error}"}
            st.rerun()

        st.info("Governance now owns overlay summary, overlay lifecycle controls, and canonical glossary authoring UI.")

        if knowledge_runtime:
            st.subheader("Knowledge Governance Debug")
            st.caption(
                "Knowledge mode: "
                + str(knowledge_runtime.get("mode") or "base_only")
                + " | runtime source: "
                + str(knowledge_runtime.get("runtime_source") or "unknown")
                + " | source hash: "
                + str(knowledge_runtime.get("source_hash_state") or "missing")
                + " | active overlay: "
                + str(knowledge_runtime.get("active_overlay_name") or "none")
                + f" | active_entry_count={knowledge_runtime.get('active_entry_count', 0)}"
                + f" | concept_count={knowledge_runtime.get('concept_count', 0)}"
            )
            if knowledge_runtime.get("seeded_at"):
                st.caption(
                    "Seed cache: "
                    + str(knowledge_runtime.get("seeded_at"))
                    + f" | seeded_concepts={knowledge_runtime.get('seeded_concept_count', 0)}"
                    + f" | seeded_canonical={knowledge_runtime.get('seeded_canonical_concept_count', 0)}"
                )
            entry_type_counts = knowledge_runtime.get("entry_type_counts") or {}
            if entry_type_counts:
                st.caption(
                    "Active overlay breakdown: "
                    + " | ".join(f"{entry_type}={count}" for entry_type, count in sorted(entry_type_counts.items()))
                )

        if knowledge_audit_logs:
            st.subheader("Knowledge Audit Log")
            st.dataframe(knowledge_audit_logs, width="stretch", hide_index=True)

        if mapping_response:
            canonical_coverage = mapping_response.get("canonical_coverage") or {}
            source_coverage = canonical_coverage.get("source") or {}
            target_coverage = canonical_coverage.get("target") or {}
            project_coverage = canonical_coverage.get("project") or {}
            if source_coverage or target_coverage or project_coverage:
                st.subheader("Canonical Coverage")
                st.dataframe(
                    [
                        {
                            "dataset": "source",
                            "coverage_ratio": source_coverage.get("coverage_ratio", 0.0),
                            "matched_columns": source_coverage.get("matched_columns", 0),
                            "total_columns": source_coverage.get("total_columns", 0),
                            "unmatched_columns": " | ".join(source_coverage.get("unmatched_columns", [])),
                        },
                        {
                            "dataset": "target",
                            "coverage_ratio": target_coverage.get("coverage_ratio", 0.0),
                            "matched_columns": target_coverage.get("matched_columns", 0),
                            "total_columns": target_coverage.get("total_columns", 0),
                            "unmatched_columns": " | ".join(target_coverage.get("unmatched_columns", [])),
                        },
                        {
                            "dataset": "project",
                            "coverage_ratio": project_coverage.get("coverage_ratio", 0.0),
                            "matched_columns": project_coverage.get("matched_columns", 0),
                            "total_columns": project_coverage.get("total_columns", 0),
                            "unmatched_columns": "shared=" + " | ".join(project_coverage.get("shared_concepts", [])),
                        },
                    ],
                    width="stretch",
                    hide_index=True,
                )
            mapping_spec_rows = current_mapping_rows(mapping_response)
            knowledge_rows = knowledge_debug_rows(mapping_response)
            if mapping_spec_rows:
                st.subheader("Knowledge and Canonical Match Insights")
                st.caption("Uses the same mapping-spec row shape as the Review tab/export so admin inspection stays column-compatible.")
                st.dataframe(mapping_spec_rows, width="stretch", hide_index=True)
            if mapping_spec_rows:
                mapping_spec_by_key = {
                    (str(item.get("source") or ""), str(item.get("target") or "")): item
                    for item in mapping_spec_rows
                }
                knowledge_row_by_key = {
                    (str(item.get("source") or ""), str(item.get("target") or "")): item
                    for item in knowledge_rows
                }
                for mapping_spec_row in mapping_spec_rows:
                    row_key = (
                        str(mapping_spec_row.get("source") or ""),
                        str(mapping_spec_row.get("target") or ""),
                    )
                    row = knowledge_row_by_key.get(row_key) or {
                        "source": row_key[0],
                        "target": row_key[1],
                        "knowledge_explanations": [],
                        "canonical_explanations": [],
                    }
                    with st.expander(f"Knowledge details: {row['source']} -> {row['target']}"):
                        mapping_spec_row = mapping_spec_by_key.get(
                            (str(row.get("source") or ""), str(row.get("target") or ""))
                        )
                        if mapping_spec_row:
                            st.caption("Mapping specification")
                            detail_fields = [
                                "source",
                                "target",
                                "confidence",
                                "confidence_label",
                                "status",
                                "validator",
                                "canonical_status_label",
                                "shared_concepts",
                                "source_concepts",
                                "target_concepts",
                                "canonical_path",
                                "llm_consulted",
                            ]
                            for field_name in detail_fields:
                                st.write(f"**{field_name}:** {mapping_spec_row.get(field_name, '')}")
                            if mapping_spec_row.get("llm_recommendation"):
                                llm_recommendation = mapping_spec_row.get("llm_recommendation") or {}
                                st.caption("LLM recommendation")
                                st.write(
                                    f"**selected_target:** {llm_recommendation.get('selected_target') or 'unmapped'}"
                                )
                                st.write(
                                    f"**confidence:** {llm_recommendation.get('confidence') if llm_recommendation.get('confidence') is not None else ''}"
                                )
                                reasoning = llm_recommendation.get("reasoning") or []
                                if reasoning:
                                    st.write("**reasoning:**")
                                    for line in reasoning:
                                        st.write(f"- {line}")
                                raw_response = str(llm_recommendation.get("raw_response") or "").strip()
                                if raw_response:
                                    with st.expander("Raw LLM response", expanded=False):
                                        st.code(raw_response)
                        if row["knowledge_explanations"]:
                            st.caption("Knowledge explanations")
                        for line in row["knowledge_explanations"]:
                            st.caption(line)
                        if row["canonical_explanations"]:
                            st.caption("Canonical explanations")
                        for line in row["canonical_explanations"]:
                            st.caption(line)
                        if not row["knowledge_explanations"] and not row["canonical_explanations"]:
                            st.caption("No extracted knowledge/canonical explanation is available for this mapping row.")

        if decision_logs:
            st.subheader("Decision Logs")
            st.dataframe(decision_logs, width="stretch", hide_index=True)
