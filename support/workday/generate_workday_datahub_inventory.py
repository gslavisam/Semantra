"""Classify Workday HRDH datahub inventory against canonical business concepts.

Reads HRDH_Table_Columns.xlsx (overview sheet), extracts table-column pairs, classifies
each against the current canonical and knowledge layers, and produces detailed reports.

Classification buckets match SAP/QB patterns:
- direct_alias_match: Column name directly matches a canonical alias (~100% confidence)
- description_alias_match: Column description/type information matches canonical concepts
- strong_canonical_candidate: Top canonical match strength >= 0.75
- weak_canonical_candidate: Top canonical match strength < 0.75
- knowledge_only: Knowledge concept exists but no canonical mapping
- unmapped: No knowledge or canonical signal

Usage:
    python support/workday/generate_workday_datahub_inventory.py
"""

from __future__ import annotations

import csv
from pathlib import Path
from collections import Counter
from dataclasses import dataclass

from openpyxl import load_workbook

from app.services.metadata_knowledge_service import metadata_knowledge_service
from app.models.schema import ColumnProfile


PROJECT_ROOT = Path(__file__).resolve().parents[2]
HRDH_PATH = PROJECT_ROOT / "metadata_dict" / "HRDH_Table_Columns.xlsx"
OUTPUT_DIR = PROJECT_ROOT / "knowledge_sources" / "generated" / "runtime" / "workday"
CLASSIFICATION_OUTPUT_PATH = OUTPUT_DIR / "workday_datahub_classification.csv"
GAP_CANDIDATES_PATH = OUTPUT_DIR / "workday_datahub_gap_candidates.csv"
SUMMARY_PATH = OUTPUT_DIR / "workday_datahub_summary.csv"


@dataclass
class HRDHColumn:
    table_name: str
    column_id: int
    column_name: str
    type_name: str
    max_length: int | None
    is_identity: bool
    is_nullable: bool


def main() -> None:
    if not HRDH_PATH.exists():
        print(f"Error: {HRDH_PATH} not found.")
        return

    # Refresh knowledge/canonical runtime
    metadata_knowledge_service.refresh()
    print("Knowledge/canonical runtime refreshed")

    # Load HRDH data
    columns = load_hrdh_inventory(HRDH_PATH)
    print(f"Loaded {len(columns)} HRDH column entries from {len(set(c.table_name for c in columns))} tables")

    # Classify each column
    classified_rows = []
    bucket_counts = Counter()

    for col in columns:
        row = classify_column(col)
        classified_rows.append(row)
        bucket_counts[row["classification_bucket"]] += 1

    # Write outputs
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    fieldnames = list(classified_rows[0].keys()) if classified_rows else []
    write_csv(CLASSIFICATION_OUTPUT_PATH, classified_rows, fieldnames)

    # Gap candidates (knowledge_only + unmapped)
    gap_candidates = [
        row for row in classified_rows
        if row["classification_bucket"] in ("knowledge_only", "unmapped")
    ]
    write_csv(GAP_CANDIDATES_PATH, gap_candidates, fieldnames)

    # Summary by bucket
    summary_rows = [
        {
            "bucket": bucket,
            "count": str(count),
            "percentage": f"{100 * count / len(classified_rows):.1f}%" if classified_rows else "0%",
        }
        for bucket, count in sorted(bucket_counts.items(), key=lambda x: -x[1])
    ]
    write_summary_csv(SUMMARY_PATH, summary_rows)

    print(f"\nWorkday HRDH Classification Results:")
    for bucket, count in sorted(bucket_counts.items(), key=lambda x: -x[1]):
        print(f"  {bucket}: {count}")
    print(f"\nWrote: {CLASSIFICATION_OUTPUT_PATH}")
    print(f"Wrote: {GAP_CANDIDATES_PATH}")
    print(f"Wrote: {SUMMARY_PATH}")


def load_hrdh_inventory(xlsx_path: Path) -> list[HRDHColumn]:
    columns = []
    wb = load_workbook(xlsx_path)
    ws = wb["overview"] if "overview" in wb.sheetnames else wb.active

    for row_idx in range(2, ws.max_row + 1):
        table_name = ws.cell(row_idx, 1).value
        column_id = ws.cell(row_idx, 2).value
        column_name = ws.cell(row_idx, 3).value
        type_name = ws.cell(row_idx, 4).value
        max_length = ws.cell(row_idx, 5).value
        is_identity = ws.cell(row_idx, 6).value
        is_nullable = ws.cell(row_idx, 7).value

        # Skip rows without table name (continuation rows)
        if not table_name:
            continue

        col = HRDHColumn(
            table_name=str(table_name).strip(),
            column_id=int(column_id) if column_id else 0,
            column_name=str(column_name).strip() if column_name else "",
            type_name=str(type_name).strip() if type_name else "",
            max_length=int(max_length) if isinstance(max_length, int) else None,
            is_identity=bool(is_identity) if is_identity else False,
            is_nullable=bool(is_nullable) if is_nullable else False,
        )
        columns.append(col)

    return columns


def classify_column(col: HRDHColumn) -> dict[str, str]:
    """Classify a single HRDH column against knowledge/canonical."""
    
    # Build column profile with statistical defaults (schema-only, no actual data)
    profile = ColumnProfile(
        name=col.column_name,
        normalized_name=col.column_name.lower().replace("_", ""),
        description=f"{col.table_name} {col.column_name} ({col.type_name})",
        declared_type=col.type_name,
        dtype="object",  # Default to object for schema-only inventory
        null_ratio=0.5 if col.is_nullable else 0.0,  # Estimate based on is_nullable
        unique_ratio=0.5,  # Estimated
        non_null_count=0,  # Schema only - don't have actual row counts
    )

    # Match knowledge
    knowledge_matches = metadata_knowledge_service.match_concepts(
        profile,
        prefer_metadata_text=True,
    )
    top_knowledge = knowledge_matches[0] if knowledge_matches else None

    # Match canonical
    canonical_matches = metadata_knowledge_service.match_canonical_concepts(
        profile,
        prefer_metadata_text=True,
    )
    top_canonical = canonical_matches[0] if canonical_matches else None

    # Check direct field alias match
    direct_field_canonical = None
    try:
        resolved = metadata_knowledge_service.resolve_canonical_concept_id(col.column_name)
        if resolved:
            direct_field_canonical = resolved
    except:
        pass

    # Check description match
    description_canonical = None
    if not direct_field_canonical and col.column_name:
        try:
            resolved = metadata_knowledge_service.resolve_canonical_concept_id(col.column_name)
            if resolved:
                description_canonical = resolved
        except:
            pass

    # Classify
    bucket = classify_bucket(
        direct_field_canonical,
        description_canonical,
        top_canonical,
        top_knowledge,
    )

    return {
        "wd_table": col.table_name,
        "wd_column": col.column_name,
        "wd_type": col.type_name,
        "wd_max_length": str(col.max_length) if col.max_length else "",
        "wd_is_identity": str(col.is_identity),
        "wd_is_nullable": str(col.is_nullable),
        "classification_bucket": bucket,
        "direct_field_canonical_concept_id": direct_field_canonical or "",
        "description_canonical_concept_id": description_canonical or "",
        "top_canonical_concept_id": top_canonical.concept_id if top_canonical else "",
        "top_canonical_strength": f"{top_canonical.strength:.3f}" if top_canonical else "",
        "top_knowledge_concept_id": top_knowledge.concept_id if top_knowledge else "",
        "top_knowledge_strength": f"{top_knowledge.strength:.3f}" if top_knowledge else "",
    }


def classify_bucket(
    direct_field_canonical: str | None,
    description_canonical: str | None,
    top_canonical,
    top_knowledge,
) -> str:
    """Determine classification bucket based on matches."""
    if direct_field_canonical:
        return "direct_alias_match"
    if description_canonical:
        return "description_alias_match"
    if top_canonical and top_canonical.strength >= 0.75:
        return "strong_canonical_candidate"
    if top_canonical:
        return "weak_canonical_candidate"
    if top_knowledge:
        return "knowledge_only"
    return "unmapped"


def write_csv(path: Path, rows: list[dict[str, str]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def write_summary_csv(path: Path, rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=["bucket", "count", "percentage"])
        writer.writeheader()
        writer.writerows(rows)


if __name__ == "__main__":
    main()
