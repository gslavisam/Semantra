"""Export the full SAP workbook inventory against the current canonical runtime.

Reads metadata_dict/sap_tables_mostUsed.xlsx, classifies each SAP table-column row
against the current knowledge and canonical runtime, and writes reviewable CSV
artifacts under knowledge_sources/generated/runtime/sap/.

Usage:
    python support/sap/generate_sap_canonical_inventory.py [--limit 500] [--output-dir PATH]
"""

from __future__ import annotations

import argparse
import csv
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from app.models.schema import ColumnProfile
from app.services.metadata_knowledge_service import metadata_knowledge_service
from app.utils.normalization import semantic_token_set


PROJECT_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_WORKBOOK_PATH = PROJECT_ROOT / "metadata_dict" / "sap_tables_mostUsed.xlsx"
DEFAULT_OUTPUT_DIR = PROJECT_ROOT / "knowledge_sources" / "generated" / "runtime" / "sap"


@dataclass(frozen=True)
class SapInventoryRow:
    table: str
    module: str
    field: str
    description: str
    data_element: str
    domain: str
    table_description: str


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Classify the full SAP workbook inventory against current canonical concepts.")
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK_PATH, help="Path to sap_tables_mostUsed.xlsx")
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR, help="Directory for generated CSV outputs")
    parser.add_argument("--limit", type=int, default=0, help="Optional row limit for diagnostic runs")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    output_dir = args.output_dir
    output_dir.mkdir(parents=True, exist_ok=True)

    metadata_knowledge_service.refresh()
    table_descriptions, _field_descriptions = metadata_knowledge_service._load_sap_table_descriptions()
    rows = list(load_sap_inventory_rows(args.workbook, table_descriptions, limit=args.limit or None))

    classification_rows: list[dict[str, str]] = []
    gap_rows: list[dict[str, str]] = []
    classification_counts = Counter()
    module_counts: dict[str, Counter[str]] = defaultdict(Counter)

    for row in rows:
        record = classify_row(row)
        classification_rows.append(record)
        classification_counts[record["classification_bucket"]] += 1
        module_counts[row.module or "UNKNOWN"][record["classification_bucket"]] += 1
        if record["classification_bucket"] in {"knowledge_only", "unmapped"}:
            gap_rows.append(record)

    write_csv(
        output_dir / "sap_full_inventory_classification.csv",
        classification_rows,
        [
            "sap_module",
            "sap_table",
            "sap_field",
            "sap_description",
            "sap_data_element",
            "sap_domain",
            "sap_table_description",
            "direct_field_canonical_concept_id",
            "description_canonical_concept_id",
            "top_canonical_concept_id",
            "top_canonical_strength",
            "top_canonical_matches",
            "top_knowledge_concept_id",
            "top_knowledge_strength",
            "top_knowledge_matches",
            "classification_bucket",
            "review_recommendation",
        ],
    )
    write_csv(
        output_dir / "sap_canonical_gap_candidates.csv",
        gap_rows,
        [
            "sap_module",
            "sap_table",
            "sap_field",
            "sap_description",
            "sap_data_element",
            "sap_domain",
            "sap_table_description",
            "direct_field_canonical_concept_id",
            "description_canonical_concept_id",
            "top_canonical_concept_id",
            "top_canonical_strength",
            "top_canonical_matches",
            "top_knowledge_concept_id",
            "top_knowledge_strength",
            "top_knowledge_matches",
            "classification_bucket",
            "review_recommendation",
        ],
    )
    write_summary_csv(output_dir / "sap_full_inventory_summary.csv", len(rows), classification_counts, module_counts)

    print(f"SAP inventory rows processed: {len(rows)}")
    for bucket, count in sorted(classification_counts.items()):
        print(f"{bucket}: {count}")
    print(f"Wrote: {output_dir / 'sap_full_inventory_classification.csv'}")
    print(f"Wrote: {output_dir / 'sap_canonical_gap_candidates.csv'}")
    print(f"Wrote: {output_dir / 'sap_full_inventory_summary.csv'}")


def load_sap_inventory_rows(
    workbook_path: Path,
    table_descriptions: dict[str, str],
    *,
    limit: int | None = None,
) -> list[SapInventoryRow]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    worksheet = workbook["Tbls_Clm"]
    header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
    header_index = {
        str(value or "").strip(): index
        for index, value in enumerate(header_row)
        if str(value or "").strip()
    }
    required_headers = ("Table", "Modul", "Field", "Description", "Data Element", "Domain")
    missing_headers = [header for header in required_headers if header not in header_index]
    if missing_headers:
        missing = ", ".join(missing_headers)
        raise ValueError(f"SAP workbook is missing required Tbls_Clm headers: {missing}")

    rows: list[SapInventoryRow] = []
    for workbook_row in worksheet.iter_rows(min_row=2, values_only=True):
        table = str(workbook_row[header_index["Table"]] or "").strip()
        field = str(workbook_row[header_index["Field"]] or "").strip()
        if not table or not field:
            continue
        row = SapInventoryRow(
            table=table,
            module=str(workbook_row[header_index["Modul"]] or "").strip(),
            field=field,
            description=str(workbook_row[header_index["Description"]] or "").strip(),
            data_element=str(workbook_row[header_index["Data Element"]] or "").strip(),
            domain=str(workbook_row[header_index["Domain"]] or "").strip(),
            table_description=table_descriptions.get(table.lower(), ""),
        )
        rows.append(row)
        if limit and len(rows) >= limit:
            break
    return rows


def classify_row(row: SapInventoryRow) -> dict[str, str]:
    profile = build_profile(row)
    knowledge_matches = sorted(
        metadata_knowledge_service.match_concepts(profile, prefer_metadata_text=True),
        key=lambda match: (-match.strength, match.concept_id),
    )
    canonical_matches = sorted(
        metadata_knowledge_service.match_canonical_concepts(profile, prefer_metadata_text=True),
        key=lambda match: (-match.strength, match.concept_id),
    )

    direct_field_canonical = metadata_knowledge_service.resolve_canonical_concept_id(row.field)
    description_canonical = metadata_knowledge_service.resolve_canonical_concept_id(row.description) if row.description else None

    top_knowledge = knowledge_matches[0] if knowledge_matches else None
    top_canonical = canonical_matches[0] if canonical_matches else None
    bucket, recommendation = classify_bucket(direct_field_canonical, description_canonical, top_canonical, top_knowledge)

    return {
        "sap_module": row.module,
        "sap_table": row.table,
        "sap_field": row.field,
        "sap_description": row.description,
        "sap_data_element": row.data_element,
        "sap_domain": row.domain,
        "sap_table_description": row.table_description,
        "direct_field_canonical_concept_id": direct_field_canonical or "",
        "description_canonical_concept_id": description_canonical or "",
        "top_canonical_concept_id": top_canonical.concept_id if top_canonical else "",
        "top_canonical_strength": f"{top_canonical.strength:.2f}" if top_canonical else "",
        "top_canonical_matches": format_matches(canonical_matches),
        "top_knowledge_concept_id": top_knowledge.concept_id if top_knowledge else "",
        "top_knowledge_strength": f"{top_knowledge.strength:.2f}" if top_knowledge else "",
        "top_knowledge_matches": format_matches(knowledge_matches),
        "classification_bucket": bucket,
        "review_recommendation": recommendation,
    }


def build_profile(row: SapInventoryRow) -> ColumnProfile:
    declared_type = " ".join(part for part in (row.data_element, row.domain) if part)
    tokenized_name = sorted(token for token in semantic_token_set(row.field.replace("_", " ")) if token)
    return ColumnProfile(
        name=row.field,
        normalized_name=row.field.replace("_", " "),
        description=row.description,
        declared_type=declared_type,
        dtype="object",
        null_ratio=0.0,
        unique_ratio=1.0,
        avg_length=float(len(row.field)),
        non_null_count=1,
        sample_values=[],
        distinct_sample_values=[],
        detected_patterns=["text"],
        tokenized_name=tokenized_name,
    )


def classify_bucket(
    direct_field_canonical: str | None,
    description_canonical: str | None,
    top_canonical,
    top_knowledge,
) -> tuple[str, str]:
    if direct_field_canonical:
        return "direct_alias_match", "Already maps directly to an existing canonical concept via SAP field alias."
    if description_canonical:
        return "description_alias_match", "Maps to an existing canonical concept through the SAP field description."
    if top_canonical and top_canonical.strength >= 0.75:
        return "strong_canonical_candidate", "Strong current canonical candidate; review for promotion or alias hardening if needed."
    if top_canonical:
        return "weak_canonical_candidate", "Weak current canonical candidate; requires steward review before treating as mapped."
    if top_knowledge:
        return "knowledge_only", "Knowledge support exists, but no current canonical concept is reachable."
    return "unmapped", "No current knowledge or canonical path was found; candidate for SAP ingest enrichment or canonical gap review."


def format_matches(matches: list, limit: int = 3) -> str:
    formatted: list[str] = []
    for match in matches[:limit]:
        formatted.append(f"{match.concept_id}:{match.strength:.2f}")
    return " | ".join(formatted)


def write_csv(path: Path, rows: list[dict[str, str]], fieldnames: list[str]) -> None:
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def write_summary_csv(
    path: Path,
    total_rows: int,
    classification_counts: Counter[str],
    module_counts: dict[str, Counter[str]],
) -> None:
    fieldnames = ["scope", "scope_value", "total_rows", "classification_bucket", "row_count", "ratio"]
    rows: list[dict[str, str]] = []
    for bucket, row_count in sorted(classification_counts.items()):
        rows.append(
            {
                "scope": "overall",
                "scope_value": "ALL",
                "total_rows": str(total_rows),
                "classification_bucket": bucket,
                "row_count": str(row_count),
                "ratio": f"{(row_count / total_rows):.4f}" if total_rows else "0.0000",
            }
        )
    for module, counts in sorted(module_counts.items()):
        module_total = sum(counts.values())
        for bucket, row_count in sorted(counts.items()):
            rows.append(
                {
                    "scope": "module",
                    "scope_value": module,
                    "total_rows": str(module_total),
                    "classification_bucket": bucket,
                    "row_count": str(row_count),
                    "ratio": f"{(row_count / module_total):.4f}" if module_total else "0.0000",
                }
            )
    write_csv(path, rows, fieldnames)


if __name__ == "__main__":
    main()