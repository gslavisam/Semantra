"""Run an end-to-end SAP canonical coverage exercise.

This script orchestrates the SAP inventory pipeline and computes measurable coverage KPIs
for a full SAP workbook pass (often ~10k fields).

Modes:
- audit: generate inventory + review queue only (read-only on canonical glossary)
- promote: run safe promotion scripts, re-classify, and then measure

Outputs under knowledge_sources/generated/runtime/sap/:
- sap_full_coverage_exercise_summary.csv
- sap_snapshot_regression_results.csv
"""

from __future__ import annotations

import argparse
import csv
import os
import subprocess
import sys
from collections import Counter
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[2]
BACKEND_ROOT = PROJECT_ROOT / "backend"
if str(BACKEND_ROOT) not in sys.path:
    sys.path.insert(0, str(BACKEND_ROOT))

from app.models.schema import ColumnProfile, SchemaProfile
from app.services.mapping_service import generate_mapping_candidates
from app.services.virtual_target_service import build_virtual_target_schema

SAP_RUNTIME_DIR = PROJECT_ROOT / "knowledge_sources" / "generated" / "runtime" / "sap"
CLASSIFICATION_PATH = SAP_RUNTIME_DIR / "sap_full_inventory_classification.csv"
SUMMARY_OUTPUT_PATH = SAP_RUNTIME_DIR / "sap_full_coverage_exercise_summary.csv"
SNAPSHOT_RESULTS_PATH = SAP_RUNTIME_DIR / "sap_snapshot_regression_results.csv"
WORKBOOK_PATH = PROJECT_ROOT / "metadata_dict" / "sap_tables_mostUsed.xlsx"
SNAPSHOT_DIR = PROJECT_ROOT / "benchmarks" / "vendor_coverage" / "sap"

BUCKET_ORDER = (
    "direct_alias_match",
    "description_alias_match",
    "strong_canonical_candidate",
    "weak_canonical_candidate",
    "knowledge_only",
    "unmapped",
)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Run SAP full coverage exercise and KPI report.")
    parser.add_argument(
        "--mode",
        choices=("audit", "promote"),
        default="audit",
        help="audit=read-only inventory and KPIs; promote=apply safe promotions before KPI recompute",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=0,
        help="Optional row limit for fast diagnostic passes; 0 means full workbook.",
    )
    parser.add_argument(
        "--python",
        default=sys.executable,
        help="Python executable used for orchestrated child scripts.",
    )
    parser.add_argument(
        "--snapshot-regression",
        action="store_true",
        help="Run canonical snapshot regression (can be expensive with large canonical target sets).",
    )
    parser.add_argument(
        "--generate-unmapped-auto-overlay",
        action="store_true",
        help="Generate non-destructive auto-enrichment candidates/overlay for unmapped SAP rows.",
    )
    parser.add_argument(
        "--unmapped-auto-min-strength",
        type=float,
        default=0.75,
        help="Minimum top canonical strength for auto-unmapped candidate generation.",
    )
    parser.add_argument(
        "--unmapped-auto-min-margin",
        type=float,
        default=0.15,
        help="Minimum top-vs-second margin for auto-unmapped candidate generation.",
    )
    parser.add_argument(
        "--unmapped-auto-buckets",
        default="unmapped,knowledge_only",
        help="Buckets included by auto-unmapped generator.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    SAP_RUNTIME_DIR.mkdir(parents=True, exist_ok=True)

    run_pipeline(
        mode=args.mode,
        limit=args.limit,
        python_executable=args.python,
        generate_unmapped_auto_overlay=args.generate_unmapped_auto_overlay,
        unmapped_auto_min_strength=args.unmapped_auto_min_strength,
        unmapped_auto_min_margin=args.unmapped_auto_min_margin,
        unmapped_auto_buckets=args.unmapped_auto_buckets,
    )

    classification_rows = list(read_csv(CLASSIFICATION_PATH))
    coverage_metrics = compute_coverage_metrics(classification_rows)

    snapshot_metrics: list[dict[str, str]] = []
    if args.snapshot_regression:
        snapshot_metrics = run_snapshot_regression()
        write_snapshot_results(snapshot_metrics)

    write_exercise_summary(
        coverage_metrics,
        snapshot_metrics,
        args.mode,
        args.limit,
        snapshot_regression_enabled=args.snapshot_regression,
    )
    print_report(
        coverage_metrics,
        snapshot_metrics,
        args.mode,
        args.limit,
        snapshot_regression_enabled=args.snapshot_regression,
    )


def run_pipeline(
    mode: str,
    limit: int,
    python_executable: str,
    *,
    generate_unmapped_auto_overlay: bool,
    unmapped_auto_min_strength: float,
    unmapped_auto_min_margin: float,
    unmapped_auto_buckets: str,
) -> None:
    commands: list[list[str]] = [
        [
            python_executable,
            "support/sap/generate_sap_canonical_inventory.py",
            *(["--limit", str(limit)] if limit > 0 else []),
        ],
    ]

    if mode == "promote":
        commands.extend(
            [
                [python_executable, "support/sap/promote_sap_canonical_matches.py"],
                [python_executable, "support/sap/promote_sap_canonical_expansions.py"],
                [python_executable, "support/sap/materialize_sap_canonical_contexts.py"],
                [
                    python_executable,
                    "support/sap/generate_sap_canonical_inventory.py",
                    *(["--limit", str(limit)] if limit > 0 else []),
                ],
            ]
        )

    commands.append([python_executable, "support/sap/prioritize_sap_review_queue.py"])

    if generate_unmapped_auto_overlay:
        commands.append(
            [
                python_executable,
                "support/sap/generate_sap_unmapped_context_patch_overlay.py",
                "--min-strength",
                f"{unmapped_auto_min_strength:.2f}",
                "--min-margin",
                f"{unmapped_auto_min_margin:.2f}",
                "--include-buckets",
                unmapped_auto_buckets,
            ]
        )

    env = os.environ.copy()
    existing_pythonpath = env.get("PYTHONPATH", "")
    env["PYTHONPATH"] = str(BACKEND_ROOT) if not existing_pythonpath else str(BACKEND_ROOT) + os.pathsep + existing_pythonpath

    for command in commands:
        completed = subprocess.run(command, cwd=PROJECT_ROOT, env=env, check=False)
        if completed.returncode != 0:
            raise RuntimeError(f"Pipeline step failed ({completed.returncode}): {' '.join(command)}")


def compute_coverage_metrics(rows: list[dict[str, str]]) -> dict[str, float]:
    total = len(rows)
    bucket_counts = Counter(row.get("classification_bucket", "") for row in rows)

    direct_alias = bucket_counts.get("direct_alias_match", 0)
    description_alias = bucket_counts.get("description_alias_match", 0)
    strong_candidate = bucket_counts.get("strong_canonical_candidate", 0)
    weak_candidate = bucket_counts.get("weak_canonical_candidate", 0)
    knowledge_only = bucket_counts.get("knowledge_only", 0)
    unmapped = bucket_counts.get("unmapped", 0)

    mapped_strict = direct_alias + description_alias
    mapped_strong = mapped_strict + strong_candidate
    mapped_with_review = mapped_strong + weak_candidate
    coverage_any_path = mapped_with_review + knowledge_only

    metrics: dict[str, float] = {
        "total_fields": float(total),
        "mapped_strict": float(mapped_strict),
        "mapped_strong": float(mapped_strong),
        "mapped_with_review": float(mapped_with_review),
        "coverage_any_path": float(coverage_any_path),
        "knowledge_only": float(knowledge_only),
        "unmapped": float(unmapped),
    }

    for bucket in BUCKET_ORDER:
        metrics[f"bucket_{bucket}"] = float(bucket_counts.get(bucket, 0))

    for key in ("mapped_strict", "mapped_strong", "mapped_with_review", "coverage_any_path", "knowledge_only", "unmapped"):
        metrics[f"{key}_ratio"] = (metrics[key] / total) if total else 0.0

    return metrics


def run_snapshot_regression() -> list[dict[str, str]]:
    canonical_target_schema = build_virtual_target_schema("canonical")
    field_description_map = load_field_descriptions(WORKBOOK_PATH)

    results: list[dict[str, str]] = []
    for snapshot_path in sorted(SNAPSHOT_DIR.glob("*_mapping_snapshot.csv")):
        rows = list(read_csv(snapshot_path))
        expected_pairs: list[tuple[str, str]] = []
        for row in rows:
            source_field = (row.get("source_field") or "").strip()
            expected_target = (row.get("expected_target") or "").strip()
            if source_field and expected_target:
                expected_pairs.append((source_field, expected_target))

        if not expected_pairs:
            continue

        source_schema = SchemaProfile(
            dataset_id=snapshot_path.stem,
            dataset_name=snapshot_path.stem,
            row_count=max(len(expected_pairs), 1),
            columns=[build_profile(field, field_description_map.get(field, "")) for field, _ in expected_pairs],
        )

        response = generate_mapping_candidates(source_schema, canonical_target_schema, write_decision_log=False)
        by_source = {mapping.source: mapping for mapping in response.mappings}

        total = 0
        exact = 0
        accepted = 0
        accepted_exact = 0
        high_conf = 0

        for source_field, expected_target in expected_pairs:
            total += 1
            mapping = by_source.get(source_field)
            if mapping is None:
                continue
            predicted = mapping.target or ""
            is_exact = predicted == expected_target
            if is_exact:
                exact += 1
            if mapping.status == "accepted":
                accepted += 1
                if is_exact:
                    accepted_exact += 1
            if mapping.confidence_label == "high_confidence":
                high_conf += 1

        results.append(
            {
                "snapshot": snapshot_path.name,
                "fields": str(total),
                "exact_target_hits": str(exact),
                "exact_target_ratio": f"{(exact / total):.4f}" if total else "0.0000",
                "accepted_rows": str(accepted),
                "accepted_ratio": f"{(accepted / total):.4f}" if total else "0.0000",
                "accepted_exact_hits": str(accepted_exact),
                "accepted_exact_ratio": f"{(accepted_exact / total):.4f}" if total else "0.0000",
                "high_conf_rows": str(high_conf),
                "high_conf_ratio": f"{(high_conf / total):.4f}" if total else "0.0000",
            }
        )

    return results


def load_field_descriptions(workbook_path: Path) -> dict[str, str]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    worksheet = workbook["Tbls_Clm"]

    header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
    header_index = {
        str(value or "").strip(): index
        for index, value in enumerate(header_row)
        if str(value or "").strip()
    }
    if "Field" not in header_index or "Description" not in header_index:
        return {}

    field_counts: dict[str, Counter[str]] = {}
    for workbook_row in worksheet.iter_rows(min_row=2, values_only=True):
        field = str(workbook_row[header_index["Field"]] or "").strip()
        description = str(workbook_row[header_index["Description"]] or "").strip()
        if not field or not description:
            continue
        field_counter = field_counts.setdefault(field, Counter())
        field_counter[description] += 1

    chosen: dict[str, str] = {}
    for field, descriptions in field_counts.items():
        chosen[field] = descriptions.most_common(1)[0][0]
    return chosen


def build_profile(name: str, description: str) -> ColumnProfile:
    return ColumnProfile(
        name=name,
        normalized_name=name.lower(),
        description=description,
        declared_type="",
        dtype="string",
        null_ratio=0.0,
        unique_ratio=1.0,
        avg_length=float(len(name)),
        non_null_count=100,
        sample_values=[],
        distinct_sample_values=[],
        detected_patterns=[],
        tokenized_name=[],
    )


def write_snapshot_results(rows: list[dict[str, str]]) -> None:
    fieldnames = [
        "snapshot",
        "fields",
        "exact_target_hits",
        "exact_target_ratio",
        "accepted_rows",
        "accepted_ratio",
        "accepted_exact_hits",
        "accepted_exact_ratio",
        "high_conf_rows",
        "high_conf_ratio",
    ]
    write_csv(SNAPSHOT_RESULTS_PATH, rows, fieldnames)


def write_exercise_summary(
    coverage_metrics: dict[str, float],
    snapshot_metrics: list[dict[str, str]],
    mode: str,
    limit: int,
    snapshot_regression_enabled: bool,
) -> None:
    snapshot_count = len(snapshot_metrics)
    avg_exact_ratio = average(float(item["exact_target_ratio"]) for item in snapshot_metrics)
    avg_accepted_exact_ratio = average(float(item["accepted_exact_ratio"]) for item in snapshot_metrics)

    rows = [
        {"metric": "mode", "value": mode},
        {"metric": "row_limit", "value": str(limit or 0)},
        {"metric": "snapshot_regression_enabled", "value": str(snapshot_regression_enabled).lower()},
        {"metric": "total_fields", "value": str(int(coverage_metrics["total_fields"]))},
        {"metric": "mapped_strict", "value": str(int(coverage_metrics["mapped_strict"]))},
        {"metric": "mapped_strict_ratio", "value": f"{coverage_metrics['mapped_strict_ratio']:.4f}"},
        {"metric": "mapped_strong", "value": str(int(coverage_metrics["mapped_strong"]))},
        {"metric": "mapped_strong_ratio", "value": f"{coverage_metrics['mapped_strong_ratio']:.4f}"},
        {"metric": "mapped_with_review", "value": str(int(coverage_metrics["mapped_with_review"]))},
        {"metric": "mapped_with_review_ratio", "value": f"{coverage_metrics['mapped_with_review_ratio']:.4f}"},
        {"metric": "coverage_any_path", "value": str(int(coverage_metrics["coverage_any_path"]))},
        {"metric": "coverage_any_path_ratio", "value": f"{coverage_metrics['coverage_any_path_ratio']:.4f}"},
        {"metric": "knowledge_only", "value": str(int(coverage_metrics["knowledge_only"]))},
        {"metric": "knowledge_only_ratio", "value": f"{coverage_metrics['knowledge_only_ratio']:.4f}"},
        {"metric": "unmapped", "value": str(int(coverage_metrics["unmapped"]))},
        {"metric": "unmapped_ratio", "value": f"{coverage_metrics['unmapped_ratio']:.4f}"},
        {"metric": "snapshot_count", "value": str(snapshot_count)},
        {"metric": "snapshot_avg_exact_target_ratio", "value": f"{avg_exact_ratio:.4f}"},
        {"metric": "snapshot_avg_accepted_exact_ratio", "value": f"{avg_accepted_exact_ratio:.4f}"},
    ]

    for bucket in BUCKET_ORDER:
        rows.append({"metric": f"bucket_{bucket}", "value": str(int(coverage_metrics[f'bucket_{bucket}']))})

    write_csv(SUMMARY_OUTPUT_PATH, rows, ["metric", "value"])

    if not snapshot_regression_enabled:
        write_csv(
            SNAPSHOT_RESULTS_PATH,
            [{
                "snapshot": "skipped",
                "fields": "0",
                "exact_target_hits": "0",
                "exact_target_ratio": "0.0000",
                "accepted_rows": "0",
                "accepted_ratio": "0.0000",
                "accepted_exact_hits": "0",
                "accepted_exact_ratio": "0.0000",
                "high_conf_rows": "0",
                "high_conf_ratio": "0.0000",
            }],
            [
                "snapshot",
                "fields",
                "exact_target_hits",
                "exact_target_ratio",
                "accepted_rows",
                "accepted_ratio",
                "accepted_exact_hits",
                "accepted_exact_ratio",
                "high_conf_rows",
                "high_conf_ratio",
            ],
        )


def print_report(
    coverage_metrics: dict[str, float],
    snapshot_metrics: list[dict[str, str]],
    mode: str,
    limit: int,
    snapshot_regression_enabled: bool,
) -> None:
    print("=== SAP Full Coverage Exercise ===")
    print(f"mode={mode} limit={limit or 0}")
    print(f"total_fields={int(coverage_metrics['total_fields'])}")
    print(
        "mapped_strict="
        f"{int(coverage_metrics['mapped_strict'])} ({coverage_metrics['mapped_strict_ratio']:.2%})"
    )
    print(
        "mapped_strong="
        f"{int(coverage_metrics['mapped_strong'])} ({coverage_metrics['mapped_strong_ratio']:.2%})"
    )
    print(
        "mapped_with_review="
        f"{int(coverage_metrics['mapped_with_review'])} ({coverage_metrics['mapped_with_review_ratio']:.2%})"
    )
    print(
        "coverage_any_path="
        f"{int(coverage_metrics['coverage_any_path'])} ({coverage_metrics['coverage_any_path_ratio']:.2%})"
    )
    print(f"unmapped={int(coverage_metrics['unmapped'])} ({coverage_metrics['unmapped_ratio']:.2%})")

    if snapshot_regression_enabled and snapshot_metrics:
        avg_exact = average(float(item["exact_target_ratio"]) for item in snapshot_metrics)
        avg_accepted_exact = average(float(item["accepted_exact_ratio"]) for item in snapshot_metrics)
        print(f"snapshot_count={len(snapshot_metrics)}")
        print(f"snapshot_avg_exact_target_ratio={avg_exact:.2%}")
        print(f"snapshot_avg_accepted_exact_ratio={avg_accepted_exact:.2%}")
    elif not snapshot_regression_enabled:
        print("snapshot_regression=skipped")

    print(f"Wrote: {SUMMARY_OUTPUT_PATH}")
    print(f"Wrote: {SNAPSHOT_RESULTS_PATH}")


def read_csv(path: Path) -> Iterable[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        yield from csv.DictReader(handle)


def write_csv(path: Path, rows: list[dict[str, str]], fieldnames: list[str]) -> None:
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def average(values: Iterable[float]) -> float:
    values_list = list(values)
    return (sum(values_list) / len(values_list)) if values_list else 0.0


if __name__ == "__main__":
    main()
